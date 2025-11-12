import pandas as pd
from io import BytesIO
from datetime import datetime
import os
from logger import logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def export_to_excel(registros, colunas, nome_arquivo):
    """
    Exporta registros para um arquivo Excel com formatação
    """
    try:
        # Criar DataFrame
        df = pd.DataFrame(registros, columns=colunas)
        
        # Criar buffer em memória
        output = BytesIO()
        
        # Salvar primeiro com pandas
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados', index=False)
            
            # Ajustar largura das colunas
            worksheet = writer.sheets['Dados']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
        
        # Agora aplicar formatação condicional
        output.seek(0)
        workbook = load_workbook(output)
        worksheet = workbook['Dados']
        
        # Definir fill amarelo para células com "S/DePara"
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Encontrar todas as colunas que terminam com "_Codigo"
        colunas_codigo = []
        for idx, cell in enumerate(worksheet[1]):  # Linha 1 é o cabeçalho
            if cell.value and str(cell.value).endswith('_Codigo'):
                colunas_codigo.append(idx)
        
        # Aplicar formatação para todas as colunas de código
        for coluna_idx in colunas_codigo:
            for row in worksheet.iter_rows(min_row=2):  # Pular cabeçalho
                cell = row[coluna_idx]
                if cell.value == "S/DePara":
                    cell.fill = yellow_fill
        
        # Salvar de volta no buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"Arquivo Excel exportado com {len(registros)} registros e {len(colunas)} colunas")
        return output
    
    except Exception as e:
        logger.error(f"Erro ao exportar para Excel: {str(e)}")
        raise e

def import_from_excel(arquivo):
    """
    Importa dados de um arquivo Excel
    """
    try:
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Substituir NaN por None para melhor compatibilidade
        df = df.where(pd.notnull(df), None)
        
        # Converter para lista de dicionários
        registros = df.to_dict('records')
        colunas = df.columns.tolist()
        
        logger.info(f"Importados {len(registros)} registros com {len(colunas)} colunas: {colunas}")
        
        # Log do primeiro registro para debug
        if registros:
            logger.info(f"Primeiro registro: {registros[0]}")
        
        return registros, colunas
    
    except Exception as e:
        logger.error(f"Erro ao importar do Excel: {str(e)}")
        raise e