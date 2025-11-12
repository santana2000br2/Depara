from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import tempfile
import os
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

pessoacodfabricante_bp = Blueprint("pessoacodfabricante", __name__)

def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela Pessoa do banco homólogo onde Pessoa_TipoPessoa = 'J'"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        
        # Verificar se a tabela existe
        cursor.execute("""
            SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME = 'Pessoa'
        """)
        tabela_existe = cursor.fetchone()
        
        if not tabela_existe:
            logger.warning(f"Tabela Pessoa não existe no banco homólogo: {banco_homo}")
            return []
        
        # Query corrigida conforme solicitado - apenas códigos
        cursor.execute("SELECT Pessoa_Codigo FROM Pessoa WHERE Pessoa_TipoPessoa = 'J'")
        registros = cursor.fetchall()
        
        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF (Pessoa_TipoPessoa = 'J')")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []

# REMOVIDA: função obter_descricao_wf - não temos campo de descrição para atualizar
# REMOVIDA: função atualizar_descricoes_apos_importacao - não temos descrição para atualizar

@pessoacodfabricante_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('pessoacodfabricante.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM PessoaCodFabricante_DePara")
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('pessoacodfabricante.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em pessoacodfabricante: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('pessoacodfabricante.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@pessoacodfabricante_bp.route('/exportar')
def exportar_pessoacodfabricante():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        # Executar consulta - USANDO COLUNAS CORRETAS DA TABELA
        cursor = conexao.cursor()
        
        # Primeiro verificar a estrutura da tabela
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'PessoaCodFabricante_DePara'
        """)
        colunas_existentes = [row[0] for row in cursor.fetchall()]
        logger.info(f"Colunas existentes na tabela: {colunas_existentes}")
        
        # Construir query baseada nas colunas existentes - INCLUIR ID
        colunas_selecionar = []
        # CORREÇÃO: Incluir ID sempre que existir
        if 'id' in colunas_existentes:
            colunas_selecionar.append('id')
        if 'fabr_cd' in colunas_existentes:
            colunas_selecionar.append('fabr_cd')
        if 'fabr_nm' in colunas_existentes:
            colunas_selecionar.append('fabr_nm')
        if 'fabr_cdmont' in colunas_existentes:
            colunas_selecionar.append('fabr_cdmont')
        if 'ProdutoMarca_PessoaCodFabricante' in colunas_existentes:
            colunas_selecionar.append('ProdutoMarca_PessoaCodFabricante')
        if 'PessoaCodFabricante_Descricao' in colunas_existentes:
            colunas_selecionar.append('PessoaCodFabricante_Descricao')
        if 'PessoaCodFabricante_MarcaCod' in colunas_existentes:
            colunas_selecionar.append('PessoaCodFabricante_MarcaCod')
        if 'PessoaCodFabricante_Letra' in colunas_existentes:
            colunas_selecionar.append('PessoaCodFabricante_Letra')
        
        if not colunas_selecionar:
            flash('Nenhuma coluna válida encontrada na tabela.', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        query = f"SELECT {', '.join(colunas_selecionar)} FROM PessoaCodFabricante_DePara"
        logger.info(f"Executando query: {query}")
        
        cursor.execute(query)
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]
        
        # Mapear nomes das colunas para os nomes amigáveis - INCLUIR ID
        mapeamento_colunas = {
            'id': 'ID',
            'fabr_cd': 'Codigo de Origem',
            'fabr_nm': 'Descrição de origem',
            'fabr_cdmont': 'Codigo Montadora',
            'ProdutoMarca_PessoaCodFabricante': 'ProdutoMarca_PessoaCodFabricante',
            'PessoaCodFabricante_Descricao': 'PessoaCodFabricante_Descricao',
            'PessoaCodFabricante_MarcaCod': 'PessoaCodFabricante_MarcaCod',
            'PessoaCodFabricante_Letra': 'PessoaCodFabricante_Letra'
        }
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook e worksheet de forma mais robusta
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="PessoaCodFabricante_DePara")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna ProdutoMarca_PessoaCodFabricante
                if colunas_originais[col_num-1] == 'ProdutoMarca_PessoaCodFabricante':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "PessoaCodFabricante_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar pessoacodfabricante: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('pessoacodfabricante.index'))

@pessoacodfabricante_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_pessoacodfabricante_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet
        ws = wb.create_sheet(title="PessoaCodFabricante_Filtrado")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna ProdutoMarca_PessoaCodFabricante
                if header == 'ProdutoMarca_PessoaCodFabricante':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="PessoaCodFabricante_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar pessoacodfabricante filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@pessoacodfabricante_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela Pessoa do banco homólogo (BancoHomo) onde Pessoa_TipoPessoa = 'J'"""
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter o BancoHomo diretamente do banco de dados
        banco_homo = obter_banco_homo(projeto_id)
        
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")
        
        # Conectar ao banco homólogo
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        cursor = conexao.cursor()
        
        # Verificar se a tabela existe no banco homólogo
        cursor.execute("""
            SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME = 'Pessoa'
        """)
        tabela_existe = cursor.fetchone()
        
        if not tabela_existe:
            flash('Tabela Pessoa não encontrada no banco homólogo.', 'error')
            return redirect(url_for('pessoacodfabricante.index'))
        
        # Executar consulta na tabela Pessoa do banco homólogo - COM A CONDIÇÃO CORRETA
        cursor.execute("SELECT Pessoa_Codigo, Pessoa_nome FROM Pessoa WHERE Pessoa_TipoPessoa = 'J'")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        # Converter para lista de dicionários
        registros_list = [dict(zip(colunas, row)) for row in registros]
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Criar DataFrame pandas
        df = pd.DataFrame(registros_list, columns=colunas)
        
        # Criar buffer para o arquivo Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='PessoaCodFabricante_WF', index=False)
        
        output.seek(0)
        
        # Nome do arquivo
        nome_arquivo = "PessoaCodFabricante_WF.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('pessoacodfabricante.index'))


@pessoacodfabricante_bp.route('/importar', methods=['POST'])
def importar_pessoacodfabricante():
    try:
        # Verificar se foi enviado um arquivo
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['file']
        
        # Verificar se filename existe e não está vazio
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        # Verificação mais segura da extensão
        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Iniciando importação para o banco: {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Primeiro verificar a estrutura da tabela
        cursor.execute("""
            SELECT COLUMN_NAME, DATA_TYPE 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'PessoaCodFabricante_DePara'
        """)
        colunas_info = {row[0]: row[1] for row in cursor.fetchall()}
        colunas_existentes = list(colunas_info.keys())
        logger.info(f"Colunas existentes na tabela: {colunas_existentes}")
        logger.info(f"Tipos de dados das colunas: {colunas_info}")
        
        # Importar dados do Excel
        registros, colunas_excel = import_from_excel(arquivo)
        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")
        
        # MAPEAMENTO: Nomes das colunas na planilha para nomes das colunas no banco - INCLUIR ID
        mapeamento_colunas = {
            'ID': 'id',  # CORREÇÃO: Adicionar mapeamento para ID
            'Codigo de Origem': 'fabr_cd',
            'Descrição de origem': 'fabr_nm',
            'Codigo Montadora': 'fabr_cdmont',
            'ProdutoMarca_PessoaCodFabricante': 'ProdutoMarca_PessoaCodFabricante',
            'PessoaCodFabricante_Descricao': 'PessoaCodFabricante_Descricao',
            'PessoaCodFabricante_MarcaCod': 'PessoaCodFabricante_MarcaCod',
            'PessoaCodFabricante_Letra': 'PessoaCodFabricante_Letra',
            # Incluir também os nomes originais caso a planilha os use
            'id': 'id',
            'fabr_cd': 'fabr_cd',
            'fabr_nm': 'fabr_nm',
            'fabr_cdmont': 'fabr_cdmont'
        }
        
        # Colunas do banco que precisamos (apenas as que existem na tabela) - INCLUIR ID
        colunas_banco = []
        # CORREÇÃO: Incluir ID sempre que existir
        if 'id' in colunas_existentes:
            colunas_banco.append('id')
        for coluna in ['fabr_cd', 'fabr_nm', 'fabr_cdmont', 'ProdutoMarca_PessoaCodFabricante', 
                      'PessoaCodFabricante_Descricao', 'PessoaCodFabricante_MarcaCod', 'PessoaCodFabricante_Letra']:
            if coluna in colunas_existentes:
                colunas_banco.append(coluna)
        
        if not colunas_banco:
            return jsonify({'success': False, 'message': 'Nenhuma coluna válida encontrada na tabela de destino'})
        
        # Verificar se todas as colunas necessárias estão presentes (usando o mapeamento)
        colunas_faltantes = []
        for coluna_banco in colunas_banco:
            # Encontrar o nome correspondente na planilha
            coluna_planilha = None
            for chave, valor in mapeamento_colunas.items():
                if valor == coluna_banco and chave in colunas_excel:
                    coluna_planilha = chave
                    break
            
            if not coluna_planilha:
                colunas_faltantes.append(coluna_banco)
        
        if colunas_faltantes:
            return jsonify({
                'success': False, 
                'message': f'Colunas necessárias faltando no arquivo: {", ".join(colunas_faltantes)}. Certifique-se de que a planilha contém as colunas: {", ".join(colunas_banco)}'
            })
        
        # Função para converter valores conforme o tipo de dados do banco
        def converter_valor_para_tipo(coluna, valor):
            if valor is None or valor == '':
                return None
            
            tipo_dado = colunas_info.get(coluna)
            
            # Remover espaços extras se for string
            if isinstance(valor, str):
                valor = valor.strip()
                if valor == '':
                    return None
            
            # Converter baseado no tipo de dados do banco
            if tipo_dado in ['int', 'bigint', 'smallint', 'tinyint']:
                try:
                    return int(float(valor)) if valor else None
                except (ValueError, TypeError):
                    return None
            elif tipo_dado in ['float', 'real', 'decimal', 'numeric']:
                try:
                    return float(valor) if valor else None
                except (ValueError, TypeError):
                    return None
            elif tipo_dado in ['varchar', 'nvarchar', 'text', 'char', 'nchar']:
                return str(valor) if valor else None
            else:
                # Para tipos desconhecidos, manter como está
                return valor
        
        # Filtrar e mapear os registros
        registros_mapeados = []
        for registro in registros:
            registro_mapeado = {}
            
            # Para cada coluna do banco, buscar o valor correspondente na planilha
            for coluna_banco in colunas_banco:
                # Encontrar o nome da coluna na planilha
                coluna_planilha = None
                for chave, valor in mapeamento_colunas.items():
                    if valor == coluna_banco and chave in colunas_excel:
                        coluna_planilha = chave
                        break
                
                if coluna_planilha:
                    valor = registro.get(coluna_planilha)
                    # Converter valor para o tipo apropriado
                    valor_convertido = converter_valor_para_tipo(coluna_banco, valor)
                    registro_mapeado[coluna_banco] = valor_convertido
                else:
                    registro_mapeado[coluna_banco] = None
            
            registros_mapeados.append(registro_mapeado)
        
        logger.info(f"Registros mapeados: {len(registros_mapeados)}")
        
        # VERIFICAÇÃO ANTES: Contar registros antes da importação
        cursor.execute("SELECT COUNT(*) FROM PessoaCodFabricante_DePara")
        result_antes = cursor.fetchone()
        count_antes = result_antes[0] if result_antes and result_antes[0] is not None else 0
        logger.info(f"Registros na tabela ANTES da importação: {count_antes}")
        
        # ESTRATÉGIA CORRIGIDA: Usar ID como chave primária
        try:
            contador_atualizacoes = 0
            contador_insercoes = 0
            
            for registro in registros_mapeados:
                id_registro = registro.get('id')
                
                # CORREÇÃO: Se temos ID, tentar UPDATE, senão INSERT
                if id_registro:
                    # Verificar se o registro existe pelo ID
                    cursor.execute("SELECT COUNT(*) FROM PessoaCodFabricante_DePara WHERE id = ?", (id_registro,))
                    result = cursor.fetchone()
                    existe = result[0] > 0 if result and result[0] is not None else False
                    
                    logger.info(f"Verificando registro id={id_registro}, existe={existe}")
                    
                    if existe:
                        # UPDATE do registro existente pelo ID
                        update_values = []
                        update_fields = []
                        
                        # Atualizar todos os campos exceto o ID
                        for coluna in colunas_banco:
                            if coluna != 'id' and registro.get(coluna) is not None:
                                update_fields.append(f"{coluna} = ?")
                                update_values.append(registro.get(coluna))
                        
                        if update_fields:
                            update_values.append(id_registro)
                            query = f"UPDATE PessoaCodFabricante_DePara SET {', '.join(update_fields)} WHERE id = ?"
                            logger.info(f"Executando UPDATE pelo ID: {query} com valores: {update_values}")
                            cursor.execute(query, update_values)
                            contador_atualizacoes += 1
                            logger.info(f"UPDATE realizado para id={id_registro}")
                        else:
                            logger.warning(f"Nenhum campo para atualizar no registro id={id_registro}")
                    else:
                        # INSERT de novo registro (ID não existe)
                        insert_fields = []
                        insert_placeholders = []
                        insert_values = []
                        
                        for coluna in colunas_banco:
                            if registro.get(coluna) is not None:  # Só incluir campos não nulos
                                insert_fields.append(coluna)
                                insert_placeholders.append("?")
                                insert_values.append(registro.get(coluna))
                        
                        if insert_fields:
                            query = f"INSERT INTO PessoaCodFabricante_DePara ({', '.join(insert_fields)}) VALUES ({', '.join(insert_placeholders)})"
                            logger.info(f"Executando INSERT: {query} com valores: {insert_values}")
                            cursor.execute(query, insert_values)
                            contador_insercoes += 1
                            logger.info(f"INSERT realizado para id={id_registro}")
                else:
                    # INSERT de novo registro (sem ID)
                    insert_fields = []
                    insert_placeholders = []
                    insert_values = []
                    
                    for coluna in colunas_banco:
                        if registro.get(coluna) is not None:  # Só incluir campos não nulos
                            insert_fields.append(coluna)
                            insert_placeholders.append("?")
                            insert_values.append(registro.get(coluna))
                    
                    if insert_fields:
                        query = f"INSERT INTO PessoaCodFabricante_DePara ({', '.join(insert_fields)}) VALUES ({', '.join(insert_placeholders)})"
                        logger.info(f"Executando INSERT (sem ID): {query} com valores: {insert_values}")
                        cursor.execute(query, insert_values)
                        contador_insercoes += 1
                        logger.info(f"INSERT realizado (sem ID)")
            
            logger.info(f"UPDATEs executados: {contador_atualizacoes} registros")
            logger.info(f"INSERTs executados: {contador_insercoes} registros")
            
            # COMMIT
            logger.info("Executando COMMIT...")
            conexao.commit()
            logger.info("COMMIT executado com sucesso")
            
        except Exception as e:
            logger.error(f"Erro durante operações de banco: {str(e)}")
            conexao.rollback()
            logger.info("ROLLBACK executado")
            raise e
        
        # VERIFICAÇÃO DEPOIS: Contar registros após a importação
        cursor.execute("SELECT COUNT(*) FROM PessoaCodFabricante_DePara")
        result_depois = cursor.fetchone()
        count_depois = result_depois[0] if result_depois and result_depois[0] is not None else 0
        logger.info(f"Registros na tabela DEPOIS da importação: {count_depois}")
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True, 
            'message': f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total na base: {count_depois} registros.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao importar pessoacodfabricante: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False, 
            'message': f'Erro na importação: {str(e)}'
        })

@pessoacodfabricante_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline"""
    logger.info("=== UPDATE REGISTRO PESSOACODFABRICANTE ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados recebidos: {data}")
        
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')
        
        if not record_id or not field:
            logger.error("ID ou campo não fornecidos")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})
        
        # VERIFICAR SE O CAMPO É PERMITIDO - CORREÇÃO: Campo principal é ProdutoMarca_PessoaCodFabricante
        colunas_permitidas = ['ProdutoMarca_PessoaCodFabricante', 'PessoaCodFabricante_Descricao', 'PessoaCodFabricante_MarcaCod', 'PessoaCodFabricante_Letra']
        if field not in colunas_permitidas:
            logger.error(f"Tentativa de editar campo não permitido: {field}")
            return jsonify({'success': False, 'message': f'Campo {field} não é permitido para edição'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            logger.error("Nenhum projeto selecionado na sessão")
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        
        if not banco_usuario:
            logger.error("Banco não configurado para este projeto")
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Atualizando registro {record_id}, campo {field} para valor '{value}' no banco {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error(f"Falha na conexão com o banco: {banco_usuario}")
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Primeiro verificar se a coluna existe na tabela
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'PessoaCodFabricante_DePara' AND COLUMN_NAME = ?
        """, (field,))
        coluna_existe = cursor.fetchone()
        
        if not coluna_existe:
            logger.error(f"Coluna {field} não existe na tabela")
            return jsonify({'success': False, 'message': f'Coluna {field} não existe na tabela'})
        
        # Verificar se o registro existe
        cursor.execute(f"SELECT TOP 1 {field} FROM PessoaCodFabricante_DePara WHERE id = ?", (record_id,))
        resultado = cursor.fetchone()
        if not resultado:
            logger.error(f"Registro não encontrado: {record_id}")
            return jsonify({'success': False, 'message': 'Registro não encontrado'})
        
        # Atualizar registro - usando id como chave primária
        query = f"UPDATE PessoaCodFabricante_DePara SET {field} = ? WHERE id = ?"
        logger.info(f"Executando query: {query} com valores: ({value}, {record_id})")
        
        cursor.execute(query, (value, record_id))
        
        # Verificar se alguma linha foi afetada
        if cursor.rowcount == 0:
            logger.warning(f"Nenhuma linha afetada pela atualização do registro {record_id}")
            if conexao:
                conexao.rollback()
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})
        
        # Commit da transação
        conexao.commit()
        
        logger.info(f"Registro {record_id} atualizado com sucesso")
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
        
    except Exception as e:
        logger.error(f"Erro ao atualizar registro: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Fazer rollback em caso de erro apenas se a conexão existir
        if conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")
        
        return jsonify({'success': False, 'message': f'Erro ao atualizar registro: {str(e)}'})
    
    finally:
        # Fechar recursos de forma segura
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            logger.error(f"Erro ao fechar cursor: {e}")
        
        try:
            if conexao:
                conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexão: {e}")

@pessoacodfabricante_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez - SEM atualização de descrição"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()
        
        # Verificar colunas existentes na tabela
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'PessoaCodFabricante_DePara'
        """)
        colunas_existentes = [row[0] for row in cursor.fetchall()]
        
        # CORREÇÃO: Campo principal é ProdutoMarca_PessoaCodFabricante
        colunas_permitidas = []
        for coluna in ['ProdutoMarca_PessoaCodFabricante', 'PessoaCodFabricante_Descricao', 'PessoaCodFabricante_MarcaCod', 'PessoaCodFabricante_Letra']:
            if coluna in colunas_existentes:
                colunas_permitidas.append(coluna)

        success_count = 0
        error_count = 0
        error_messages = []

        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')

                if not record_id or not field:
                    error_count += 1
                    continue

                if field not in colunas_permitidas:
                    error_count += 1
                    continue

                # Atualiza o campo alterado
                query = f"UPDATE PessoaCodFabricante_DePara SET {field} = ? WHERE id = ?"
                cursor.execute(query, (value, record_id))
                
                # REMOVIDO: atualização automática da descrição - não temos esse campo
                
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")

            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")

        conexao.commit()
        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")

        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }

        if error_messages:
            response['error_details'] = error_messages[:10]

        return jsonify(response)

    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}")
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})

    finally:
        if cursor:
            cursor.close()
        if conexao:
            conexao.close()

# REMOVIDO: endpoint get_descricao_wf - não temos descrição para buscar

# Função para dashboard (se necessário)
def dados_pessoacodfabricante(banco_usuario):
    if not banco_usuario:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    conexao = conectar_segunda_base(banco_usuario)
    if not conexao:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    try:
        cursor = conexao.cursor()
        
        # Primeira consulta - total de registros
        cursor.execute("SELECT COUNT(*) FROM PessoaCodFabricante_DePara")
        result = cursor.fetchone()
        qtd = result[0] if result and result[0] is not None else 0

        # CORREÇÃO: Campo principal é ProdutoMarca_PessoaCodFabricante
        # Segunda consulta - registros pendentes
        cursor.execute(
            "SELECT COUNT(*) FROM PessoaCodFabricante_DePara WHERE ProdutoMarca_PessoaCodFabricante = 'S/DePara'"
        )
        result_pendente = cursor.fetchone()
        qtdPendente = result_pendente[0] if result_pendente and result_pendente[0] is not None else 0

        # Cálculo do percentual
        percentualConclusao = ((qtd - qtdPendente) / qtd * 100) if qtd > 0 else 0

        return {
            "qtd": qtd,
            "qtdPendente": qtdPendente,
            "percentualConclusao": round(percentualConclusao, 1),
        }
    except Exception as e:
        logger.error(f"Erro ao calcular dados PessoaCodFabricante_DePara: {e}")
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}
    finally:
        if conexao:
            conexao.close()