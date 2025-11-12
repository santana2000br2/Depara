# equipe.py
from flask import (
    Blueprint, render_template, redirect, url_for, session, flash,
    request, jsonify, send_file
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

equipe_bp = Blueprint('equipe', __name__)

# -------------------------
# Helpers / utilitários
# -------------------------
def obter_banco_homo(projeto_id):
    """
    Retorna o BancoHomo configurado para o projeto (mesma lógica usada nos outros módulos).
    """
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("obter_banco_homo: falha ao conectar ao banco principal")
            return None
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if row and row[0]:
            return row[0]
        return None
    except Exception as e:
        logger.error(f"obter_banco_homo (equipe) -> {e}", exc_info=True)
        return None


def obter_codigos_wf(banco_homo):
    """
    Retorna lista de códigos existentes na tabela 'equipe' do banco homólogo.
    """
    try:
        if not banco_homo:
            return []
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"obter_codigos_wf: falha na conexão com banco_homo={banco_homo}")
            return []
        cursor = conexao.cursor()
        cursor.execute("SELECT Equipe_Codigo FROM equipe")
        rows = cursor.fetchall()
        cursor.close()
        conexao.close()
        codigos = [str(r[0]) for r in rows if r and r[0] is not None]
        logger.info(f"obter_codigos_wf: encontrados {len(codigos)} códigos em equipe (WF)")
        return codigos
    except Exception as e:
        logger.error(f"obter_codigos_wf (equipe) -> {e}", exc_info=True)
        return []


def obter_descricao_wf(banco_homo, codigo):
    """
    Retorna a descrição do código na tabela 'equipe' do banco homólogo.
    """
    try:
        if not banco_homo or not codigo:
            return None
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"obter_descricao_wf: falha na conexão banco_homo={banco_homo}")
            return None
        cursor = conexao.cursor()
        cursor.execute("SELECT Equipe_Descricao FROM equipe WHERE Equipe_Codigo = ?", codigo)
        row = cursor.fetchone()
        cursor.close()
        conexao.close()
        return row[0] if row and row[0] is not None else None
    except Exception as e:
        logger.error(f"obter_descricao_wf (equipe) -> {e}", exc_info=True)
        return None


def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """
    Percorre a tabela Equipe_DePara e atualiza Equipe_Descricao a partir do banco homólogo.
    Chamado após importações/updates em lote.
    """
    conexao = None
    cursor = None
    conexao_aux = None
    cursor_aux = None
    try:
        if not banco_homo:
            logger.info("atualizar_descricoes_apos_importacao: banco_homo não configurado, pulando.")
            return

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("atualizar_descricoes_apos_importacao: falha conectar ao banco do usuário")
            return

        cursor = conexao.cursor()
        # Seleciona registros que tem código WF configurado (não S/DePara)
        cursor.execute("""
            SELECT id, Equipe_Codigo, Equipe_Descricao
            FROM Equipe_DePara
            WHERE Equipe_Codigo IS NOT NULL AND Equipe_Codigo != 'S/DePara'
        """)
        rows = cursor.fetchall()
        logger.info(f"atualizar_descricoes_apos_importacao: {len(rows)} registros encontrados para verificação")

        if not rows:
            cursor.close()
            conexao.close()
            return

        # conexao para o banco homólogo
        conexao_aux = conectar_segunda_base(banco_homo)
        if not conexao_aux:
            logger.error("atualizar_descricoes_apos_importacao: falha conectar banco_homo")
            cursor.close()
            conexao.close()
            return
        cursor_aux = conexao_aux.cursor()

        updates = 0
        for r in rows:
            id_reg = r[0]
            codigo_wf = r[1]
            desc_atual = r[2] if len(r) > 2 else None
            if not codigo_wf:
                continue
            cursor_aux.execute("SELECT Equipe_Descricao FROM equipe WHERE Equipe_Codigo = ?", (codigo_wf,))
            res = cursor_aux.fetchone()
            desc_wf = res[0] if res and res[0] is not None else None
            if desc_wf and desc_wf != desc_atual:
                try:
                    cursor.execute("UPDATE Equipe_DePara SET Equipe_Descricao = ? WHERE id = ?", (desc_wf, id_reg))
                    updates += 1
                except Exception as ex:
                    logger.warning(f"Falha ao atualizar descricao id={id_reg}: {ex}")
        if updates > 0:
            conexao.commit()
            logger.info(f"atualizar_descricoes_apos_importacao: atualizadas {updates} descrições")
        else:
            logger.info("atualizar_descricoes_apos_importacao: nenhuma atualização necessária")

        # fechar conexões
        cursor_aux.close()
        conexao_aux.close()
        cursor.close()
        conexao.close()
    except Exception as e:
        logger.error(f"atualizar_descricoes_apos_importacao (equipe) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass


# -------------------------
# ROTAS
# -------------------------
@equipe_bp.route("/")
def index():
    """
    Página principal: lista registros de Equipe_DePara e passa codigos WF para o template.
    """
    conexao = None
    cursor = None
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        projeto_nome = projeto.get('NomeProjeto', 'N/A')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('equipe.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Equipe_DePara")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros_dict = [dict(zip(colunas, r)) for r in registros]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        cursor.close()
        conexao.close()

        return render_template('equipe.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf)
    except Exception as e:
        logger.error(f"index (equipe) -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro ao carregar equipe: {e}', 'error')
        return render_template('equipe.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])


@equipe_bp.route("/exportar")
def exportar_equipe():
    """
    Exporta toda a tabela Equipe_DePara para Excel com coloração na coluna Equipe_Codigo.
    """
    conexao = None
    cursor = None
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('equipe.index'))

        cursor = conexao.cursor()
        cursor.execute("SELECT eqp_cd, eqp_ds, Equipe_Codigo, Equipe_Descricao FROM Equipe_DePara")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="Equipe_DePara")

        # cabeçalhos amigáveis (map)
        mapeamento_colunas = {
            'eqp_cd': 'Codigo de Origem',
            'eqp_ds': 'Descrição de origem',
            'Equipe_Codigo': 'Equipe_Codigo',
            'Equipe_Descricao': 'Equipe_Descricao'
        }
        colunas_amigaveis = [mapeamento_colunas.get(c, c) for c in colunas]

        # cabeçalho
        for idx, c in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=idx, value=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for rnum, row in enumerate(registros, 2):
            for cnum, val in enumerate(row, 1):
                v = val.strip() if isinstance(val, str) else val
                cell = ws.cell(row=rnum, column=cnum, value=v)
                # Equipe_Codigo está na posição 3 (0-based -> index 2)
                if cnum == 3:
                    if not v or v == '':
                        cell.fill = laranja
                    elif v == 'S/DePara':
                        cell.fill = amarelo
                    elif str(v) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        # ajustar largura
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        cursor.close()
        conexao.close()

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Equipe_DePara.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"exportar_equipe (equipe) -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro na exportação: {e}', 'error')
        return redirect(url_for('equipe.index'))


@equipe_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_equipe_filtrados():
    """
    Exporta somente os registros filtrados enviados pelo front-end (com coloração).
    """
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400

        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="Equipe_Filtrado")

        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for rnum, registro in enumerate(registros_filtrados, 2):
            for cnum, header in enumerate(headers, 1):
                val = registro.get(header, '')
                if isinstance(val, str):
                    val = val.strip()
                cell = ws.cell(row=rnum, column=cnum, value=val)
                if header == 'Equipe_Codigo':
                    if not val or val == '':
                        cell.fill = laranja
                    elif val == 'S/DePara':
                        cell.fill = amarelo
                    elif val and str(val) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Equipe_Filtrado.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"exportar_equipe_filtrados (equipe) -> {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro: {e}'}), 500


@equipe_bp.route('/export_wf')
def export_wf():
    """
    Exporta a tabela 'equipe' do banco homólogo (WF) com os campos solicitados.
    SELECT: Equipe_Codigo, Equipe_Descricao, Equipe_Ativo
    """
    conexao = None
    cursor = None
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('equipe.index'))

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('equipe.index'))

        cursor = conexao.cursor()
        cursor.execute("SELECT Equipe_Codigo, Equipe_Descricao, Equipe_Ativo FROM equipe")
        rows = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros = [dict(zip(colunas, r)) for r in rows]
        cursor.close()
        conexao.close()

        df = pd.DataFrame(registros, columns=colunas)
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Equipe_WF', index=False)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Equipe_WF.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"export_wf (equipe) -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro na exportação WF: {e}', 'error')
        return redirect(url_for('equipe.index'))


@equipe_bp.route('/importar', methods=['POST'])
def importar_equipe():
    """
    Importa planilha Excel para Equipe_DePara.
    Aceita colunas com nomes amigáveis (Codigo de Origem, Descrição de origem) ou técnicos.
    Faz UPDATE quando eqp_cd já existe, senão INSERT.
    Depois tenta atualizar descrições a partir do WF (se banco homólogo configurado).
    """
    conexao = None
    cursor = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        filename = arquivo.filename
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use .xlsx ou .xls'})

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para o projeto'})

        logger.info(f"importar_equipe: iniciando importação para banco {banco_usuario}")

        # Ler Excel
        try:
            df = pd.read_excel(arquivo)
            df = df.where(pd.notnull(df), None)
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()
        except Exception as e:
            logger.error(f"importar_equipe: erro ao ler excel -> {e}", exc_info=True)
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {e}'})

        logger.info(f"importar_equipe: arquivo lido: {len(registros)} registros, colunas: {colunas_excel}")

        # Mapeamento nomes amigáveis -> técnicos
        mapeamento_colunas = {
            'Codigo de Origem': 'eqp_cd',
            'Descrição de origem': 'eqp_ds',
            'eqp_cd': 'eqp_cd',
            'eqp_ds': 'eqp_ds',
            'Equipe_Codigo': 'Equipe_Codigo',
            'Equipe_Descricao': 'Equipe_Descricao'
        }

        colunas_normalizadas = [mapeamento_colunas.get(c, c) for c in colunas_excel]
        df.columns = colunas_normalizadas

        # Colunas obrigatórias
        colunas_necessarias = {'eqp_cd', 'eqp_ds', 'Equipe_Codigo', 'Equipe_Descricao'}
        if not colunas_necessarias.issubset(set(colunas_normalizadas)):
            missing = colunas_necessarias - set(colunas_normalizadas)
            return jsonify({'success': False, 'message': f'Colunas necessárias faltando no arquivo: {", ".join(missing)}'})

        registros = df.to_dict('records')

        # Tratar tamanhos e normalizar
        registros_filtrados = []
        for reg in registros:
            r = {}
            r['eqp_cd'] = None if reg.get('eqp_cd') is None else str(reg.get('eqp_cd'))[:100]
            r['eqp_ds'] = None if reg.get('eqp_ds') is None else str(reg.get('eqp_ds'))[:200]
            r['Equipe_Codigo'] = None if reg.get('Equipe_Codigo') is None else str(reg.get('Equipe_Codigo'))[:100]
            r['Equipe_Descricao'] = None if reg.get('Equipe_Descricao') is None else str(reg.get('Equipe_Descricao'))[:200]
            registros_filtrados.append(r)

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # Estatísticas
        cursor.execute("SELECT COUNT(*) FROM Equipe_DePara")
        antes = cursor.fetchone()
        count_antes = antes[0] if antes else 0

        contador_update = 0
        contador_insert = 0

        for reg in registros_filtrados:
            chave = reg.get('eqp_cd')
            if chave:
                cursor.execute("SELECT id FROM Equipe_DePara WHERE eqp_cd = ?", (chave,))
                existe = cursor.fetchone()
                if existe:
                    cursor.execute("""
                        UPDATE Equipe_DePara
                        SET eqp_ds = ?, Equipe_Codigo = ?, Equipe_Descricao = ?
                        WHERE eqp_cd = ?
                    """, (reg.get('eqp_ds'), reg.get('Equipe_Codigo'), reg.get('Equipe_Descricao'), chave))
                    contador_update += 1
                else:
                    cursor.execute("""
                        INSERT INTO Equipe_DePara (eqp_cd, eqp_ds, Equipe_Codigo, Equipe_Descricao)
                        VALUES (?, ?, ?, ?)
                    """, (reg.get('eqp_cd'), reg.get('eqp_ds'), reg.get('Equipe_Codigo'), reg.get('Equipe_Descricao')))
                    contador_insert += 1
            else:
                # Sem chave de origem: insere mesmo assim
                cursor.execute("""
                    INSERT INTO Equipe_DePara (eqp_cd, eqp_ds, Equipe_Codigo, Equipe_Descricao)
                    VALUES (?, ?, ?, ?)
                """, (reg.get('eqp_cd'), reg.get('eqp_ds'), reg.get('Equipe_Codigo'), reg.get('Equipe_Descricao')))
                contador_insert += 1

        conexao.commit()
        logger.info(f"importar_equipe: atualizados={contador_update}, inseridos={contador_insert}")

        # Atualizar descricoes a partir do banco homólogo, se configurado
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            logger.info("importar_equipe: atualizando descricoes a partir do banco homólogo")
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        cursor.execute("SELECT COUNT(*) FROM Equipe_DePara")
        depois = cursor.fetchone()
        count_depois = depois[0] if depois else 0

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': f'Importação concluída! {contador_update} atualizados, {contador_insert} inseridos. Antes: {count_antes}, Depois: {count_depois}.'})
    except Exception as e:
        logger.error(f"importar_equipe (equipe) -> {e}", exc_info=True)
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if 'cursor' in locals() and cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro na importação: {e}'}), 500


@equipe_bp.route('/update', methods=['POST'])
def update_registro():
    """
    Atualiza um único registro (edição inline). Suporta Equipe_Codigo e Equipe_Descricao.
    Se Equipe_Codigo for alterado, tenta atualizar Equipe_Descricao automaticamente a partir do WF.
    """
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not record_id or not field:
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'}), 400

        campos_permitidos = ['Equipe_Codigo', 'Equipe_Descricao']
        if field not in campos_permitidos:
            return jsonify({'success': False, 'message': 'Campo não permitido para edição'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400

        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'}), 500

        cursor = conexao.cursor()

        # validação simples para Equipe_Codigo: números ou 'S/DePara'
        if field == 'Equipe_Codigo':
            if value is not None and value != 'S/DePara' and not str(value).isdigit():
                return jsonify({'success': False, 'message': 'Equipe_Codigo deve ser numérico ou "S/DePara"'}), 400

        cursor.execute(f"UPDATE Equipe_DePara SET {field} = ? WHERE id = ?", (value, record_id))
        conexao.commit()
        logger.info(f"update_registro: id={record_id} campo={field} valor={value}")

        # se atualizou o código, tentar atualizar descrição a partir do WF
        if field == 'Equipe_Codigo' and value and value != 'S/DePara':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                desc = obter_descricao_wf(banco_homo, value)
                if desc:
                    try:
                        cursor.execute("UPDATE Equipe_DePara SET Equipe_Descricao = ? WHERE id = ?", (desc, record_id))
                        conexao.commit()
                        logger.info(f"update_registro: descrição atualizada automaticamente para id={record_id}")
                    except Exception as ex:
                        logger.warning(f"update_registro: falha ao atualizar descricao automaticamente: {ex}")

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
    except Exception as e:
        logger.error(f"update_registro (equipe) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {e}'}), 500


@equipe_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """
    Atualização em lote. Recebe JSON com 'updates': [{id, field, value}, ...]
    """
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização enviada'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400

        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão: {banco_usuario}'}), 500

        cursor = conexao.cursor()

        contador = 0
        for upd in updates:
            record_id = upd.get('id')
            field = upd.get('field')
            value = upd.get('value')
            if not record_id or not field:
                continue
            if field not in ['Equipe_Codigo', 'Equipe_Descricao']:
                continue
            # validação simples
            if field == 'Equipe_Codigo' and value is not None and value != 'S/DePara' and not str(value).isdigit():
                continue
            cursor.execute(f"UPDATE Equipe_DePara SET {field} = ? WHERE id = ?", (value, record_id))
            contador += 1

        conexao.commit()
        # após aplicar em lote, tentar atualizar descrições a partir do banco homólogo
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': f'{contador} registros atualizados com sucesso'})
    except Exception as e:
        logger.error(f"update_batch (equipe) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar em lote: {e}'}), 500


@equipe_bp.route('/get_descricao_wf/<codigo>', methods=['GET'])
def get_descricao_wf_endpoint(codigo):
    """
    Endpoint AJAX para retornar a descrição do WF dado um código.
    """
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'descricao': None})
        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        desc = obter_descricao_wf(banco_homo, codigo)
        return jsonify({'descricao': desc})
    except Exception as e:
        logger.error(f"get_descricao_wf_endpoint (equipe) -> {e}", exc_info=True)
        return jsonify({'descricao': None})
