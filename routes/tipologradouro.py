from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

tipologradouro_bp = Blueprint("tipologradouro", __name__)

def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela TipoLogradouro do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoLogradouro_Codigo FROM TipoLogradouro")
        registros = cursor.fetchall()
        
        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela TipoLogradouro do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoLogradouro_Descricao FROM TipoLogradouro WHERE TipoLogradouro_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None

def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições após importação baseado nos códigos WF"""
    try:
        if not banco_homo:
            return
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições")
            return
        
        cursor = conexao.cursor()
        
        # Buscar registros que têm código WF (não são S/DePara)
        cursor.execute("""
            SELECT id, TipoLogradouro_Codigo, TipoLogradouro_Descricao 
            FROM TipoLogradouro_DePara 
            WHERE TipoLogradouro_Codigo IS NOT NULL AND TipoLogradouro_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()
        
        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)
            
            if descricao_wf and descricao_wf != descricao_atual:
                # Atualizar descrição
                cursor.execute("""
                    UPDATE TipoLogradouro_DePara 
                    SET TipoLogradouro_Descricao = ? 
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")
        
        conexao.commit()
        cursor.close()
        conexao.close()
        
        logger.info(f"Atualizações automáticas de descrição: {atualizacoes} registros")
        
    except Exception as e:
        logger.error(f"Erro ao atualizar descrições após importação: {str(e)}")

@tipologradouro_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('tipologradouro.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM TipoLogradouro_DePara")
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('tipologradouro.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em tipo logradouro: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('tipologradouro.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@tipologradouro_bp.route('/exportar')
def exportar_tipologradouro():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('tipologradouro.index'))
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('tipologradouro.index'))
        
        # Executar consulta - Excluir coluna id da exportação
        cursor = conexao.cursor()
        cursor.execute("""
            SELECT logradouro_sigla, logradouro_nm, TipoLogradouro_Codigo, 
                   TipoLogradouro_Sigla, TipoLogradouro_Descricao, tabela 
            FROM TipoLogradouro_DePara
        """)
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]
        
        # Mapear nomes das colunas para os nomes amigáveis
        mapeamento_colunas = {
            'logradouro_sigla': 'Logradouro Sigla',
            'logradouro_nm': 'Logradouro Nome',
            'TipoLogradouro_Codigo': 'TipoLogradouro_Codigo',
            'TipoLogradouro_Sigla': 'TipoLogradouro_Sigla',
            'TipoLogradouro_Descricao': 'TipoLogradouro_Descricao',
            'tabela': 'Tabela'
        }
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # CORREÇÃO: Criar workbook e worksheet de forma mais robusta
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="TipoLogradouro_DePara")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna TipoLogradouro_Codigo (coluna 3)
                if col_num == 3:  # TipoLogradouro_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "TipoLogradouro_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tipo logradouro: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('tipologradouro.index'))

@tipologradouro_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_tipologradouro_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet
        ws = wb.create_sheet(title="TipoLogradouro_Filtrado")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna TipoLogradouro_Codigo
                if header == 'TipoLogradouro_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="TipoLogradouro_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tipo logradouro filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@tipologradouro_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela TipoLogradouro do banco homólogo (BancoHomo)"""
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter o BancoHomo diretamente do banco de dados
        banco_homo = obter_banco_homo(projeto_id)
        
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('tipologradouro.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")
        
        # Conectar ao banco homólogo
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('tipologradouro.index'))
        
        # Executar consulta na tabela TipoLogradouro do banco homólogo
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoLogradouro_Codigo, TipoLogradouro_Descricao, TipoLogradouro_Ativo FROM TipoLogradouro")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        # Converter para lista de dicionários
        registros_list = [dict(zip(colunas, row)) for row in registros]
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Criar DataFrame pandas
        df = pd.DataFrame(registros_list, columns=colunas)
        
        # Criar buffer para o arquivo Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TipoLogradouro_WF', index=False)
        
        output.seek(0)
        
        # Nome do arquivo
        nome_arquivo = "TipoLogradouro_WF.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('tipologradouro.index'))

@tipologradouro_bp.route('/importar', methods=['POST'])
def importar_tipologradouro():
    """Importa a planilha de TipoLogradouro e atualiza/inclui registros corretamente"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        if not (arquivo.filename.endswith('.xlsx') or arquivo.filename.endswith('.xls')):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # === Leitura da planilha (forçando strings) ===
        df = pd.read_excel(arquivo, dtype=str)
        df = df.replace({pd.NaT: None, 'nan': None, 'NaN': None, '': None})
        df = df.where(pd.notnull(df), None)
        registros = df.to_dict('records')
        colunas_excel = df.columns.tolist()
        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")

        # === Mapeamento de colunas ===
        mapeamento_colunas = {
            'Logradouro Sigla': 'logradouro_sigla',
            'Logradouro Nome': 'logradouro_nm',
            'TipoLogradouro_Codigo': 'TipoLogradouro_Codigo',
            'TipoLogradouro_Sigla': 'TipoLogradouro_Sigla',
            'TipoLogradouro_Descricao': 'TipoLogradouro_Descricao',
            'Tabela': 'tabela',
            'logradouro_sigla': 'logradouro_sigla',
            'logradouro_nm': 'logradouro_nm',
            'tabela': 'tabela'
        }

        colunas_banco = [
            'logradouro_sigla', 'logradouro_nm',
            'TipoLogradouro_Codigo', 'TipoLogradouro_Sigla',
            'TipoLogradouro_Descricao', 'tabela'
        ]

        # === Normalização e filtragem ===
        registros_mapeados = []
        for registro in registros:
            novo = {}
            for coluna_banco in colunas_banco:
                valor = None
                for chave, destino in mapeamento_colunas.items():
                    if destino == coluna_banco and chave in colunas_excel:
                        valor = registro.get(chave)
                        break
                if valor is not None:
                    valor = str(valor).strip() if not isinstance(valor, str) else valor.strip()
                    if valor == '':
                        valor = None
                novo[coluna_banco] = valor
            # Só adiciona se tiver pelo menos sigla ou nome
            if novo.get('logradouro_sigla') or novo.get('logradouro_nm'):
                registros_mapeados.append(novo)

        logger.info(f"Registros mapeados após limpeza: {len(registros_mapeados)}")

        # === Importação (UPDATE/INSERT) com matching robusto ===
        atualizados = 0
        inseridos = 0
        for i, registro in enumerate(registros_mapeados):
            sigla = registro.get('logradouro_sigla') or ''
            nome = registro.get('logradouro_nm') or ''
            sigla = sigla.strip()
            nome = nome.strip()

            # Normaliza para comparação (maiusculas)
            sigla_comp = sigla.upper() if sigla else ''
            nome_comp = nome.upper() if nome else ''

            # Se não houver chave válida, pula
            if not sigla_comp and not nome_comp:
                continue

            # Monta SELECT dependendo das chaves disponíveis
            if sigla_comp and nome_comp:
                # Usar ambos para matching
                cursor.execute("""
                    SELECT id FROM TipoLogradouro_DePara
                    WHERE UPPER(LTRIM(RTRIM(ISNULL(logradouro_sigla, '')))) = ?
                      AND UPPER(LTRIM(RTRIM(ISNULL(logradouro_nm, '')))) = ?
                """, (sigla_comp, nome_comp))
            elif sigla_comp:
                cursor.execute("""
                    SELECT id FROM TipoLogradouro_DePara
                    WHERE UPPER(LTRIM(RTRIM(ISNULL(logradouro_sigla, '')))) = ?
                """, (sigla_comp,))
            else:  # nome_comp
                cursor.execute("""
                    SELECT id FROM TipoLogradouro_DePara
                    WHERE UPPER(LTRIM(RTRIM(ISNULL(logradouro_nm, '')))) = ?
                """, (nome_comp,))

            existe = cursor.fetchone()

            if existe:
                # UPDATE por id
                cursor.execute("""
                    UPDATE TipoLogradouro_DePara
                    SET TipoLogradouro_Codigo = ?, 
                        TipoLogradouro_Sigla = ?, 
                        TipoLogradouro_Descricao = ?, 
                        tabela = ?
                    WHERE id = ?
                """, (
                    registro.get('TipoLogradouro_Codigo'),
                    registro.get('TipoLogradouro_Sigla'),
                    registro.get('TipoLogradouro_Descricao'),
                    registro.get('tabela'),
                    existe[0]
                ))
                atualizados += 1
            else:
                # INSERT
                cursor.execute("""
                    INSERT INTO TipoLogradouro_DePara
                    (logradouro_sigla, logradouro_nm, TipoLogradouro_Codigo, 
                     TipoLogradouro_Sigla, TipoLogradouro_Descricao, tabela)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    registro.get('logradouro_sigla'),
                    registro.get('logradouro_nm'),
                    registro.get('TipoLogradouro_Codigo'),
                    registro.get('TipoLogradouro_Sigla'),
                    registro.get('TipoLogradouro_Descricao'),
                    registro.get('tabela')
                ))
                inseridos += 1

            if i % 100 == 0:
                conexao.commit()

        conexao.commit()
        logger.info(f"Importação finalizada — {atualizados} atualizados, {inseridos} inseridos.")

        # === Atualizar descrições via WF (se aplicável) ===
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        # fechar recursos
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conexao.close()
        except Exception:
            pass

        return jsonify({
            'success': True,
            'message': f'Importação concluída! {atualizados} atualizados, {inseridos} inseridos.'
        })

    except Exception as e:
        logger.error(f"Erro na importação de TipoLogradouro: {e}", exc_info=True)
        if 'conexao' in locals() and conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")
        return jsonify({'success': False, 'message': f'Erro na importação: {str(e)}'})



@tipologradouro_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados batch recebidos: {len(data.get('updates', []))} atualizações")
        
        updates = data.get('updates', [])
        
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        success_count = 0
        error_count = 0
        error_messages = []
        
        # Processar cada atualização
        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')
                
                if not record_id or not field:
                    error_count += 1
                    error_messages.append(f"ID ou campo não fornecidos para atualização: {update}")
                    continue
                
                # Atualizar registro
                query = f"UPDATE TipoLogradouro_DePara SET {field} = ? WHERE id = ?"
                cursor.execute(query, (value, record_id))
                
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")
                    
            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")
        
        # Commit de todas as atualizações
        conexao.commit()
        
        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")
        
        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }
        
        if error_messages:
            response['error_details'] = error_messages[:10]  # Limitar a 10 mensagens de erro
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}")
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})
    
    finally:
        if cursor:
            cursor.close()
        if conexao:
            conexao.close()

@tipologradouro_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})
        
        descricao = obter_descricao_wf(banco_homo, codigo)
        
        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })
            
    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})