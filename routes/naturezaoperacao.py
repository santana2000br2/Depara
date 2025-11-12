from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd

naturezaoperacao_bp = Blueprint("naturezaoperacao", __name__)

def obter_banco_homo(projeto_id):
    """Retorna o BancoHomo para o projeto (mesma lógica dos outros módulos)."""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if resultado and resultado[0]:
            return resultado[0]
        return None
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo (naturezaoperacao): {e}")
        return None

def obter_codigos_wf(banco_homo):
    """Retorna lista de códigos existentes na tabela WF (NaturezaOperacao)."""
    try:
        if not banco_homo:
            return []
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        cursor = conexao.cursor()
        cursor.execute("SELECT NaturezaOperacao_Codigo FROM NaturezaOperacao")
        registros = cursor.fetchall()
        cursor.close()
        conexao.close()
        codigos = [str(r[0]) for r in registros if r and r[0] is not None]
        logger.info(f"Encontrados {len(codigos)} códigos WF (NaturezaOperacao)")
        return codigos
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF (naturezaoperacao): {e}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Retorna a descrição da tabela WF para um código específico."""
    try:
        if not banco_homo or not codigo:
            return None
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        cursor = conexao.cursor()
        cursor.execute("SELECT NaturezaOperacao_Descricao FROM NaturezaOperacao WHERE NaturezaOperacao_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        cursor.close()
        conexao.close()
        return resultado[0] if resultado and resultado[0] else None
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF (naturezaoperacao) para código {codigo}: {e}")
        return None

@naturezaoperacao_bp.route("/")
def index():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        projeto_nome = projeto.get('NomeProjeto', 'N/A')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('naturezaoperacao.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM NaturezaOperacao_DePara")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros_dict = [dict(zip(colunas, r)) for r in registros]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        cursor.close()
        conexao.close()

        return render_template('naturezaoperacao.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf,
                               banco_homo=banco_homo)
    except Exception as e:
        logger.error(f"Erro em naturezaoperacao.index: {e}")
        flash(f'Erro: {e}', 'error')
        return render_template('naturezaoperacao.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@naturezaoperacao_bp.route('/exportar')
def exportar_depara():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('naturezaoperacao.index'))

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('naturezaoperacao.index'))

        cursor = conexao.cursor()
        cursor.execute("""
            SELECT me_cd, me_ds, dep_cd, plan_cd, int_cd, Tipo,
                   NaturezaOperacao_Codigo, NaturezaOperacao_Descricao,
                   Departamento_Codigo, Procedure_Origem
            FROM NaturezaOperacao_DePara
        """)
        registros = cursor.fetchall()
        colunas_originais = [c[0] for c in cursor.description]

        # Mapear colunas para nomes amigáveis
        mapeamento = {
            'me_cd': 'Codigo de Origem',
            'me_ds': 'Descrição de origem',
            'dep_cd': 'dep_cd',
            'plan_cd': 'plan_cd',
            'int_cd': 'int_cd',
            'Tipo': 'Tipo',
            'NaturezaOperacao_Codigo': 'NaturezaOperacao_Codigo',
            'NaturezaOperacao_Descricao': 'NaturezaOperacao_Descricao',
            'Departamento_Codigo': 'Departamento_Codigo',
            'Procedure_Origem': 'Procedure_Origem'
        }
        colunas_amigaveis = [mapeamento.get(c, c) for c in colunas_originais]

        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="NaturezaOperacao_DePara")

        # Cabeçalhos
        for idx, h in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Preencher dados e colorir coluna NaturezaOperacao_Codigo (índice conforme SELECT)
        # SELECT order: me_cd(1), me_ds(2), dep_cd(3), plan_cd(4), int_cd(5), Tipo(6),
        # NaturezaOperacao_Codigo(7), NaturezaOperacao_Descricao(8), Departamento_Codigo(9), Procedure_Origem(10)
        for row_num, row in enumerate(registros, 2):
            for col_num, val in enumerate(row, 1):
                if isinstance(val, str):
                    cell_value = val.strip()
                else:
                    cell_value = val
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                if col_num == 7:  # NaturezaOperacao_Codigo
                    if not cell_value or cell_value == '':
                        cell.fill = laranja
                    elif cell_value == 'S/DePara':
                        cell.fill = amarelo
                    elif str(cell_value) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        # Ajustar largura
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for c in column:
                try:
                    if c.value and len(str(c.value)) > max_length:
                        max_length = len(str(c.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        cursor.close()
        conexao.close()

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="NaturezaOperacao_DePara.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar naturezaoperacao: {e}")
        flash(f'Erro na exportação: {e}', 'error')
        return redirect(url_for('naturezaoperacao.index'))

@naturezaoperacao_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_filtrados():
    """Exporta apenas os registros filtrados com coloração."""
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        projeto = session['projeto_selecionado']
        projeto_id = projeto.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="NaturezaOperacao_Filtrado")

        # Cabeçalhos
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for rnum, registro in enumerate(registros_filtrados, 2):
            for cnum, header in enumerate(headers, 1):
                val = registro.get(header, '')
                if isinstance(val, str):
                    val = val.strip()
                cell = ws.cell(row=rnum, column=cnum, value=val)
                if header == 'NaturezaOperacao_Codigo':
                    if not val or val == '':
                        cell.fill = laranja
                    elif val == 'S/DePara':
                        cell.fill = amarelo
                    elif val and str(val) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="NaturezaOperacao_Filtrado.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar naturezaoperacao filtrado: {e}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {e}'}), 500


@naturezaoperacao_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela WF (NaturezaOperacao)"""
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('naturezaoperacao.index'))

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('naturezaoperacao.index'))

        cursor = conexao.cursor()
        cursor.execute("select NaturezaOperacao_Codigo, NaturezaOperacao_Descricao, NaturezaOperacao_Ativo from NaturezaOperacao")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros_list = [dict(zip(colunas, r)) for r in registros]
        cursor.close()
        conexao.close()

        df = pd.DataFrame(registros_list, columns=colunas)
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='NaturezaOperacao_WF', index=False)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="NaturezaOperacao_WF.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar WF (naturezaoperacao): {e}")
        flash(f'Erro na exportação WF: {e}', 'error')
        return redirect(url_for('naturezaoperacao.index'))

@naturezaoperacao_bp.route('/importar', methods=['POST'])
def importar():
    try:
        # arquivo
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato inválido. Use .xlsx ou .xls'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()

        colunas_importacao = [
            'me_cd',
            'me_ds',
            'dep_cd',
            'plan_cd',
            'int_cd',
            'Tipo',
            'NaturezaOperacao_Codigo',
            'NaturezaOperacao_Descricao',
            'Departamento_Codigo',
            'Procedure_Origem'
        ]

        # Ler Excel
        try:
            df = pd.read_excel(arquivo)
            df = df.where(pd.notnull(df), None)
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()
        except Exception as e:
            logger.error(f"Erro ao ler Excel (naturezaoperacao): {e}")
            return jsonify({'success': False, 'message': f'Erro ao ler Excel: {e}'})

        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")

        # Mapeamento amigável -> técnico
        mapeamento = {
            'Codigo de Origem': 'me_cd',
            'Descrição de origem': 'me_ds',
            'dep_cd': 'dep_cd',
            'plan_cd': 'plan_cd',
            'int_cd': 'int_cd',
            'Tipo': 'Tipo',
            'NaturezaOperacao_Codigo': 'NaturezaOperacao_Codigo',
            'NaturezaOperacao_Descricao': 'NaturezaOperacao_Descricao',
            'Departamento_Codigo': 'Departamento_Codigo',
            'Procedure_Origem': 'Procedure_Origem'
        }

        colunas_normalizadas = []
        for c in colunas_excel:
            colunas_normalizadas.append(mapeamento.get(c, c))
        df.columns = colunas_normalizadas

        colunas_excel_set = set(colunas_normalizadas)
        colunas_necessarias = set(colunas_importacao)
        if not colunas_necessarias.issubset(colunas_excel_set):
            missing = colunas_necessarias - colunas_excel_set
            return jsonify({'success': False, 'message': f'Colunas necessárias faltando no arquivo: {", ".join(missing)}'})

        registros = df.to_dict('records')

        # Tratar e limitar tamanhos
        registros_filtrados = []
        for reg in registros:
            r = {}
            for col in colunas_importacao:
                val = reg.get(col)
                if val is None:
                    r[col] = None
                else:
                    s = str(val)
                    if col in ['me_cd', 'dep_cd', 'plan_cd', 'int_cd', 'Tipo', 'Procedure_Origem']:
                        r[col] = s[:150]
                    elif col in ['me_ds']:
                        r[col] = s[:500]
                    elif col in ['NaturezaOperacao_Codigo', 'Departamento_Codigo']:
                        r[col] = s[:100]
                    elif col == 'NaturezaOperacao_Descricao':
                        r[col] = s[:100]
                    else:
                        r[col] = s
            registros_filtrados.append(r)

        logger.info(f"Registros tratados: {len(registros_filtrados)}")

        # Contagem antes
        cursor.execute("SELECT COUNT(*) FROM NaturezaOperacao_DePara")
        antes = cursor.fetchone()
        count_antes = antes[0] if antes else 0
        logger.info(f"Antes: {count_antes}")

        contador_update = 0
        contador_insert = 0

        # Inserir ou atualizar (por me_cd)
        for reg in registros_filtrados:
            chave_origem = reg.get('me_cd')
            codigo_wf = reg.get('NaturezaOperacao_Codigo')
            desc_wf = reg.get('NaturezaOperacao_Descricao')
            if chave_origem:
                cursor.execute("SELECT id FROM NaturezaOperacao_DePara WHERE me_cd = ?", (chave_origem,))
                existe = cursor.fetchone()
                if existe:
                    cursor.execute("""
                        UPDATE NaturezaOperacao_DePara
                        SET me_ds = ?, dep_cd = ?, plan_cd = ?, int_cd = ?, Tipo = ?,
                            NaturezaOperacao_Codigo = ?, NaturezaOperacao_Descricao = ?, Departamento_Codigo = ?, Procedure_Origem = ?
                        WHERE me_cd = ?
                    """, (
                        reg.get('me_ds'),
                        reg.get('dep_cd'),
                        reg.get('plan_cd'),
                        reg.get('int_cd'),
                        reg.get('Tipo'),
                        codigo_wf,
                        desc_wf,
                        reg.get('Departamento_Codigo'),
                        reg.get('Procedure_Origem'),
                        chave_origem
                    ))
                    contador_update += 1
                else:
                    placeholders = ', '.join(['?' for _ in colunas_importacao])
                    col_str = ', '.join(colunas_importacao)
                    valores = [reg.get(c) for c in colunas_importacao]
                    cursor.execute(f"INSERT INTO NaturezaOperacao_DePara ({col_str}) VALUES ({placeholders})", valores)
                    contador_insert += 1
            else:
                # sem chave de origem -> inserir linha mesmo assim
                placeholders = ', '.join(['?' for _ in colunas_importacao])
                col_str = ', '.join(colunas_importacao)
                valores = [reg.get(c) for c in colunas_importacao]
                cursor.execute(f"INSERT INTO NaturezaOperacao_DePara ({col_str}) VALUES ({placeholders})", valores)
                contador_insert += 1

        conexao.commit()
        logger.info(f"UPDATEs: {contador_update}, INSERTs: {contador_insert}")

        # Atualizar descrições com base no banco homólogo (se configurado)
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            logger.info("Atualizando NaturezaOperacao_Descricao a partir do WF...")
            conexao_aux = conectar_segunda_base(banco_usuario)
            if conexao_aux:
                cursor_aux = conexao_aux.cursor()
                cursor_aux.execute("""
                    SELECT id, NaturezaOperacao_Codigo, NaturezaOperacao_Descricao
                    FROM NaturezaOperacao_DePara
                    WHERE NaturezaOperacao_Codigo IS NOT NULL AND NaturezaOperacao_Codigo != 'S/DePara'
                """)
                rows = cursor_aux.fetchall()
                for r in rows:
                    idr, cod, desc_atual = r
                    desc_from_wf = obter_descricao_wf(banco_homo, cod)
                    if desc_from_wf and desc_from_wf != desc_atual:
                        cursor_aux.execute("UPDATE NaturezaOperacao_DePara SET NaturezaOperacao_Descricao = ? WHERE id = ?", (desc_from_wf, idr))
                conexao_aux.commit()
                cursor_aux.close()
                conexao_aux.close()

        # Contagem depois
        cursor.execute("SELECT COUNT(*) FROM NaturezaOperacao_DePara")
        depois = cursor.fetchone()
        count_depois = depois[0] if depois else 0

        cursor.close()
        conexao.close()

        return jsonify({'success': True, 'message': f'Importação concluída! {contador_update} atualizados, {contador_insert} inseridos. Antes: {count_antes}, Depois: {count_depois}.'})
    except Exception as e:
        logger.error(f"Erro ao importar naturezaoperacao: {e}", exc_info=True)
        # rollback/fechar com segurança
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if 'cursor' in locals() and cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro na importação: {e}'}), 500

@naturezaoperacao_bp.route('/update', methods=['POST'])
def update_registro():
    """Atualiza individualmente NaturezaOperacao_Codigo ou NaturezaOperacao_Descricao via edição inline."""
    logger.info("=== UPDATE NATUREZA OPERACAO ===")
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not record_id or not field:
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'}), 400

        campos_permitidos = ['NaturezaOperacao_Codigo', 'NaturezaOperacao_Descricao']
        if field not in campos_permitidos:
            return jsonify({'success': False, 'message': 'Campo não permitido para edição'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'}), 500

        cursor = conexao.cursor()

        if field == 'NaturezaOperacao_Codigo':
            if value != 'S/DePara' and value is not None and not str(value).isdigit():
                return jsonify({'success': False, 'message': 'NaturezaOperacao_Codigo deve ser somente números ou "S/DePara"'}), 400

        cursor.execute(f"UPDATE NaturezaOperacao_DePara SET {field} = ? WHERE id = ?", (value, record_id))
        conexao.commit()

        # se alterou o código, tentar atualizar descrição a partir do WF
        if field == 'NaturezaOperacao_Codigo':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo and value and value != 'S/DePara':
                descricao = obter_descricao_wf(banco_homo, value)
                if descricao:
                    try:
                        cursor.execute("UPDATE NaturezaOperacao_DePara SET NaturezaOperacao_Descricao = ? WHERE id = ?", (descricao, record_id))
                        conexao.commit()
                        logger.info(f"Descrição atualizada automaticamente id={record_id}")
                    except Exception as ex:
                        logger.warning(f"Falha ao atualizar descrição automaticamente: {ex}")

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
    except Exception as e:
        logger.error(f"Erro no update (naturezaoperacao): {e}", exc_info=True)
        # rollback e fechar com segurança
        if conexao is not None:
            try:
                conexao.rollback()
            except Exception:
                pass
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conexao is not None:
            try:
                conexao.close()
            except Exception:
                pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {e}'}), 500
