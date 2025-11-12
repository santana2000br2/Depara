# profissao.py
from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

profissao_bp = Blueprint("profissao", __name__)


def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None

        cursor = conn.cursor()
        # passar parâmetros como tupla
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", (projeto_id,))
        resultado = cursor.fetchone()

        cursor.close()
        conn.close()

        if resultado and resultado[0]:
            return resultado[0]
        return None

    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None


def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela Profissao do banco homólogo"""
    try:
        if not banco_homo:
            return []

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []

        cursor = conexao.cursor()
        cursor.execute("SELECT Profissao_Codigo FROM Profissao")
        registros = cursor.fetchall()

        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro and registro[0] is not None]

        cursor.close()
        conexao.close()

        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos

    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []


def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela Profissao do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None

        cursor = conexao.cursor()
        cursor.execute("SELECT Profissao_Descricao FROM Profissao WHERE Profissao_Codigo = ?", (codigo,))
        resultado = cursor.fetchone()

        cursor.close()
        conexao.close()

        if resultado and resultado[0]:
            return resultado[0]
        return None

    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None


def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza Profissao_Descricao registro a registro usando o banco homólogo (WF)."""
    try:
        if not banco_homo:
            logger.error("Banco homólogo não informado para atualização de descrições.")
            return False

        # Conectar ao banco do usuário
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar ao banco do usuário para atualizar descrições.")
            return False
        cursor = conexao.cursor()

        # Buscar registros que têm código WF (não são S/DePara) - trazer id para atualizar por id
        cursor.execute("""
            SELECT id, Profissao_Codigo, Profissao_Descricao
            FROM Profissao_DePara
            WHERE Profissao_Codigo IS NOT NULL AND Profissao_Codigo <> 'S/DePara' AND LTRIM(RTRIM(Profissao_Codigo)) <> ''
        """)
        registros = cursor.fetchall()

        if not registros:
            logger.info("Nenhum registro com código válido encontrado para atualização.")
            cursor.close()
            conexao.close()
            return True

        # Conectar ao banco WF (homologação) uma vez
        conexao_wf = conectar_segunda_base(banco_homo)
        if not conexao_wf:
            logger.error(f"Falha ao conectar ao banco homólogo: {banco_homo}")
            cursor.close()
            conexao.close()
            return False
        cursor_wf = conexao_wf.cursor()

        atualizacoes = 0
        for registro in registros:
            id_registro = registro[0]
            codigo_wf = registro[1]
            descricao_atual = registro[2]

            if codigo_wf is None:
                continue

            # Normalizar código antes de buscar
            codigo_norm = str(codigo_wf).strip()

            if codigo_norm == '' or codigo_norm.upper() == 'S/DEPARA':
                continue

            try:
                # Buscar descrição correspondente no WF
                cursor_wf.execute("SELECT Profissao_Descricao FROM Profissao WHERE LTRIM(RTRIM(Profissao_Codigo)) = ?", (codigo_norm,))
                resultado = cursor_wf.fetchone()
                descricao_wf = resultado[0] if resultado else None

                # Se encontrou descrição e for diferente da atual, atualiza somente esse id
                if descricao_wf and descricao_wf != descricao_atual:
                    cursor.execute("""
                        UPDATE Profissao_DePara
                        SET Profissao_Descricao = ?
                        WHERE id = ?
                    """, (descricao_wf, id_registro))
                    atualizacoes += 1
                    logger.info(f"Atualizado id {id_registro} (codigo {codigo_norm}) para descrição: {descricao_wf}")

            except Exception as e:
                logger.warning(f"Erro ao consultar/descrever código {codigo_norm} no WF: {e}")

        # Commit e fechamento
        conexao.commit()
        cursor_wf.close()
        conexao_wf.close()
        cursor.close()
        conexao.close()

        logger.info(f"Atualizações automáticas de descrição concluídas: {atualizacoes} registros atualizados.")
        return True

    except Exception as e:
        logger.error(f"Erro ao atualizar descrições após importação: {e}", exc_info=True)
        try:
            if 'conexao' in locals() and conexao:
                conexao.rollback()
        except Exception:
            pass
        return False



# -------------------------
# Rotas
# -------------------------
@profissao_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('profissao.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Profissao_DePara")
        registros = cursor.fetchall()

        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description] if cursor.description else []

        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]

        logger.info(f"Encontrados {len(registros_dict)} registros")

        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # Fechar recursos
        cursor.close()
        conexao.close()

        return render_template('profissao.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf,
                               banco_homo=banco_homo)

    except Exception as e:
        logger.error(f"Erro em profissao: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('profissao.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])


@profissao_bp.route('/exportar')
def exportar_profissao():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('profissao.index'))

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('profissao.index'))

        # Executar consulta - Excluir coluna id da exportação
        cursor = conexao.cursor()
        cursor.execute("SELECT prof_cd, prof_ds, Profissao_Codigo, Profissao_Descricao FROM Profissao_DePara")
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description] if cursor.description else []

        # Mapear nomes das colunas para os nomes amigáveis
        mapeamento_colunas = {
            'prof_cd': 'Cod. Profissão Origem',
            'prof_ds': 'Profissão Descrição',
            'Profissao_Codigo': 'Profissao_Codigo',
            'Profissao_Descricao': 'Profissao_Descricao'
        }

        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]

        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # CORREÇÃO: Criar workbook e worksheet de forma mais robusta
        wb = Workbook()

        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)

        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="Profissao_DePara")

        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios

        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()

                cell = ws.cell(row=row_num, column=col_num, value=valor)

                # Aplicar cores na coluna Profissao_Codigo (coluna 3)
                if col_num == 3:  # Profissao_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        # Ajustar largura das colunas automaticamente (forma segura)
        for column in ws.columns:
            if not column or column[0] is None:
                continue
            try:
                column_letter = getattr(column[0], "column_letter", None)
                if not column_letter:
                    continue
            except Exception:
                continue

            max_length = 0
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            try:
                ws.column_dimensions[column_letter].width = adjusted_width
            except Exception:
                pass

        # Fechar recursos
        cursor.close()
        conexao.close()

        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = "Profissao_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar profissao: {str(e)}", exc_info=True)
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('profissao.index'))


@profissao_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_profissao_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])

        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400

        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # Criar workbook
        wb = Workbook()

        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)

        # Criar uma nova worksheet
        ws = wb.create_sheet(title="Profissao_Filtrado")

        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios

        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()

                cell = ws.cell(row=row_num, column=col_num, value=valor)

                # Aplicar cores na coluna Profissao_Codigo
                if header == 'Profissao_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        # Ajustar largura das colunas automaticamente (forma segura)
        for column in ws.columns:
            if not column or column[0] is None:
                continue
            try:
                column_letter = getattr(column[0], "column_letter", None)
                if not column_letter:
                    continue
            except Exception:
                continue

            max_length = 0
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            try:
                ws.column_dimensions[column_letter].width = adjusted_width
            except Exception:
                pass

        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Profissao_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar profissao filtrada: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500


@profissao_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela Profissao do banco homólogo (BancoHomo)"""
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        # Obter o BancoHomo diretamente do banco de dados
        banco_homo = obter_banco_homo(projeto_id)

        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('profissao.index'))

        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")

        # Conectar ao banco homólogo
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('profissao.index'))

        # Executar consulta na tabela Profissao do banco homólogo
        cursor = conexao.cursor()
        cursor.execute("SELECT Profissao_Codigo, Profissao_Descricao, Profissao_Ativo FROM Profissao")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description] if cursor.description else []

        # Converter para lista de dicionários
        registros_list = [dict(zip(colunas, row)) for row in registros]

        # Fechar recursos
        cursor.close()
        conexao.close()

        # Criar DataFrame pandas
        df = pd.DataFrame(registros_list, columns=colunas)

        # Criar buffer para o arquivo Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Profissao_WF', index=False)

        output.seek(0)

        # Nome do arquivo
        nome_arquivo = "Profissao_WF.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}", exc_info=True)
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('profissao.index'))


@profissao_bp.route('/importar', methods=['POST'])
def importar_profissao():
    """
    Importa a planilha de Profissões e faz UPDATE/INSERT corretos conforme a coluna prof_cd.
    Também atualiza automaticamente Profissao_Descricao com base na base WF.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado.'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado.'})
        if not (arquivo.filename.endswith('.xlsx') or arquivo.filename.endswith('.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use .xlsx ou .xls.'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado.'})

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto.'})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # === Leitura robusta da planilha ===
        df = pd.read_excel(arquivo, dtype=str)
        df = df.replace({pd.NaT: None, 'nan': None, 'NaN': None, '': None})
        df = df.where(pd.notnull(df), None)
        registros = df.to_dict('records')

        # === Mapeamento de colunas ===
        mapeamento_colunas = {
            'Cod. Profissão Origem': 'prof_cd',
            'Profissão Descrição': 'prof_ds',
            'Profissao_Codigo': 'Profissao_Codigo',
            'Profissao_Descricao': 'Profissao_Descricao',
            'prof_cd': 'prof_cd',
            'prof_ds': 'prof_ds'
        }

        colunas_banco = ['prof_cd', 'prof_ds', 'Profissao_Codigo', 'Profissao_Descricao']

        # === Conversão ===
        registros_mapeados = []
        for reg in registros:
            novo = {}
            for coluna_banco in colunas_banco:
                valor = None
                for chave, destino in mapeamento_colunas.items():
                    if destino == coluna_banco and chave in df.columns:
                        valor = reg.get(chave)
                        break
                if valor is not None:
                    valor = str(valor).strip()
                    if valor == '':
                        valor = None
                    elif coluna_banco == 'Profissao_Codigo' and valor.upper() == 'S/DEPARA':
                        valor = 'S/DePara'
                novo[coluna_banco] = valor
            if novo.get('prof_cd'):
                registros_mapeados.append(novo)

        logger.info(f"Registros mapeados após limpeza: {len(registros_mapeados)}")

        # === UPDATE/INSERT ===
        atualizados = 0
        inseridos = 0
        for i, registro in enumerate(registros_mapeados):
            prof_cd = registro['prof_cd']
            prof_ds = registro['prof_ds']
            codigo = registro['Profissao_Codigo']
            descricao = registro['Profissao_Descricao']

            cursor.execute("SELECT id FROM Profissao_DePara WHERE LTRIM(RTRIM(prof_cd)) = ?", (prof_cd,))
            existe = cursor.fetchone()

            if existe:
                cursor.execute("""
                    UPDATE Profissao_DePara
                    SET prof_ds = ?, Profissao_Codigo = ?, Profissao_Descricao = ?
                    WHERE LTRIM(RTRIM(prof_cd)) = ?
                """, (prof_ds, codigo, descricao, prof_cd))
                atualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO Profissao_DePara (prof_cd, prof_ds, Profissao_Codigo, Profissao_Descricao)
                    VALUES (?, ?, ?, ?)
                """, (prof_cd, prof_ds, codigo, descricao))
                inseridos += 1

            if i % 100 == 0:
                conexao.commit()

        conexao.commit()

        # === Atualizar descrições via WF ===
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        msg = f"Importação concluída! {atualizados} registros atualizados, {inseridos} inseridos."
        logger.info(msg)
        return jsonify({'success': True, 'message': msg})

    except Exception as e:
        logger.error(f"Erro na importação de profissões: {e}", exc_info=True)
        if 'conexao' in locals() and conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conexao' in locals() and conexao:
            conexao.close()



@profissao_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline"""
    logger.info("=== UPDATE REGISTRO PROFISSAO ENDPOINT ACESSADO ===")

    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados recebidos: {data}")

        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not record_id or not field:
            logger.error("ID ou campo não fornecidos")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})

        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            logger.error("Nenhum projeto selecionado na sessão")
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            logger.error("Banco não configurado para este projeto")
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        logger.info(f"Atualizando registro {record_id}, campo {field} para valor '{value}' no banco {banco_usuario}")

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error(f"Falha na conexão com o banco: {banco_usuario}")
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()

        # Verificar se a tabela existe e tem a coluna
        try:
            cursor.execute(f"SELECT TOP 1 {field} FROM Profissao_DePara WHERE id = ?", (record_id,))
            resultado = cursor.fetchone()
            if not resultado:
                logger.error(f"Registro não encontrado: {record_id}")
                return jsonify({'success': False, 'message': 'Registro não encontrado'})
        except Exception as e:
            logger.error(f"Erro ao verificar registro: {str(e)}")
            return jsonify({'success': False, 'message': f'Campo {field} não existe na tabela'})

        # Atualizar registro - usando id como chave primária
        query = f"UPDATE Profissao_DePara SET {field} = ? WHERE id = ?"
        logger.info(f"Executando query: {query} com valores: ({value}, {record_id})")

        cursor.execute(query, (value, record_id))

        # Se atualizar o código WF, buscar descrição na base homóloga e gravar Profissao_Descricao
        if field == 'Profissao_Codigo' and value and value != 'S/DePara':
            try:
                banco_homo = obter_banco_homo(projeto_id)
                if banco_homo:
                    conexao_wf = conectar_segunda_base(banco_homo)
                    if conexao_wf:
                        cursor_wf = conexao_wf.cursor()
                        cursor_wf.execute("SELECT Profissao_Descricao FROM Profissao WHERE Profissao_Codigo = ?", (value,))
                        res = cursor_wf.fetchone()
                        descricao_wf = res[0] if res else None
                        cursor_wf.close()
                        conexao_wf.close()
                        if descricao_wf:
                            cursor.execute("UPDATE Profissao_DePara SET Profissao_Descricao = ? WHERE id = ?", (descricao_wf, record_id))
            except Exception as e:
                logger.warning(f"Não foi possível obter Profissao_Descricao do WF para {value}: {e}")

        # Verificar se alguma linha foi afetada (considerar também atualizações subsequentes)
        if cursor.rowcount == 0:
            logger.warning(f"Nenhuma linha afetada pela atualização do registro {record_id}")
            try:
                conexao.rollback()
            except Exception:
                pass
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})

        # Commit da transação
        conexao.commit()

        logger.info(f"Registro {record_id} atualizado com sucesso")
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})

    except Exception as e:
        logger.error(f"Erro ao atualizar registro: {str(e)}", exc_info=True)
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")

        # Fazer rollback em caso de erro apenas se a conexão existir
        if conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")

        return jsonify({'success': False, 'message': f'Erro ao atualizar registro: {str(e)}'})

    finally:
        # Fechar recursos de forma segura
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            logger.error(f"Erro ao fechar cursor: {e}")

        try:
            if conexao:
                conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexão: {e}")


@profissao_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")

    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados batch recebidos: {len(data.get('updates', []))} atualizações")

        updates = data.get('updates', [])

        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})

        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()
        success_count = 0
        error_count = 0
        error_messages = []

        # Processar cada atualização
        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')

                if not record_id or not field:
                    error_count += 1
                    error_messages.append(f"ID ou campo não fornecidos para atualização: {update}")
                    continue

                # Atualizar registro
                query = f"UPDATE Profissao_DePara SET {field} = ? WHERE id = ?"
                cursor.execute(query, (value, record_id))

                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")

            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")

        # Commit de todas as atualizações
        conexao.commit()

        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")

        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }

        if error_messages:
            response['error_details'] = error_messages[:10]  # Limitar a 10 mensagens de erro

        return jsonify(response)

    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}", exc_info=True)
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})

    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao:
                conexao.close()
        except Exception:
            pass


@profissao_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})

        descricao = obter_descricao_wf(banco_homo, codigo)

        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })

    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})
