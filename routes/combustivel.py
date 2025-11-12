from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import tempfile
import os

combustivel_bp = Blueprint("combustivel", __name__)

def safe_convert_id(value):
    """Converte valores para inteiro de forma segura"""
    if value is None:
        return None
    try:
        if isinstance(value, float):
            return int(value)
        elif isinstance(value, str):
            cleaned = value.strip()
            if cleaned == '':
                return None
            # Remove caracteres não numéricos, exceto ponto e vírgula
            cleaned = ''.join(c for c in cleaned if c.isdigit() or c in '.-')
            if cleaned in ['', '.', '-']:
                return None
            return int(float(cleaned)) if '.' in cleaned else int(cleaned)
        elif isinstance(value, int):
            return value
        else:
            return int(value)
    except (ValueError, TypeError, AttributeError):
        raise ValueError(f"Valor não pode ser convertido para inteiro: {value}")

def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", (projeto_id,))
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela Combustivel do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Combustivel_Codigo FROM Combustivel")
        registros = cursor.fetchall()
        
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela Combustivel do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Combustivel_Descricao FROM Combustivel WHERE Combustivel_Codigo = ?", (codigo,))
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None

def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições após importação baseado nos códigos WF"""
    try:
        if not banco_homo:
            return
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições")
            return
        
        cursor = conexao.cursor()
        
        cursor.execute("""
            SELECT id, Combustivel_Codigo, Combustivel_Descricao 
            FROM Combustivel_DePara 
            WHERE Combustivel_Codigo IS NOT NULL AND Combustivel_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()
        
        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)
            
            if descricao_wf and descricao_wf != descricao_atual:
                cursor.execute("""
                    UPDATE Combustivel_DePara 
                    SET Combustivel_Descricao = ? 
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")
        
        conexao.commit()
        cursor.close()
        conexao.close()
        
        logger.info(f"Atualizações automáticas de descrição: {atualizacoes} registros")
        
    except Exception as e:
        logger.error(f"Erro ao atualizar descrições após importação: {str(e)}")

def safe_fetchone(cursor):
    """Função segura para fetchone que evita None is not subscriptable"""
    try:
        result = cursor.fetchone()
        if result is not None and result[0] is not None:
            return result[0]
        return 0
    except Exception as e:
        logger.error(f"Erro em safe_fetchone: {str(e)}")
        return 0

@combustivel_bp.route("/")
def index():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('combustivel.html', 
                                 registros=[], 
                                 colunas=[], 
                                 projeto_nome=projeto_nome, 
                                 banco_usuario=banco_usuario, 
                                 codigos_wf=[])
        
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Combustivel_DePara")
        registros = cursor.fetchall()
        
        colunas = [column[0] for column in cursor.description]
        
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos WF para validação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = []
        if banco_homo:
            codigos_wf = obter_codigos_wf(banco_homo)
        else:
            logger.warning("Banco homólogo não encontrado para o projeto")
        
        cursor.close()
        conexao.close()
        
        return render_template('combustivel.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em combustivel: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('combustivel.html', 
                             registros=[], 
                             colunas=[], 
                             projeto_nome='N/A', 
                             banco_usuario='N/A', 
                             codigos_wf=[])

@combustivel_bp.route('/exportar')
def exportar_combustivel():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('combustivel.index'))
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('combustivel.index'))
        
        cursor = conexao.cursor()
        cursor.execute("SELECT id, comb_cd, comb_ds, Combustivel_Codigo, Combustivel_Descricao FROM Combustivel_DePara")
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]
        
        mapeamento_colunas = {
            'id': 'ID',
            'comb_cd': 'Codigo de Origem',
            'comb_ds': 'Descrição de origem',
            'Combustivel_Codigo': 'Combustivel_Codigo',
            'Combustivel_Descricao': 'Combustivel_Descricao'
        }
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        wb = Workbook()
        
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        ws = wb.create_sheet(title="Combustivel_DePara")
        
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                if col_num == 4:  # Combustivel_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        cursor.close()
        conexao.close()
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "Combustivel_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar combustivel: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('combustivel.index'))

@combustivel_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_combustivel_filtrados():
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        wb = Workbook()
        
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        ws = wb.create_sheet(title="Combustivel_Filtrado")
        
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                if header == 'Combustivel_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="Combustivel_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar combustivel filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@combustivel_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela Combustivel do banco homólogo (BancoHomo)"""
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('combustivel.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('combustivel.index'))
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Combustivel_Codigo, Combustivel_Descricao, Combustivel_Ativo FROM Combustivel")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        registros_list = [dict(zip(colunas, row)) for row in registros]
        
        cursor.close()
        conexao.close()
        
        df = pd.DataFrame(registros_list, columns=colunas)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Combustivel_WF', index=False)
        
        output.seek(0)
        
        nome_arquivo = "Combustivel_WF.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('combustivel.index'))

@combustivel_bp.route('/importar', methods=['POST'])
def importar_combustivel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['file']
        
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})
        
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Iniciando importação para o banco: {banco_usuario}")
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        try:
            # Ler o arquivo Excel mantendo os tipos originais
            df = pd.read_excel(arquivo, dtype=str, na_values=['', ' ', 'NULL', 'null'])
            df = df.where(pd.notnull(df), None)
            
            # Converter todas as strings para strip
            for col in df.columns:
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
                
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel: {str(e)}")
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {str(e)}'})
        
        registros = df.to_dict('records')
        colunas_excel = df.columns.tolist()
        
        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")
        
        # MAPEAMENTO DIRETO DAS COLUNAS - CORRIGIDO
        mapeamento_colunas = {
            'ID': 'id',
            'Codigo de Origem': 'comb_cd',
            'Descrição de origem': 'comb_ds',
            'Combustivel_Codigo': 'Combustivel_Codigo',
            'Combustivel_Descricao': 'Combustivel_Descricao'
        }
        
        colunas_banco = ['id', 'comb_cd', 'comb_ds', 'Combustivel_Codigo', 'Combustivel_Descricao']
        
        # VERIFICAÇÃO SIMPLIFICADA DAS COLUNAS
        colunas_faltantes = []
        for coluna_planilha, coluna_banco in mapeamento_colunas.items():
            if coluna_planilha not in colunas_excel:
                colunas_faltantes.append(coluna_banco)
        
        if colunas_faltantes:
            return jsonify({
                'success': False, 
                'message': f'Colunas necessárias faltando no arquivo: {", ".join(colunas_faltantes)}. '
                          f'Certifique-se de que a planilha contém as colunas: {", ".join(mapeamento_colunas.keys())}'
            })
        
        # MAPEAMENTO DOS REGISTROS - VERSÃO SIMPLIFICADA
        registros_mapeados = []
        for registro in registros:
            registro_mapeado = {}
            
            # Mapear cada coluna diretamente
            registro_mapeado['id'] = safe_convert_id(registro.get('ID'))
            registro_mapeado['comb_cd'] = registro.get('Codigo de Origem')
            registro_mapeado['comb_ds'] = registro.get('Descrição de origem')
            registro_mapeado['Combustivel_Codigo'] = registro.get('Combustivel_Codigo')
            registro_mapeado['Combustivel_Descricao'] = registro.get('Combustivel_Descricao')
            
            # Limpar valores string
            for key, value in registro_mapeado.items():
                if value is not None and isinstance(value, str):
                    value = value.strip()
                    if value == '':
                        registro_mapeado[key] = None
                    else:
                        registro_mapeado[key] = value
            
            registros_mapeados.append(registro_mapeado)
        
        logger.info(f"Registros mapeados: {len(registros_mapeados)}")
        
        # Verificar se todos os registros têm ID e comb_cd
        for i, registro in enumerate(registros_mapeados):
            if registro.get('id') is None:
                return jsonify({
                    'success': False, 
                    'message': f'Registro na linha {i+2} não tem ID válido.'
                })
            if registro.get('comb_cd') is None:
                return jsonify({
                    'success': False, 
                    'message': f'Registro na linha {i+2} não tem Código de Origem válido.'
                })
        
        # LÓGICA DE INSERÇÃO/ATUALIZAÇÃO
        cursor.execute("SELECT COUNT(*) FROM Combustivel_DePara")
        result_antes = cursor.fetchone()
        count_antes = safe_fetchone(cursor) if result_antes else 0
        logger.info(f"Registros na tabela ANTES da importação: {count_antes}")
        
        try:
            contador_atualizacoes = 0
            contador_insercoes = 0
            
            for registro in registros_mapeados:
                registro_id = registro.get('id')
                codigo_origem = registro.get('comb_cd')
                
                if registro_id and codigo_origem:
                    cursor.execute("SELECT id FROM Combustivel_DePara WHERE id = ?", (registro_id,))
                    resultado = cursor.fetchone()
                    
                    if resultado:
                        cursor.execute("""
                            UPDATE Combustivel_DePara 
                            SET comb_cd = ?, 
                                comb_ds = ?, 
                                Combustivel_Codigo = ?, 
                                Combustivel_Descricao = ?
                            WHERE id = ?
                        """, (
                            codigo_origem,
                            registro.get('comb_ds'),
                            registro.get('Combustivel_Codigo'),
                            registro.get('Combustivel_Descricao'),
                            registro_id
                        ))
                        contador_atualizacoes += 1
                    else:
                        cursor.execute("""
                            INSERT INTO Combustivel_DePara 
                            (id, comb_cd, comb_ds, Combustivel_Codigo, Combustivel_Descricao)
                            VALUES (?, ?, ?, ?, ?)
                        """, (
                            registro_id,
                            codigo_origem,
                            registro.get('comb_ds'),
                            registro.get('Combustivel_Codigo'),
                            registro.get('Combustivel_Descricao')
                        ))
                        contador_insercoes += 1
            
            logger.info(f"UPDATEs executados: {contador_atualizacoes} registros")
            logger.info(f"INSERTs executados: {contador_insercoes} registros")
            
            logger.info("Executando COMMIT...")
            conexao.commit()
            logger.info("COMMIT executado com sucesso")
            
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                logger.info("Atualizando descrições com base no banco homólogo...")
                atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)
            
        except Exception as e:
            logger.error(f"Erro durante operações de banco: {str(e)}")
            conexao.rollback()
            logger.info("ROLLBACK executado")
            raise e
        
        cursor.execute("SELECT COUNT(*) FROM Combustivel_DePara")
        result_depois = cursor.fetchone()
        count_depois = safe_fetchone(cursor) if result_depois else 0
        logger.info(f"Registros na tabela DEPOIS da importação: {count_depois}")
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True, 
            'message': f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total na base: {count_depois} registros.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao importar combustivel: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False, 
            'message': f'Erro na importação: {str(e)}'
        })

def importar_arquivo_wf(cursor, conexao, registros, colunas_excel, banco_usuario, projeto_id):
    """Processa arquivo que contém apenas dados WF para fazer merge com dados existentes"""
    try:
        # Mapeamento flexível para arquivo WF
        mapeamento_wf = {
            'combustivel_codigo': ['combustivel_codigo', 'codigo', 'código', 'codigowf', 'códigowf', 'cd_wf'],
            'combustivel_descricao': ['combustivel_descricao', 'descricao', 'descrição', 'descricaowf', 'descriçãowf', 'ds_wf'],
            'combustivel_ativo': ['combustivel_ativo', 'ativo', 'status']
        }
        
        # Encontrar mapeamentos
        colunas_encontradas = {}
        for coluna_wf, possiveis_nomes in mapeamento_wf.items():
            encontrou = False
            for coluna_excel in colunas_excel:
                coluna_excel_clean = coluna_excel.lower().strip().replace(' ', '').replace('_', '')
                for possivel_nome in possiveis_nomes:
                    possivel_nome_clean = possivel_nome.lower().strip().replace(' ', '').replace('_', '')
                    if coluna_excel_clean == possivel_nome_clean:
                        colunas_encontradas[coluna_wf] = coluna_excel
                        encontrou = True
                        break
                if encontrou:
                    break
        
        # Verificar se temos pelo menos o código WF
        if 'combustivel_codigo' not in colunas_encontradas:
            return jsonify({
                'success': False, 
                'message': 'Arquivo WF não contém a coluna Combustivel_Codigo. Colunas encontradas: ' + ', '.join(colunas_excel)
            })
        
        # Obter dados atuais da tabela DePara
        cursor.execute("SELECT id, comb_cd, comb_ds FROM Combustivel_DePara")
        dados_atuais = cursor.fetchall()
        
        if not dados_atuais:
            return jsonify({
                'success': False, 
                'message': 'Não há dados de origem na tabela DePara. Primeiro exporte e preencha os dados de origem.'
            })
        
        # Criar dicionário para lookup por código de origem
        dados_por_codigo_origem = {str(row[1]): {'id': row[0], 'comb_cd': row[1], 'comb_ds': row[2]} for row in dados_atuais}
        
        # Processar registros WF
        registros_processados = []
        for registro in registros:
            codigo_wf = registro.get(colunas_encontradas.get('combustivel_codigo'))
            descricao_wf = registro.get(colunas_encontradas.get('combustivel_descricao'))
            ativo_wf = registro.get(colunas_encontradas.get('combustivel_ativo'))
            
            if codigo_wf:
                # Tentar encontrar correspondência por código de origem
                # Neste caso, assumimos que o código WF pode ser igual ao código de origem
                # Ou você pode implementar uma lógica de matching mais sofisticada
                codigo_origem = codigo_wf  # Simplificação - pode precisar de ajuste
                
                if codigo_origem in dados_por_codigo_origem:
                    registro_processado = dados_por_codigo_origem[codigo_origem].copy()
                    registro_processado['Combustivel_Codigo'] = codigo_wf
                    registro_processado['Combustivel_Descricao'] = descricao_wf
                    registros_processados.append(registro_processado)
        
        logger.info(f"Registros processados para merge: {len(registros_processados)}")
        
        if not registros_processados:
            return jsonify({
                'success': False, 
                'message': 'Não foi possível fazer o merge automático. Verifique se os códigos WF correspondem aos códigos de origem.'
            })
        
        # Atualizar registros na base
        contador_atualizacoes = 0
        for registro in registros_processados:
            cursor.execute("""
                UPDATE Combustivel_DePara 
                SET Combustivel_Codigo = ?, Combustivel_Descricao = ?
                WHERE id = ?
            """, (
                registro.get('Combustivel_Codigo'),
                registro.get('Combustivel_Descricao'),
                registro.get('id')
            ))
            contador_atualizacoes += 1
        
        conexao.commit()
        
        # Atualizar descrições com base no banco homólogo
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            logger.info("Atualizando descrições com base no banco homólogo...")
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True, 
            'message': f'Merge concluído! {contador_atualizacoes} registros atualizados com dados WF.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao processar arquivo WF: {str(e)}")
        conexao.rollback()
        raise e

def importar_arquivo_depara(cursor, conexao, registros, colunas_excel, banco_usuario, projeto_id):
    """Processa arquivo DePara completo (com colunas de origem e destino)"""
    try:
        # VERIFICAÇÃO MAIS FLEXÍVEL DAS COLUNAS
        colunas_banco_obrigatorias = ['id', 'comb_cd']
        colunas_banco_opcionais = ['comb_ds', 'Combustivel_Codigo', 'Combustivel_Descricao']
        colunas_banco = colunas_banco_obrigatorias + colunas_banco_opcionais
        
        # Mapeamento flexível de colunas
        mapeamento_flexivel = {
            # Para coluna 'id'
            'id': ['id', 'codigo', 'código', 'index', 'idx'],
            # Para coluna 'comb_cd' 
            'comb_cd': ['comb_cd', 'codigo origem', 'código origem', 'codigo_origem', 'código_origem', 
                       'codigoorigem', 'cd_origem', 'cod_origem'],
            # Para coluna 'comb_ds'
            'comb_ds': ['comb_ds', 'descricao origem', 'descrição origem', 'descricao_origem', 
                       'descrição_origem', 'descricaoorigem', 'ds_origem', 'desc_origem'],
            # Para coluna 'Combustivel_Codigo'
            'Combustivel_Codigo': ['combustivel_codigo', 'codigo wf', 'código wf', 'codigo_wf', 
                                  'código_wf', 'codigowf', 'cd_wf', 'cod_wf'],
            # Para coluna 'Combustivel_Descricao'
            'Combustivel_Descricao': ['combustivel_descricao', 'descricao wf', 'descrição wf', 
                                     'descricao_wf', 'descrição_wf', 'descricaowf', 'ds_wf', 'desc_wf']
        }
        
        # Encontrar mapeamentos
        colunas_encontradas = {}
        colunas_faltantes = []
        
        for coluna_banco in colunas_banco_obrigatorias:
            encontrou = False
            for coluna_excel in colunas_excel:
                coluna_excel_clean = coluna_excel.lower().strip().replace(' ', '').replace('_', '').replace('-', '')
                
                # Verificar mapeamento flexível
                for possivel_nome in mapeamento_flexivel.get(coluna_banco, []):
                    possivel_nome_clean = possivel_nome.lower().strip().replace(' ', '').replace('_', '').replace('-', '')
                    if coluna_excel_clean == possivel_nome_clean:
                        colunas_encontradas[coluna_banco] = coluna_excel
                        encontrou = True
                        break
                
                if encontrou:
                    break
            
            if not encontrou:
                colunas_faltantes.append(coluna_banco)
        
        # Verificar colunas opcionais
        for coluna_banco in colunas_banco_opcionais:
            encontrou = False
            for coluna_excel in colunas_excel:
                coluna_excel_clean = coluna_excel.lower().strip().replace(' ', '').replace('_', '').replace('-', '')
                
                for possivel_nome in mapeamento_flexivel.get(coluna_banco, []):
                    possivel_nome_clean = possivel_nome.lower().strip().replace(' ', '').replace('_', '').replace('-', '')
                    if coluna_excel_clean == possivel_nome_clean:
                        colunas_encontradas[coluna_banco] = coluna_excel
                        encontrou = True
                        break
                
                if encontrou:
                    break
        
        if colunas_faltantes:
            return jsonify({
                'success': False, 
                'message': f'Colunas obrigatórias faltando: {", ".join(colunas_faltantes)}. '
                          f'Colunas encontradas no arquivo: {", ".join(colunas_excel)}. '
                          f'As colunas obrigatórias são: ID e Código de Origem.'
            })
        
        # MAPEAMENTO DOS REGISTROS
        registros_mapeados = []
        for registro in registros:
            registro_mapeado = {}
            
            for coluna_banco in colunas_banco:
                coluna_planilha = colunas_encontradas.get(coluna_banco)
                
                if coluna_planilha:
                    valor = registro.get(coluna_planilha)
                    
                    if coluna_banco == 'id':
                        try:
                            valor = safe_convert_id(valor)
                        except ValueError as e:
                            return jsonify({
                                'success': False, 
                                'message': f'Erro de conversão do ID: {str(e)}'
                            })
                    elif valor is not None and isinstance(valor, str):
                        valor = valor.strip()
                        if valor == '':
                            valor = None
                    
                    registro_mapeado[coluna_banco] = valor
                else:
                    # Para colunas opcionais não encontradas, definir como None
                    registro_mapeado[coluna_banco] = None
            
            registros_mapeados.append(registro_mapeado)
        
        logger.info(f"Registros mapeados: {len(registros_mapeados)}")
        
        # Verificar se todos os registros têm ID e comb_cd
        for i, registro in enumerate(registros_mapeados):
            if registro.get('id') is None:
                return jsonify({
                    'success': False, 
                    'message': f'Registro na linha {i+2} não tem ID válido.'
                })
            if registro.get('comb_cd') is None:
                return jsonify({
                    'success': False, 
                    'message': f'Registro na linha {i+2} não tem Código de Origem válido.'
                })
        
        # CONTINUA COM A LÓGICA EXISTENTE DE INSERÇÃO/ATUALIZAÇÃO
        cursor.execute("SELECT COUNT(*) FROM Combustivel_DePara")
        result_antes = cursor.fetchone()
        count_antes = safe_fetchone(cursor) if result_antes else 0
        logger.info(f"Registros na tabela ANTES da importação: {count_antes}")
        
        try:
            contador_atualizacoes = 0
            contador_insercoes = 0
            
            for registro in registros_mapeados:
                registro_id = registro.get('id')
                codigo_origem = registro.get('comb_cd')
                
                if registro_id and codigo_origem:
                    cursor.execute("SELECT id FROM Combustivel_DePara WHERE id = ?", (registro_id,))
                    resultado = cursor.fetchone()
                    
                    if resultado:
                        cursor.execute("""
                            UPDATE Combustivel_DePara 
                            SET comb_cd = ?, 
                                comb_ds = ?, 
                                Combustivel_Codigo = ?, 
                                Combustivel_Descricao = ?
                            WHERE id = ?
                        """, (
                            codigo_origem,
                            registro.get('comb_ds'),
                            registro.get('Combustivel_Codigo'),
                            registro.get('Combustivel_Descricao'),
                            registro_id
                        ))
                        contador_atualizacoes += 1
                    else:
                        cursor.execute("""
                            INSERT INTO Combustivel_DePara 
                            (id, comb_cd, comb_ds, Combustivel_Codigo, Combustivel_Descricao)
                            VALUES (?, ?, ?, ?, ?)
                        """, (
                            registro_id,
                            codigo_origem,
                            registro.get('comb_ds'),
                            registro.get('Combustivel_Codigo'),
                            registro.get('Combustivel_Descricao')
                        ))
                        contador_insercoes += 1
            
            logger.info(f"UPDATEs executados: {contador_atualizacoes} registros")
            logger.info(f"INSERTs executados: {contador_insercoes} registros")
            
            logger.info("Executando COMMIT...")
            conexao.commit()
            logger.info("COMMIT executado com sucesso")
            
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                logger.info("Atualizando descrições com base no banco homólogo...")
                atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)
            
        except Exception as e:
            logger.error(f"Erro durante operações de banco: {str(e)}")
            conexao.rollback()
            logger.info("ROLLBACK executado")
            raise e
        
        cursor.execute("SELECT COUNT(*) FROM Combustivel_DePara")
        result_depois = cursor.fetchone()
        count_depois = safe_fetchone(cursor) if result_depois else 0
        logger.info(f"Registros na tabela DEPOIS da importação: {count_depois}")
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True, 
            'message': f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total na base: {count_depois} registros.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao importar arquivo DePara: {str(e)}")
        conexao.rollback()
        raise e

@combustivel_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline"""
    logger.info("=== UPDATE REGISTRO COMBUSTIVEL ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        logger.info(f"Dados recebidos: {data}")
        
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')
        
        if not record_id or not field:
            logger.error("ID ou campo não fornecidos")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})
        
        colunas_permitidas = ['Combustivel_Codigo', 'Combustivel_Descricao']
        if field not in colunas_permitidas:
            logger.error(f"Tentativa de editar campo não permitido: {field}")
            return jsonify({'success': False, 'message': f'Campo {field} não é permitido para edição'})
        
        if 'projeto_selecionado' not in session:
            logger.error("Nenhum projeto selecionado na sessão")
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            logger.error("Banco não configurado para este projeto")
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        banco_homo = obter_banco_homo(projeto_id)
        
        logger.info(f"Atualizando registro {record_id}, campo {field} para valor '{value}' no banco {banco_usuario}")
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error(f"Falha na conexão com o banco: {banco_usuario}")
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        try:
            cursor.execute(f"SELECT TOP 1 {field} FROM Combustivel_DePara WHERE id = ?", (record_id,))
            resultado = cursor.fetchone()
            if not resultado:
                logger.error(f"Registro não encontrado: {record_id}")
                return jsonify({'success': False, 'message': 'Registro não encontrado'})
        except Exception as e:
            logger.error(f"Erro ao verificar registro: {str(e)}")
            return jsonify({'success': False, 'message': f'Campo {field} não existe na tabela'})
        
        query = f"UPDATE Combustivel_DePara SET {field} = ? WHERE id = ?"
        logger.info(f"Executando query: {query} com valores: ({value}, {record_id})")
        
        cursor.execute(query, (value, record_id))
        
        if cursor.rowcount == 0:
            logger.warning(f"Nenhuma linha afetada pela atualização do registro {record_id}")
            if conexao:
                conexao.rollback()
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})
        
        conexao.commit()
        
        logger.info(f"Registro {record_id} atualizado com sucesso")
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
        
    except Exception as e:
        logger.error(f"Erro ao atualizar registro: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        if conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")
        
        return jsonify({'success': False, 'message': f'Erro ao atualizar registro: {str(e)}'})
    
    finally:
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            logger.error(f"Erro ao fechar cursor: {e}")
        
        try:
            if conexao:
                conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexão: {e}")

@combustivel_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        banco_homo = obter_banco_homo(projeto_id)

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()
        success_count = 0
        error_count = 0
        error_messages = []

        colunas_permitidas = ['Combustivel_Codigo', 'Combustivel_Descricao']

        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')

                if not record_id or not field:
                    error_count += 1
                    continue

                if field not in colunas_permitidas:
                    error_count += 1
                    continue

                query = f"UPDATE Combustivel_DePara SET {field} = ? WHERE id = ?"
                cursor.execute(query, (value, record_id))
                
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")

            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")

        conexao.commit()
        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")

        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }

        if error_messages:
            response['error_details'] = error_messages[:10]

        return jsonify(response)

    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}")
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})

    finally:
        if cursor:
            cursor.close()
        if conexao:
            conexao.close()

@combustivel_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})
        
        descricao = obter_descricao_wf(banco_homo, codigo)
        
        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })
            
    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})

def dados_combustivel(banco_usuario):
    if not banco_usuario:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    conexao = conectar_segunda_base(banco_usuario)
    if not conexao:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    try:
        cursor = conexao.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM Combustivel_DePara")
        qtd = safe_fetchone(cursor)

        cursor.execute(
            "SELECT COUNT(*) FROM Combustivel_DePara WHERE Combustivel_Codigo = 'S/DePara'"
        )
        qtdPendente = safe_fetchone(cursor)

        percentualConclusao = ((qtd - qtdPendente) / qtd * 100) if qtd > 0 else 0

        return {
            "qtd": qtd,
            "qtdPendente": qtdPendente,
            "percentualConclusao": round(percentualConclusao, 1),
        }
    except Exception as e:
        logger.error(f"Erro ao calcular dados Combustivel_DePara: {e}")
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}
    finally:
        if conexao:
            conexao.close()