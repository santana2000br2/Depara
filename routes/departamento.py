from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import tempfile
import os
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

departamento_bp = Blueprint("departamento", __name__)

def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados (mesma lógica usada em outros módulos)"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela departamento do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        # Seleciona o campo de codigo da tabela WF conforme informado
        cursor.execute("SELECT Departamento_Codigo FROM departamento")
        registros = cursor.fetchall()
        
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF (departamento)")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF (departamento): {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela departamento do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Departamento_Descricao FROM departamento WHERE Departamento_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo} (departamento): {str(e)}")
        return None

def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições do DePara após importação baseado nos códigos WF"""
    try:
        if not banco_homo:
            return
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições do departamento")
            return
        
        cursor = conexao.cursor()
        
        # Buscar registros que têm código WF (não são S/DePara)
        cursor.execute("""
            SELECT id, Departamento_Codigo, Departamento_Descricao 
            FROM Departamento_DePara 
            WHERE Departamento_Codigo IS NOT NULL AND Departamento_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()
        
        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)
            
            if descricao_wf and descricao_wf != descricao_atual:
                cursor.execute("""
                    UPDATE Departamento_DePara
                    SET Departamento_Descricao = ?
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")
        
        conexao.commit()
        cursor.close()
        conexao.close()
        
        logger.info(f"Atualizações automáticas de descrição (departamento): {atualizacoes} registros")
        
    except Exception as e:
        logger.error(f"Erro ao atualizar descrições de departamento após importação: {str(e)}")

@departamento_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão (igual aos outros processos)
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('departamento.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Departamento_DePara")
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros (departamento)")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('departamento.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em departamento.index: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('departamento.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@departamento_bp.route('/exportar')
def exportar_departamento():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('departamento.index'))
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('departamento.index'))
        
        # Executar consulta - Excluir coluna id da exportação
        cursor = conexao.cursor()
        cursor.execute("""
            SELECT dep_cd, dep_nm, dep_ativo, Departamento_Codigo, 
                   Departamento_Descricao, Departamento_Sigla, Origem 
            FROM Departamento_DePara
        """)
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]
        
        # Mapear nomes das colunas para nomes amigáveis
        mapeamento_colunas = {
            'dep_cd': 'Codigo de Origem',
            'dep_nm': 'Descrição de origem',
            'dep_ativo': 'Dep_Ativo',
            'Departamento_Codigo': 'Departamento_Codigo',
            'Departamento_Descricao': 'Departamento_Descricao',
            'Departamento_Sigla': 'Departamento_Sigla',
            'Origem': 'Origem'
        }
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        # Remover sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        ws = wb.create_sheet(title="Departamento_DePara")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna Departamento_Codigo
                # Identificar qual índice corresponde a Departamento_Codigo na query acima
                # A query selecionou: dep_cd(1), dep_nm(2), dep_ativo(3), Departamento_Codigo(4), ...
                if col_num == 4:  # Departamento_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Salvar em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "Departamento_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar departamento: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('departamento.index'))

@departamento_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_departamento_filtrados():
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        # Obter projeto e banco
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        banco_usuario = projeto_selecionado.get('DadosGX')
        
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        ws = wb.create_sheet(title="Departamento_Filtrado")
        
        # Cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                if isinstance(valor, str):
                    valor = valor.strip()
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Se a coluna é 'Departamento_Codigo' aplicar cores
                if header == 'Departamento_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="Departamento_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar departamento filtrado: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@departamento_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela departamento da base homóloga (WF)"""
    try:
        # Obter projeto
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('departamento.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo} (departamento)")
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('departamento.index'))
        
        cursor = conexao.cursor()
        # SELECT informado por você:
        cursor.execute("select Departamento_Codigo, Departamento_Descricao, TipoDepartamento_Codigo, Departamento_Contabil, Departamento_Ativo from departamento")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        registros_list = [dict(zip(colunas, row)) for row in registros]
        cursor.close()
        conexao.close()
        
        df = pd.DataFrame(registros_list, columns=colunas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Departamento_WF', index=False)
        output.seek(0)
        
        nome_arquivo = "Departamento_WF.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF (departamento): {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('departamento.index'))

@departamento_bp.route('/importar', methods=['POST'])
def importar_departamento():
    try:
        # Verificar se foi enviado um arquivo
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['file']
        
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Iniciando importação de departamento para o banco: {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Colunas esperadas
        colunas_importacao = [
            'dep_cd', 
            'dep_nm', 
            'dep_ativo',
            'Departamento_Codigo',
            'Departamento_Descricao',
            'Departamento_Sigla',
            'Origem'
        ]
        
        # Ler o Excel
        try:
            df = pd.read_excel(arquivo)
            df = df.where(pd.notnull(df), None)
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel (departamento): {str(e)}")
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {str(e)}'})
        
        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")
        
        # 🔄 Mapeamento de nomes amigáveis -> nomes do banco
        mapeamento_colunas_excel = {
            'Codigo de Origem': 'dep_cd',
            'Descrição de origem': 'dep_nm',
            'Dep_Ativo': 'dep_ativo',
            'Departamento_Codigo': 'Departamento_Codigo',
            'Departamento_Descricao': 'Departamento_Descricao',
            'Departamento_Sigla': 'Departamento_Sigla',
            'Origem': 'Origem'
        }

        # Normalizar nomes das colunas
        colunas_excel_normalizadas = []
        for col in colunas_excel:
            if col in mapeamento_colunas_excel:
                colunas_excel_normalizadas.append(mapeamento_colunas_excel[col])
            else:
                colunas_excel_normalizadas.append(col)

        df.columns = colunas_excel_normalizadas

        # Validar colunas necessárias
        colunas_excel_set = set(colunas_excel_normalizadas)
        colunas_necessarias = set(colunas_importacao)
        if not colunas_necessarias.issubset(colunas_excel_set):
            missing = colunas_necessarias - colunas_excel_set
            return jsonify({
                'success': False, 
                'message': f'Colunas necessárias faltando no arquivo: {", ".join(missing)}'
            })
        
        # Converter DataFrame novamente para dicionários normalizados
        registros = df.to_dict('records')
        
        # Filtrar e ajustar tamanhos
        registros_filtrados = []
        for registro in registros:
            registro_filtrado = {}
            for col in colunas_importacao:
                valor = registro.get(col)
                if valor is None:
                    registro_filtrado[col] = None
                else:
                    str_valor = str(valor)
                    if col in ['dep_cd', 'dep_ativo', 'Departamento_Codigo', 'Departamento_Sigla']:
                        registro_filtrado[col] = str_valor[:100]
                    elif col in ['dep_nm', 'Departamento_Descricao', 'Origem']:
                        registro_filtrado[col] = str_valor[:150]
                    else:
                        registro_filtrado[col] = str_valor
            registros_filtrados.append(registro_filtrado)
        
        logger.info(f"Registros filtrados e tratados: {len(registros_filtrados)}")
        
        # Contagem antes da importação
        cursor.execute("SELECT COUNT(*) FROM Departamento_DePara")
        result_antes = cursor.fetchone()
        count_antes = result_antes[0] if result_antes else 0
        logger.info(f"Registros antes da importação: {count_antes}")
        
        contador_atualizacoes = 0
        contador_insercoes = 0
        
        # Atualizar/Inserir registros
        for registro in registros_filtrados:
            codigo_origem = registro.get('dep_cd')
            novo_codigo_wf = registro.get('Departamento_Codigo')
            nova_descricao = registro.get('Departamento_Descricao')
            
            if codigo_origem:
                cursor.execute("SELECT id FROM Departamento_DePara WHERE dep_cd = ?", (codigo_origem,))
                resultado = cursor.fetchone()
                if resultado:
                    cursor.execute("""
                        UPDATE Departamento_DePara
                        SET dep_nm = ?, dep_ativo = ?, Departamento_Codigo = ?, 
                            Departamento_Descricao = ?, Departamento_Sigla = ?, Origem = ?
                        WHERE dep_cd = ?
                    """, (
                        registro.get('dep_nm'),
                        registro.get('dep_ativo'),
                        novo_codigo_wf,
                        nova_descricao,
                        registro.get('Departamento_Sigla'),
                        registro.get('Origem'),
                        codigo_origem
                    ))
                    contador_atualizacoes += 1
                else:
                    colunas_str = ', '.join(colunas_importacao)
                    placeholders = ', '.join(['?' for _ in colunas_importacao])
                    valores = [registro.get(c) for c in colunas_importacao]
                    cursor.execute(f"INSERT INTO Departamento_DePara ({colunas_str}) VALUES ({placeholders})", valores)
                    contador_insercoes += 1
            else:
                colunas_str = ', '.join(colunas_importacao)
                placeholders = ', '.join(['?' for _ in colunas_importacao])
                valores = [registro.get(c) for c in colunas_importacao]
                cursor.execute(f"INSERT INTO Departamento_DePara ({colunas_str}) VALUES ({placeholders})", valores)
                contador_insercoes += 1
        
        conexao.commit()
        logger.info(f"UPDATEs: {contador_atualizacoes}, INSERTs: {contador_insercoes}")
        
        # Atualizar descrições com base no banco homólogo
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            logger.info("Atualizando descrições de departamento com base no banco homólogo...")
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)
        
        # Contagem depois
        cursor.execute("SELECT COUNT(*) FROM Departamento_DePara")
        result_depois = cursor.fetchone()
        count_depois = result_depois[0] if result_depois else 0
        logger.info(f"Registros depois da importação: {count_depois}")
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True,
            'message': f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total: {count_depois}.'
        })
    
    except Exception as e:
        logger.error(f"Erro ao importar departamento: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        if 'conexao' in locals() and conexao is not None:
            try:
                conexao.rollback()
                conexao.close()
            except:
                pass
        
        return jsonify({'success': False, 'message': f'Erro na importação: {str(e)}'})


@departamento_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline (aceita Departamento_Codigo e Departamento_Descricao)"""
    logger.info("=== UPDATE REGISTRO DEPARTAMENTO ENDPOINT ACESSADO ===")
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        logger.info(f"Dados recebidos (update departamento): {data}")
        
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')
        
        if not record_id or not field:
            logger.error("ID ou campo não fornecidos (update departamento)")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'}), 400
        
        # Colunas permitidas
        colunas_permitidas = ['Departamento_Codigo', 'Departamento_Descricao']
        if field not in colunas_permitidas:
            logger.error(f"Tentativa de editar campo não permitido: {field}")
            return jsonify({'success': False, 'message': 'Campo não permitido para edição'}), 400
        
        # Conectar
        projeto_selecionado = session.get('projeto_selecionado')
        if not projeto_selecionado:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'}), 500
        
        cursor = conexao.cursor()
        
        # Validações específicas
        if field == 'Departamento_Codigo':
            # permitir 'S/DePara' ou números
            if value != 'S/DePara' and value is not None and not str(value).isdigit():
                return jsonify({'success': False, 'message': 'Departamento_Codigo deve ser somente números ou "S/DePara"'}), 400
        
        # Fazer update
        cursor.execute(f"UPDATE Departamento_DePara SET {field} = ? WHERE id = ?", (value, record_id))
        conexao.commit()
        logger.info(f"Registro id={record_id} atualizado: {field} = {value}")
        
        # Se atualizou o código, tentar atualizar descrição a partir do banco homólogo
        if field == 'Departamento_Codigo':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo and value and value != 'S/DePara':
                descricao_wf = obter_descricao_wf(banco_homo, value)
                if descricao_wf:
                    cursor.execute("UPDATE Departamento_DePara SET Departamento_Descricao = ? WHERE id = ?", (descricao_wf, record_id))
                    conexao.commit()
                    logger.info(f"Descrição atualizada automaticamente para id={record_id} com valor '{descricao_wf}'")
        
        # Fechar
        cursor.close()
        conexao.close()
        
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
        
    except Exception as e:
        logger.error(f"Erro no endpoint update (departamento): {str(e)}")

        # Rollback seguro
        if conexao is not None:
            try:
               conexao.rollback()
            except Exception as ex:
                logger.warning(f"Falha ao executar rollback: {str(ex)}")

         # Fechar cursor com segurança
        if cursor is not None:
            try:
              cursor.close()
            except Exception as ex:
              logger.warning(f"Falha ao fechar cursor: {str(ex)}")

    # Fechar conexão com segurança
        if conexao is not None:
            try:
                conexao.close()
            except Exception as ex:
                logger.warning(f"Falha ao fechar conexão: {str(ex)}")

    
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {str(e)}'}), 500
