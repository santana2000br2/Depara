# municipio.py
from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

municipio_bp = Blueprint("municipio", __name__)

# -------------------------------------------------------
# Opcional: mapear variações de nomes vindas do front-end
# (mantém compatibilidade caso o HTML use 'Municipio_Descricao')
# -------------------------------------------------------
FIELD_MAP = {
    'Municipio_Codigo': 'Municipio_Codigo',
    'municipio_codigo': 'Municipio_Codigo',
    'Municipio_Nome': 'Municipio_Nome',
    'municipio_nome': 'Municipio_Nome',
    'Municipio_Descricao': 'Municipio_Nome',   # variação aceita
    'Estado_Codigo': 'Estado_Codigo',
    'uf_cd': 'uf_cd',
}
ALLOWED_FIELDS = {'Municipio_Codigo', 'Municipio_Nome', 'Estado_Codigo', 'uf_cd'}


# -------------------------------------------------------
# Helpers
# -------------------------------------------------------
def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None

        cursor = conn.cursor()
        # usar tupla para binding
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", (projeto_id,))
        resultado = cursor.fetchone()

        cursor.close()
        conn.close()

        if resultado and resultado[0]:
            return resultado[0]
        return None

    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None


def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela Municipio do banco homólogo"""
    try:
        if not banco_homo:
            return []

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []

        cursor = conexao.cursor()
        cursor.execute("SELECT Municipio_Codigo FROM Municipio")
        registros = cursor.fetchall()

        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]

        cursor.close()
        conexao.close()

        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos

    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []


def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição (Municipio_Nome) da base de homologação (Municipio)."""
    try:
        if not banco_homo or not codigo:
            return None

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None

        cursor = conexao.cursor()
        cursor.execute("SELECT Municipio_Nome FROM Municipio WHERE Municipio_Codigo = ?", (codigo,))
        resultado = cursor.fetchone()

        cursor.close()
        conexao.close()

        if resultado and resultado[0]:
            return resultado[0]
        return None
    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF para {codigo}: {e}")
        return None


def atualizar_descricoes_automaticamente(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições na tabela Municipio_DePara baseado nos códigos WF"""
    try:
        if not banco_homo:
            return

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições")
            return

        cursor = conexao.cursor()

        # Buscar registros que têm código WF (não são S/DePara)
        cursor.execute("""
            SELECT id, Municipio_Codigo, Municipio_Nome
            FROM Municipio_DePara
            WHERE Municipio_Codigo IS NOT NULL AND Municipio_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()

        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            if not codigo_wf:
                continue
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)

            if descricao_wf and descricao_wf != descricao_atual:
                # Atualizar descrição no DePara
                cursor.execute("""
                    UPDATE Municipio_DePara
                    SET Municipio_Nome = ?
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")

        conexao.commit()
        cursor.close()
        conexao.close()

        logger.info(f"Atualizações automáticas de descrição: {atualizacoes} registros")

    except Exception as e:
        logger.error(f"Erro ao atualizar descrições automaticamente: {str(e)}")


# -------------------------------------------------------
# Rotas
# -------------------------------------------------------
@municipio_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('municipio.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Municipio_DePara")
        registros = cursor.fetchall()

        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]

        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]

        logger.info(f"Encontrados {len(registros_dict)} registros")

        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # Fechar recursos
        cursor.close()
        conexao.close()

        return render_template('municipio.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf,
                               banco_homo=banco_homo)

    except Exception as e:
        logger.error(f"Erro em municipio: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('municipio.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])


@municipio_bp.route('/exportar')
def exportar_municipio():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('municipio.index'))

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('municipio.index'))

        # Executar consulta - Excluir coluna id da exportação
        cursor = conexao.cursor()
        cursor.execute("SELECT cg_cidade, uf_cd, Estado_Codigo, Municipio_Codigo, Municipio_Nome, Municipio_IBGE FROM Municipio_DePara")
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]

        # Mapear nomes das colunas para os nomes amigáveis
        mapeamento_colunas = {
            'cg_cidade': 'Codigo Anterior',
            'uf_cd': 'UF',
            'Estado_Codigo': 'Estado_Codigo',
            'Municipio_Codigo': 'Municipio_Codigo',
            'Municipio_Nome': 'Municipio_Descricao',
            'Municipio_IBGE': 'Municipio_IBGE'
        }

        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]

        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # Criar workbook
        wb = Workbook()

        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)

        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="Municipio_DePara")

        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios

        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()

                cell = ws.cell(row=row_num, column=col_num, value=valor)

                # Aplicar cores na coluna Municipio_Codigo (coluna 4)
                if col_num == 4:  # Municipio_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except Exception:
                continue
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width

        # Fechar recursos
        cursor.close()
        conexao.close()

        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = "Municipio_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar municipio: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('municipio.index'))

@municipio_bp.route('/exportar_wf')
def exportar_wf():
    """Exporta a tabela Municipio diretamente da base de homologação (BancoHomo)."""
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        projeto_id = projeto.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco de homologação não configurado.', 'error')
            return redirect(url_for('municipio.index'))

        logger.info(f"Exportando Tabela WF da base: {banco_homo}")

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_homo}', 'error')
            return redirect(url_for('municipio.index'))

        cursor = conexao.cursor()
        cursor.execute("""
            SELECT Municipio_Codigo, Municipio_Nome, Municipio_IBGE, Municipio_Ativo
            FROM Municipio
            ORDER BY Municipio_Nome
        """)
        registros = cursor.fetchall()
        descricao_colunas = cursor.description or []
        colunas = [col[0] for col in descricao_colunas]

        # Criar Excel
        wb = Workbook()
        ws = wb.active
        # garantir que ws não seja None (evita aviso do Pylance)
        if ws is None:
            ws = wb.create_sheet(title="Municipio_WF")
        else:
            # atribuir título de forma segura
            try:
                ws.title = "Municipio_WF"
            except Exception:
                # fallback — criar uma nova sheet com título se não der para ajustar o active
                ws = wb.create_sheet(title="Municipio_WF")

        # Cabeçalhos
        for col_num, coluna in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Dados
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                ws.cell(row=row_num, column=col_num, value=valor)

        # Ajustar largura das colunas de forma segura
        for column in ws.columns:
            # coluna pode ser uma sequência vazia, então proteger
            if not column or column[0] is None:
                continue
            try:
                column_letter = getattr(column[0], "column_letter", None)
                if not column_letter:
                    continue
            except Exception:
                continue

            max_length = 0
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            try:
                ws.column_dimensions[column_letter].width = adjusted_width
            except Exception:
                # se por algum motivo não for possível ajustar, ignorar
                pass

        cursor.close()
        conexao.close()

        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Municipio_WF.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {e}", exc_info=True)
        flash(f"Erro ao exportar tabela WF: {e}", "error")
        return redirect(url_for('municipio.index'))

@municipio_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_municipio_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])

        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400

        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        # Criar workbook
        wb = Workbook()

        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)

        # Criar uma nova worksheet
        ws = wb.create_sheet(title="Municipio_Filtrado")

        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios

        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()

                cell = ws.cell(row=row_num, column=col_num, value=valor)

                # Aplicar cores na coluna Municipio_Codigo
                if header == 'Municipio_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except Exception:
                continue
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Municipio_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar municipio filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500


@municipio_bp.route('/importar', methods=['POST'])
def importar_municipio():
    try:
        # Verificar se foi enviado um arquivo
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']

        # Verificar se filename existe e não está vazio
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        # Verificação mais segura da extensão
        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})

        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        logger.info(f"Iniciando importação para o banco: {banco_usuario}")

        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()

        # Ler arquivo garantindo que todas as colunas sejam tratadas como string
        try:
            df = pd.read_excel(arquivo, dtype=str)

            # Substituir NaN, NaT e valores nulos por None
            df = df.replace({pd.NaT: None, 'nan': None, 'NaN': None, '': None})
            df = df.where(pd.notnull(df), None)

            # Converter para lista de dicionários
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()

            logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")

        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel: {str(e)}")
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {str(e)}'})

        # MAPEAMENTO: Nomes das colunas na planilha para nomes das colunas no banco
        mapeamento_colunas = {
            'Codigo Anterior': 'cg_cidade',
            'UF': 'uf_cd',
            'Estado_Codigo': 'Estado_Codigo',
            'Municipio_Codigo': 'Municipio_Codigo',
            'Municipio_Descricao': 'Municipio_Nome',  # aceitar variação
            'cg_cidade': 'cg_cidade',
            'uf_cd': 'uf_cd',
            'Municipio_Nome': 'Municipio_Nome'
        }

        # Colunas do banco que precisamos
        colunas_banco = ['cg_cidade', 'uf_cd', 'Estado_Codigo', 'Municipio_Codigo', 'Municipio_Nome']

        # Verificar se todas as colunas necessárias estão presentes (usando nomes amigáveis)
        colunas_faltantes = []
        colunas_planilha_esperadas = ['Codigo Anterior', 'UF', 'Estado_Codigo', 'Municipio_Codigo', 'Municipio Nome']

        for coluna_planilha in colunas_planilha_esperadas:
            if coluna_planilha not in colunas_excel:
                colunas_faltantes.append(coluna_planilha)

        if colunas_faltantes:
            return jsonify({
                'success': False,
                'message': f'Colunas necessárias faltando no arquivo: {", ".join(colunas_faltantes)}. Certifique-se de que a planilha contém as colunas: Codigo Anterior, UF, Estado_Codigo, Municipio_Codigo, Municipio_Descricao'
            })

        # Filtrar e mapear os registros com tratamento robusto de tipos
        registros_mapeados = []
        for registro in registros:
            registro_mapeado = {}
            for coluna_banco in colunas_banco:
                coluna_planilha = None
                for chave, valor in mapeamento_colunas.items():
                    if valor == coluna_banco and chave in colunas_excel:
                        coluna_planilha = chave
                        break

                if coluna_planilha:
                    valor = registro.get(coluna_planilha)
                    if valor is not None:
                        if not isinstance(valor, str):
                            valor = str(valor)
                        valor = valor.strip()
                        if valor == '':
                            valor = None
                        elif coluna_banco == 'Municipio_Codigo' and valor.upper() == 'S/DEPARA':
                            valor = 'S/DePara'
                    registro_mapeado[coluna_banco] = valor
                else:
                    registro_mapeado[coluna_banco] = None

            # Validar registro antes de adicionar
            if registro_mapeado.get('cg_cidade'):
                registros_mapeados.append(registro_mapeado)

        logger.info(f"Registros mapeados após filtro: {len(registros_mapeados)}")

        # VERIFICAÇÃO ANTES: Contar registros antes da importação
        cursor.execute("SELECT COUNT(*) FROM Municipio_DePara")
        result_antes = cursor.fetchone()
        count_antes = result_antes[0] if result_antes else 0
        logger.info(f"Registros na tabela ANTES da importação: {count_antes}")

        # Atualizar/Inserir registros
        try:
            contador_atualizacoes = 0
            contador_insercoes = 0
            erros_importacao = []

            for i, registro in enumerate(registros_mapeados):
                try:
                    cg_cidade = registro.get('cg_cidade')
                    novo_codigo_wf = registro.get('Municipio_Codigo')
                    nova_descricao_wf = registro.get('Municipio_Nome')
                    uf_cd = registro.get('uf_cd')
                    estado_codigo = registro.get('Estado_Codigo')

                    if cg_cidade:
                        # Verificar se o registro já existe
                        cursor.execute("SELECT id FROM Municipio_DePara WHERE cg_cidade = ?", (cg_cidade,))
                        resultado = cursor.fetchone()

                        if resultado:
                            # UPDATE do registro existente
                            cursor.execute("""
                                UPDATE Municipio_DePara
                                SET Municipio_Codigo = ?, Municipio_Nome = ?, uf_cd = ?, Estado_Codigo = ?
                                WHERE cg_cidade = ?
                            """, (novo_codigo_wf, nova_descricao_wf, uf_cd, estado_codigo, cg_cidade))
                            contador_atualizacoes += 1
                        else:
                            # INSERT apenas se for um registro novo
                            cursor.execute("""
                                INSERT INTO Municipio_DePara
                                (cg_cidade, uf_cd, Estado_Codigo, Municipio_Codigo, Municipio_Nome)
                                VALUES (?, ?, ?, ?, ?)
                            """, (
                                cg_cidade,
                                uf_cd,
                                estado_codigo,
                                novo_codigo_wf,
                                nova_descricao_wf
                            ))
                            contador_insercoes += 1

                    # COMMIT a cada 100 registros para evitar transações muito longas
                    if i % 100 == 0:
                        conexao.commit()

                except Exception as e:
                    erros_importacao.append(f"Registro {i+1} (Código: {registro.get('cg_cidade')}): {str(e)}")
                    logger.error(f"Erro no registro {i+1}: {str(e)}")

            # COMMIT final
            conexao.commit()

            logger.info(f"UPDATEs executados: {contador_atualizacoes} registros")
            logger.info(f"INSERTs executados: {contador_insercoes} registros")

            if erros_importacao:
                logger.warning(f"Erros durante importação: {len(erros_importacao)} registros com problema")

            # ATUALIZAR DESCRIÇÕES COM BASE NO BANCO HOMÓLOGO
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                logger.info("Atualizando descrições com base no banco homólogo...")
                atualizar_descricoes_automaticamente(banco_usuario, banco_homo)

        except Exception as e:
            logger.error(f"Erro durante operações de banco: {str(e)}")
            try:
                conexao.rollback()
            except:
                pass
            logger.info("ROLLBACK executado")
            raise e

        # VERIFICAÇÃO DEPOIS: Contar registros após a importação
        cursor.execute("SELECT COUNT(*) FROM Municipio_DePara")
        result_depois = cursor.fetchone()
        count_depois = result_depois[0] if result_depois else 0
        logger.info(f"Registros na tabela DEPOIS da importação: {count_depois}")

        cursor.close()
        conexao.close()

        # Mensagem de sucesso com detalhes
        mensagem = f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total na base: {count_depois} registros.'

        if erros_importacao:
            mensagem += f' {len(erros_importacao)} registros tiveram erro.'
            for erro in erros_importacao[:5]:
                logger.warning(f"Erro de importação: {erro}")

        return jsonify({
            'success': True,
            'message': mensagem
        })

    except Exception as e:
        logger.error(f"Erro ao importar municipio: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'Erro na importação: {str(e)}'
        })


@municipio_bp.route('/update', methods=['POST'])
def update_registro():
    logger.info("=== UPDATE REGISTRO MUNICIPIO ENDPOINT ACESSADO ===")
    conexao = None; cursor = None
    try:
        data = request.get_json()
        record_id = data.get('id')
        raw_field = data.get('field')
        value = data.get('value')

        if not record_id or not raw_field:
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})

        # normalizar campo e validar
        field = FIELD_MAP.get(raw_field, raw_field)
        if field not in ALLOWED_FIELDS:
            return jsonify({'success': False, 'message': f'Campo não permitido: {raw_field}'})

        # checar sessão/projeto
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        # conectar com checagem explícita
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # executar update seguro (campo validado)
        cursor.execute(f"UPDATE Municipio_DePara SET {field} = ? WHERE id = ?", (value, record_id))

        # se foi o código, sincronizar descrição no banco homólogo
        if field == 'Municipio_Codigo':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo and value and value != 'S/DePara':
                nova_descricao = obter_descricao_wf(banco_homo, value)
                if nova_descricao:
                    cursor.execute("UPDATE Municipio_DePara SET Municipio_Nome = ? WHERE id = ?", (nova_descricao, record_id))
                    logger.info(f"Descricao atualizada automaticamente para id {record_id}: {nova_descricao}")

        if cursor.rowcount == 0:
            try:
                conexao.rollback()
            except:
                pass
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})

        conexao.commit()
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})

    except Exception as e:
        logger.error("Erro ao atualizar registro", exc_info=True)
        if conexao:
            try: conexao.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)})
    finally:
        try:
            if cursor: cursor.close()
            if conexao: conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar recursos: {e}")


@municipio_bp.route('/update_batch', methods=['POST'])
def update_batch():
    logger.info("=== UPDATE BATCH MUNICIPIO ENDPOINT ACESSADO ===")
    conexao = None; cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        banco_homo = obter_banco_homo(projeto_id)
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        success_count = 0; error_count = 0; error_messages = []

        for up in updates:
            try:
                record_id = up.get('id'); raw_field = up.get('field'); value = up.get('value')
                if not record_id or not raw_field:
                    error_count += 1; error_messages.append(f"ID ou campo ausente: {up}"); continue

                field = FIELD_MAP.get(raw_field, raw_field)
                if field not in ALLOWED_FIELDS:
                    error_count += 1; error_messages.append(f"Campo não permitido: {raw_field}"); continue

                cursor.execute(f"UPDATE Municipio_DePara SET {field} = ? WHERE id = ?", (value, record_id))

                # sincronizar descrição se for código
                if field == 'Municipio_Codigo' and banco_homo and value and value != 'S/DePara':
                    nova_descricao = obter_descricao_wf(banco_homo, value)
                    if nova_descricao:
                        cursor.execute("UPDATE Municipio_DePara SET Municipio_Nome = ? WHERE id = ?", (nova_descricao, record_id))
                        logger.info(f"Descricao atualizada automaticamente para id {record_id}: {nova_descricao}")

                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")

            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")
                logger.error(f"Erro ao processar update {up}: {e}", exc_info=True)

        conexao.commit()
        return jsonify({'success': True, 'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros', 'success_count': success_count, 'error_count': error_count, 'error_details': error_messages[:10]})
    except Exception as e:
        logger.error("Erro no batch update", exc_info=True)
        if conexao:
            try: conexao.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)})
    finally:
        try:
            if cursor: cursor.close()
            if conexao: conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar recursos: {e}")


@municipio_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})

        descricao = obter_descricao_wf(banco_homo, codigo)

        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })

    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})
