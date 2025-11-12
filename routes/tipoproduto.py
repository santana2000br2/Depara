from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import tempfile
import os

tipoproduto_bp = Blueprint("tipoproduto", __name__)

def safe_convert_id(value):
    """Converte valores para inteiro de forma segura"""
    if value is None:
        return None
    try:
        if isinstance(value, float):
            return int(value)
        elif isinstance(value, str):
            # Remove espaços e tenta converter
            cleaned = value.strip()
            if cleaned == '':
                return None
            # Converte string para float primeiro, depois para int
            return int(float(cleaned))
        elif isinstance(value, int):
            return value
        else:
            return int(value)
    except (ValueError, TypeError):
        raise ValueError(f"Valor não pode ser convertido para inteiro: {value}")
    
def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela TipoProduto do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoProduto_Codigo FROM TipoProduto")
        registros = cursor.fetchall()
        
        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela TipoProduto do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoProduto_Descricao FROM TipoProduto WHERE TipoProduto_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None

def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições após importação baseado nos códigos WF"""
    try:
        if not banco_homo:
            return
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições")
            return
        
        cursor = conexao.cursor()
        
        # Buscar registros que têm código WF (não são S/DePara)
        cursor.execute("""
            SELECT id, TipoProduto_Codigo, TipoProduto_Descricao 
            FROM TipoProduto_DePara 
            WHERE TipoProduto_Codigo IS NOT NULL AND TipoProduto_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()
        
        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)
            
            if descricao_wf and descricao_wf != descricao_atual:
                # Atualizar descrição
                cursor.execute("""
                    UPDATE TipoProduto_DePara 
                    SET TipoProduto_Descricao = ? 
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")
        
        conexao.commit()
        cursor.close()
        conexao.close()
        
        logger.info(f"Atualizações automáticas de descrição: {atualizacoes} registros")
        
    except Exception as e:
        logger.error(f"Erro ao atualizar descrições após importação: {str(e)}")

def safe_fetchone(cursor):
    """Função segura para fetchone que evita None is not subscriptable"""
    try:
        result = cursor.fetchone()
        if result is not None and result[0] is not None:
            return result[0]
        return 0
    except Exception as e:
        logger.error(f"Erro em safe_fetchone: {str(e)}")
        return 0

@tipoproduto_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('tipoproduto.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM TipoProduto_DePara")
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('tipoproduto.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em tipoproduto: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('tipoproduto.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@tipoproduto_bp.route('/exportar')
def exportar_tipoproduto():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('tipoproduto.index'))
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('tipoproduto.index'))
        
        # Executar consulta - INCLUIR ID como chave primária
        # Rever trazer todas colunas do banco
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM TipoProduto_DePara")
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]
        
        # Mapear nomes das colunas para os nomes amigáveis
        mapeamento_colunas = {
            'id': 'ID',
            'tpd_cd': 'Codigo de Origem',
            'tpd_ds': 'Descrição de origem',
            'TipoProduto_Codigo': 'TipoProduto_Codigo',
            'TipoProduto_Descricao': 'TipoProduto_Descricao',
            'TipoProduto_GrupoContabilCod': 'TipoProduto_GrupoContabilCod',
            'tpd_grupocontab': 'tpd_grupocontab'
        }
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook e worksheet
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="TipoProduto_DePara")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna TipoProduto_Codigo (coluna 4)
                if col_num == 4:  # TipoProduto_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Salvar para buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "TipoProduto_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tipoproduto: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('tipoproduto.index'))

@tipoproduto_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_tipoproduto_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet
        ws = wb.create_sheet(title="TipoProduto_Filtrado")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna TipoProduto_Codigo
                if header == 'TipoProduto_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar para buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="TipoProduto_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tipoproduto filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@tipoproduto_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela TipoProduto do banco homólogo (BancoHomo)"""
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter o BancoHomo diretamente do banco de dados
        banco_homo = obter_banco_homo(projeto_id)
        
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('tipoproduto.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")
        
        # Conectar ao banco homólogo
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('tipoproduto.index'))
        
        # Executar consulta na tabela TipoProduto do banco homólogo
        cursor = conexao.cursor()
        cursor.execute("SELECT TipoProduto_Codigo, TipoProduto_Descricao, TipoProduto_GrupoContabilCod, TipoProduto_Ativo FROM TipoProduto")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        # Converter para lista de dicionários
        registros_list = [dict(zip(colunas, row)) for row in registros]
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Criar DataFrame pandas
        df = pd.DataFrame(registros_list, columns=colunas)
        
        # Criar buffer para o arquivo Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TipoProduto_WF', index=False)
        
        output.seek(0)
        
        # Nome do arquivo
        nome_arquivo = "TipoProduto_WF.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('tipoproduto.index'))

@tipoproduto_bp.route('/importar', methods=['POST'])
def importar_tipoproduto():
    try:
        # Verificar se foi enviado um arquivo
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        arquivo = request.files['file']
        
        # Verificar se filename existe e não está vazio
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})
        
        # Verificação mais segura da extensão
        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Iniciando importação para o banco: {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Importar dados do Excel
        try:
            # Forçar tipos específicos para evitar problemas de conversão
            dtype = {
                'ID': 'Int64',
                'Codigo de Origem': 'string',
                'Descrição de origem': 'string',
                'TipoProduto_Codigo': 'string',
                'TipoProduto_Descricao': 'string',
                'TipoProduto_GrupoContabilCod': 'string',
                'tpd_grupocontab': 'string'
            }
            
            # Ler o Excel com tipos específicos
            df = pd.read_excel(arquivo, dtype=dtype, na_values=['', ' ', 'NULL', 'null'])
            
            # Substituir NaN por None
            df = df.where(pd.notnull(df), None)
            
            # Converter coluna ID explicitamente
            if 'ID' in df.columns:
                df['ID'] = pd.to_numeric(df['ID'], errors='coerce').astype('Int64')
                
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel: {str(e)}")
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {str(e)}'})
        
        # Converter para lista de dicionários
        registros = df.to_dict('records')
        colunas_excel = df.columns.tolist()
        
        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")
        
        # MAPEAMENTO: Nomes das colunas na planilha para nomes das colunas no banco
        mapeamento_colunas = {
            'ID': 'id',
            'Codigo de Origem': 'tpd_cd',
            'Descrição de origem': 'tpd_ds',
            'TipoProduto_Codigo': 'TipoProduto_Codigo',
            'TipoProduto_Descricao': 'TipoProduto_Descricao',
            'TipoProduto_GrupoContabilCod': 'TipoProduto_GrupoContabilCod',
            'tpd_grupocontab': 'tpd_grupocontab',
            # Incluir também os nomes originais caso a planilha os use
            'id': 'id',
            'tpd_cd': 'tpd_cd',
            'tpd_ds': 'tpd_ds'
        }
        
        # Colunas do banco que precisamos (incluindo ID como chave)
        colunas_banco = ['id', 'tpd_cd', 'tpd_ds', 'TipoProduto_Codigo', 'TipoProduto_Descricao', 'TipoProduto_GrupoContabilCod', 'tpd_grupocontab']
        
        # Verificar se todas as colunas necessárias estão presentes (usando o mapeamento)
        colunas_faltantes = []
        for coluna_banco in colunas_banco:
            # Encontrar o nome correspondente na planilha
            coluna_planilha = None
            for chave, valor in mapeamento_colunas.items():
                if valor == coluna_banco and chave in colunas_excel:
                    coluna_planilha = chave
                    break
            
            if not coluna_planilha:
                colunas_faltantes.append(coluna_banco)
        
        if colunas_faltantes:
            return jsonify({
                'success': False, 
                'message': f'Colunas necessárias faltando no arquivo: {", ".join(colunas_faltantes)}. Certifique-se de que a planilha contém as colunas: ID, Codigo de Origem, Descrição de origem, TipoProduto_Codigo, TipoProduto_Descricao, TipoProduto_GrupoContabilCod, tpd_grupocontab'
            })
        
        # Filtrar e mapear os registros
        registros_mapeados = []
        for registro in registros:
            registro_mapeado = {}
            
            # Para cada coluna do banco, buscar o valor correspondente na planilha
            for coluna_banco in colunas_banco:
                # Encontrar o nome da coluna na planilha
                coluna_planilha = None
                for chave, valor in mapeamento_colunas.items():
                    if valor == coluna_banco and chave in colunas_excel:
                        coluna_planilha = chave
                        break
                
                if coluna_planilha:
                    valor = registro.get(coluna_planilha)
                    
                    # Conversão segura de tipos
                    if coluna_banco == 'id':  # ID deve ser inteiro
                        try:
                            valor = safe_convert_id(valor)
                        except ValueError as e:
                            return jsonify({
                                'success': False, 
                                'message': f'Erro de conversão do ID: {str(e)}'
                            })
                    elif valor is not None and isinstance(valor, str):
                        valor = valor.strip()
                        # Se for string vazia após strip, converter para None
                        if valor == '':
                            valor = None
                    
                    registro_mapeado[coluna_banco] = valor
                else:
                    registro_mapeado[coluna_banco] = None
            
            registros_mapeados.append(registro_mapeado)
        
        logger.info(f"Registros mapeados: {len(registros_mapeados)}")
        
        # VALIDAÇÃO: Verificar se todos os IDs são válidos
        for registro in registros_mapeados:
            if registro.get('id') is None:
                return jsonify({
                    'success': False, 
                    'message': 'Encontrado registro sem ID. Todos os registros devem ter um ID inteiro válido.'
                })
        
        # VERIFICAÇÃO ANTES: Contar registros antes da importação
        cursor.execute("SELECT COUNT(*) FROM TipoProduto_DePara")
        result_antes = cursor.fetchone()
        count_antes = safe_fetchone(cursor) if result_antes else 0
        logger.info(f"Registros na tabela ANTES da importação: {count_antes}")
        
        # FAZER UPDATE EM VEZ DE DELETE + INSERT
        try:
            contador_atualizacoes = 0
            contador_insercoes = 0
            
            for registro in registros_mapeados:
                registro_id = registro.get('id')
                codigo_origem = registro.get('tpd_cd')
                
                if registro_id and codigo_origem:
                    # Verificar se o registro já existe
                    cursor.execute("SELECT id FROM TipoProduto_DePara WHERE id = ?", (registro_id,))
                    resultado = cursor.fetchone()
                    
                    if resultado:
                        # UPDATE do registro existente
                        cursor.execute("""
                            UPDATE TipoProduto_DePara 
                            SET tpd_cd = ?, 
                                tpd_ds = ?, 
                                TipoProduto_Codigo = ?, 
                                TipoProduto_Descricao = ?,
                                TipoProduto_GrupoContabilCod = ?,
                                tpd_grupocontab = ?
                            WHERE id = ?
                        """, (
                            codigo_origem,
                            registro.get('tpd_ds'),
                            registro.get('TipoProduto_Codigo'),
                            registro.get('TipoProduto_Descricao'),
                            registro.get('TipoProduto_GrupoContabilCod'),
                            registro.get('tpd_grupocontab'),
                            registro_id
                        ))
                        contador_atualizacoes += 1
                    else:
                        # INSERT apenas se for um registro novo (raro caso)
                        cursor.execute("""
                            INSERT INTO TipoProduto_DePara 
                            (id, tpd_cd, tpd_ds, TipoProduto_Codigo, TipoProduto_Descricao, TipoProduto_GrupoContabilCod, tpd_grupocontab)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (
                            registro_id,
                            codigo_origem,
                            registro.get('tpd_ds'),
                            registro.get('TipoProduto_Codigo'),
                            registro.get('TipoProduto_Descricao'),
                            registro.get('TipoProduto_GrupoContabilCod'),
                            registro.get('tpd_grupocontab')
                        ))
                        contador_insercoes += 1
            
            logger.info(f"UPDATEs executados: {contador_atualizacoes} registros")
            logger.info(f"INSERTs executados: {contador_insercoes} registros")
            
            # COMMIT
            logger.info("Executando COMMIT...")
            conexao.commit()
            logger.info("COMMIT executado com sucesso")
            
            # ATUALIZAR DESCRIÇÕES COM BASE NO BANCO HOMÓLOGO
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                logger.info("Atualizando descrições com base no banco homólogo...")
                atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)
            
        except Exception as e:
            logger.error(f"Erro durante operações de banco: {str(e)}")
            conexao.rollback()
            logger.info("ROLLBACK executado")
            raise e
        
        # VERIFICAÇÃO DEPOIS: Contar registros após a importação
        cursor.execute("SELECT COUNT(*) FROM TipoProduto_DePara")
        result_depois = cursor.fetchone()
        count_depois = safe_fetchone(cursor) if result_depois else 0
        logger.info(f"Registros na tabela DEPOIS da importação: {count_depois}")
        
        cursor.close()
        conexao.close()
        
        return jsonify({
            'success': True, 
            'message': f'Importação concluída! {contador_atualizacoes} registros atualizados, {contador_insercoes} novos registros inseridos. Total na base: {count_depois} registros.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao importar tipoproduto: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False, 
            'message': f'Erro na importação: {str(e)}'
        })

@tipoproduto_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline"""
    logger.info("=== UPDATE REGISTRO TIPOPRODUTO ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados recebidos: {data}")
        
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')
        
        if not record_id or not field:
            logger.error("ID ou campo não fornecidos")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})
        
        # VERIFICAR SE O CAMPO É PERMITIDO
        colunas_permitidas = ['TipoProduto_Codigo', 'TipoProduto_GrupoContabilCod', 'tpd_grupocontab']
        if field not in colunas_permitidas:
            logger.error(f"Tentativa de editar campo não permitido: {field}")
            return jsonify({'success': False, 'message': f'Campo {field} não é permitido para edição'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            logger.error("Nenhum projeto selecionado na sessão")
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            logger.error("Banco não configurado para este projeto")
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        # Obter banco homólogo para possível atualização de descrição
        banco_homo = obter_banco_homo(projeto_id)
        
        logger.info(f"Atualizando registro {record_id}, campo {field} para valor '{value}' no banco {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error(f"Falha na conexão com o banco: {banco_usuario}")
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Verificar se a tabela existe e tem a coluna
        try:
            cursor.execute(f"SELECT TOP 1 {field} FROM TipoProduto_DePara WHERE id = ?", (record_id,))
            resultado = cursor.fetchone()
            if not resultado:
                logger.error(f"Registro não encontrado: {record_id}")
                return jsonify({'success': False, 'message': 'Registro não encontrado'})
        except Exception as e:
            logger.error(f"Erro ao verificar registro: {str(e)}")
            return jsonify({'success': False, 'message': f'Campo {field} não existe na tabela'})
        
        # Atualizar registro - usando id como chave primária
        query = f"UPDATE TipoProduto_DePara SET {field} = ? WHERE id = ?"
        logger.info(f"Executando query: {query} com valores: ({value}, {record_id})")
        
        cursor.execute(query, (value, record_id))
        
        # Se o campo alterado for TipoProduto_Codigo, buscar a descrição automaticamente
        if field == 'TipoProduto_Codigo' and banco_homo:
            nova_descricao = None
            if value and value != 'S/DePara':
                nova_descricao = obter_descricao_wf(banco_homo, value)
            
            if nova_descricao:
                cursor.execute("""
                    UPDATE TipoProduto_DePara
                    SET TipoProduto_Descricao = ?
                    WHERE id = ?
                """, (nova_descricao, record_id))
                logger.info(f"Descrição atualizada automaticamente para o código {value}: {nova_descricao}")
        
        # Verificar se alguma linha foi afetada
        if cursor.rowcount == 0:
            logger.warning(f"Nenhuma linha afetada pela atualização do registro {record_id}")
            if conexao:
                conexao.rollback()
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})
        
        # Commit da transação
        conexao.commit()
        
        logger.info(f"Registro {record_id} atualizado com sucesso")
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
        
    except Exception as e:
        logger.error(f"Erro ao atualizar registro: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Fazer rollback em caso de erro apenas se a conexão existir
        if conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")
        
        return jsonify({'success': False, 'message': f'Erro ao atualizar registro: {str(e)}'})
    
    finally:
        # Fechar recursos de forma segura
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            logger.error(f"Erro ao fechar cursor: {e}")
        
        try:
            if conexao:
                conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexão: {e}")

@tipoproduto_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez - agora também atualiza descrições automaticamente"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        # Obter banco homólogo (para buscar descrições WF)
        banco_homo = obter_banco_homo(projeto_id)

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()
        success_count = 0
        error_count = 0
        error_messages = []

        # Colunas permitidas para edição - REMOVIDO TipoProduto_Descricao
        colunas_permitidas = ['TipoProduto_Codigo', 'TipoProduto_GrupoContabilCod', 'tpd_grupocontab']

        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')

                if not record_id or not field:
                    error_count += 1
                    continue

                if field not in colunas_permitidas:
                    error_count += 1
                    continue

                # Atualiza o campo alterado
                query = f"UPDATE TipoProduto_DePara SET {field} = ? WHERE id = ?"
                cursor.execute(query, (value, record_id))
                
                # ⚙️ Se for o campo TipoProduto_Codigo, busca e atualiza automaticamente a descrição
                if field == 'TipoProduto_Codigo' and banco_homo:
                    nova_descricao = None
                    if value and value != 'S/DePara':
                        nova_descricao = obter_descricao_wf(banco_homo, value)
                    
                    if nova_descricao:
                        cursor.execute("""
                            UPDATE TipoProduto_DePara
                            SET TipoProduto_Descricao = ?
                            WHERE id = ?
                        """, (nova_descricao, record_id))
                        logger.info(f"Descrição atualizada automaticamente para o código {value}: {nova_descricao}")
                
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")

            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")

        conexao.commit()
        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")

        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }

        if error_messages:
            response['error_details'] = error_messages[:10]

        return jsonify(response)

    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}")
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})

    finally:
        if cursor:
            cursor.close()
        if conexao:
            conexao.close()

@tipoproduto_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})
        
        descricao = obter_descricao_wf(banco_homo, codigo)
        
        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })
            
    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})

# Função para dashboard (se necessário)
def dados_tipoproduto(banco_usuario):
    if not banco_usuario:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    conexao = conectar_segunda_base(banco_usuario)
    if not conexao:
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}

    try:
        cursor = conexao.cursor()
        
        # Usando função segura para evitar erro
        cursor.execute("SELECT COUNT(*) FROM TipoProduto_DePara")
        qtd = safe_fetchone(cursor)

        cursor.execute(
            "SELECT COUNT(*) FROM TipoProduto_DePara WHERE TipoProduto_Codigo = 'S/DePara'"
        )
        qtdPendente = safe_fetchone(cursor)

        percentualConclusao = ((qtd - qtdPendente) / qtd * 100) if qtd > 0 else 0

        return {
            "qtd": qtd,
            "qtdPendente": qtdPendente,
            "percentualConclusao": round(percentualConclusao, 1),
        }
    except Exception as e:
        logger.error(f"Erro ao calcular dados TipoProduto_DePara: {e}")
        return {"qtd": 0, "qtdPendente": 0, "percentualConclusao": 0}
    finally:
        if conexao:
            conexao.close()