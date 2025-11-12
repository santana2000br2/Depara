from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import import_from_excel
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

grupoproduto_bp = Blueprint("grupoproduto", __name__)

# ----------------------------
# Helpers (BancoHomo, WF lookups)
# ----------------------------
def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None


def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela GrupoProduto do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        cursor.execute("SELECT GrupoProduto_Codigo FROM GrupoProduto")
        registros = cursor.fetchall()
        
        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []


def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela GrupoProduto do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT GrupoProduto_Descricao FROM GrupoProduto WHERE GrupoProduto_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None


def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """Atualiza automaticamente as descrições após importação baseado nos códigos WF"""
    try:
        if not banco_homo:
            return
        
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("Falha ao conectar para atualizar descrições")
            return
        
        cursor = conexao.cursor()
        
        # Buscar registros que têm código WF (não são S/DePara)
        cursor.execute("""
            SELECT id, GrupoProduto_Codigo, GrupoProduto_Descricao 
            FROM GrupoProduto_DePara 
            WHERE GrupoProduto_Codigo IS NOT NULL AND GrupoProduto_Codigo != 'S/DePara'
        """)
        registros = cursor.fetchall()
        
        atualizacoes = 0
        for registro in registros:
            id_registro, codigo_wf, descricao_atual = registro
            descricao_wf = obter_descricao_wf(banco_homo, codigo_wf)
            
            if descricao_wf and descricao_wf != descricao_atual:
                # Atualizar descrição
                cursor.execute("""
                    UPDATE GrupoProduto_DePara 
                    SET GrupoProduto_Descricao = ? 
                    WHERE id = ?
                """, (descricao_wf, id_registro))
                atualizacoes += 1
                logger.info(f"Descrição atualizada para código {codigo_wf}: {descricao_wf}")
        
        conexao.commit()
        cursor.close()
        conexao.close()
        
        logger.info(f"Atualizações automáticas de descrição: {atualizacoes} registros")
        
    except Exception as e:
        logger.error(f"Erro ao atualizar descrições após importação: {str(e)}")


# ----------------------------
# Rotas
# ----------------------------
@grupoproduto_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('grupoproduto.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - INCLUIR todas as colunas para edição
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM GrupoProduto_DePara")
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('grupoproduto.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em grupoproduto: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('grupoproduto.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])


@grupoproduto_bp.route("/exportar")
def exportar_grupoproduto():
    """Exporta a tabela GrupoProduto_DePara incluindo o campo id"""
    try:
        if "projeto_selecionado" not in session:
            flash("Nenhum projeto selecionado.", "error")
            return redirect(url_for("auth.selecionar_projeto"))

        projeto = session["projeto_selecionado"]
        banco_usuario = projeto.get("DadosGX")
        projeto_id = projeto.get("ProjetoID")

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f"Falha na conexão com o banco: {banco_usuario}", "error")
            return redirect(url_for("auth.selecionar_projeto"))

        cursor = conexao.cursor()
        cursor.execute("SELECT id, grup_cd, grup_ds, GrupoProduto_Codigo, GrupoProduto_Descricao, ProdutoMarca_MarcaCod FROM GrupoProduto_DePara")
        rows = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="GrupoProduto_DePara")
        else:
            ws.title = "GrupoProduto_DePara"

        # cabeçalho
        for i, c in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                v = val.strip() if isinstance(val, str) else val
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                # destaque para código WF
                if colunas[c_idx - 1] == "GrupoProduto_Codigo":
                    if not v or v == "":
                        cell.fill = laranja
                    elif str(v) == "S/DePara":
                        cell.fill = amarelo
                    elif v and str(v) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        # Ajustar largura das colunas - CORREÇÃO DO ERRO
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        cursor.close()
        conexao.close()
        return send_file(output, as_attachment=True, download_name="GrupoProduto_DePara.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logger.error(f"Erro exportar grupoproduto: {str(e)}")
        flash(f"Erro: {str(e)}", "error")
        return redirect(url_for("grupoproduto.index"))


@grupoproduto_bp.route("/exportar_filtrados", methods=["POST"])
def exportar_grupoproduto_filtrados():
    """Exporta apenas os registros filtrados enviados pelo front-end"""
    try:
        data = request.get_json()
        registros_filtrados = data.get("registros", [])
        headers = data.get("headers", [])
        if not registros_filtrados:
            return jsonify({"success": False, "message": "Nenhum registro para exportar"}), 400
        if "projeto_selecionado" not in session:
            return jsonify({"success": False, "message": "Nenhum projeto selecionado"}), 400

        projeto = session["projeto_selecionado"]
        projeto_id = projeto.get("ProjetoID")
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        ws = wb.create_sheet(title="GrupoProduto_Filtrado")
        # headers
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for r_idx, reg in enumerate(registros_filtrados, start=2):
            for c_idx, header in enumerate(headers, start=1):
                valor = reg.get(header, "")
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                cell = ws.cell(row=r_idx, column=c_idx, value=valor)
                if header == "GrupoProduto_Codigo":
                    if not valor or valor == "":
                        cell.fill = laranja
                    elif valor == "S/DePara":
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        # ajustar largura de colunas - CORREÇÃO DO ERRO
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="GrupoProduto_Filtrado.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logger.error(f"Erro exportar filtrados grupoproduto: {str(e)}")
        return jsonify({"success": False, "message": f"Erro na exportação: {str(e)}"}), 500


@grupoproduto_bp.route("/export_wf")
def export_wf():
    """Exporta a tabela GrupoProduto do banco homólogo"""
    try:
        if "projeto_selecionado" not in session:
            flash("Nenhum projeto selecionado.", "error")
            return redirect(url_for("auth.selecionar_projeto"))

        projeto = session["projeto_selecionado"]
        projeto_id = projeto.get("ProjetoID")
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash("Banco homólogo não configurado para este projeto.", "error")
            return redirect(url_for("grupoproduto.index"))

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f"Falha na conexão com o banco homólogo: {banco_homo}", "error")
            return redirect(url_for("grupoproduto.index"))

        cursor = conexao.cursor()
        cursor.execute("SELECT GrupoProduto_Codigo, GrupoProduto_Descricao, GrupoProduto_Tipo, GrupoProduto_Ativo FROM GrupoProduto")
        rows = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros = [dict(zip(colunas, r)) for r in rows]
        cursor.close()
        conexao.close()

        df = pd.DataFrame(registros, columns=colunas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="GrupoProduto_WF", index=False)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="GrupoProduto_WF.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logger.error(f"Erro export_wf grupoproduto: {str(e)}")
        flash(f"Erro: {str(e)}", "error")
        return redirect(url_for("grupoproduto.index"))


@grupoproduto_bp.route("/importar", methods=["POST"])
def importar_grupoproduto():
    """Importa Excel e atualiza por id (sem INSERT). Após update sincroniza descrições com BancoHomo."""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "Nenhum arquivo enviado"})
        arquivo = request.files["file"]
        if not arquivo or not arquivo.filename:
            return jsonify({"success": False, "message": "Nenhum arquivo selecionado"})
        if not (arquivo.filename.endswith(".xlsx") or arquivo.filename.endswith(".xls")):
            return jsonify({"success": False, "message": "Formato inválido. Use .xlsx ou .xls"})
        if "projeto_selecionado" not in session:
            return jsonify({"success": False, "message": "Nenhum projeto selecionado"})

        projeto = session["projeto_selecionado"]
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({"success": False, "message": "Banco não configurado"})

        df = pd.read_excel(arquivo)
        df = df.where(pd.notnull(df), None)
        
        # CORREÇÃO: Converter nomes das colunas para minúsculas para comparação case-insensitive
        colunas_excel = [col.lower() for col in df.columns.tolist()]
        
        obrigatorias = {"id", "grup_cd", "grup_ds", "grupoproduto_codigo", "grupoproduto_descricao", "produtomarca_marcacod"}
        faltando = obrigatorias - set(colunas_excel)
        if faltando:
            return jsonify({"success": False, "message": f"Colunas faltando: {', '.join(faltando)}"})

        # CORREÇÃO: Mapear colunas do DataFrame para os nomes corretos (case-insensitive)
        col_mapping = {}
        for col in df.columns:
            col_lower = col.lower()
            if col_lower in obrigatorias:
                col_mapping[col_lower] = col

        registros = df.to_dict("records")
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({"success": False, "message": f"Falha na conexão com o banco: {banco_usuario}"})

        cursor = conexao.cursor()
        contador_update = 0
        contador_sem_id = 0
        for reg in registros:
            # CORREÇÃO: Usar o mapeamento para acessar os valores corretamente
            record_id = reg.get(col_mapping.get("id"))
            if not record_id:
                contador_sem_id += 1
                continue
            
            cursor.execute("""
                UPDATE GrupoProduto_DePara
                SET grup_cd = ?, grup_ds = ?, GrupoProduto_Codigo = ?, GrupoProduto_Descricao = ?, ProdutoMarca_MarcaCod = ?
                WHERE id = ?
            """, (
                reg.get(col_mapping.get("grup_cd")),
                reg.get(col_mapping.get("grup_ds")),
                reg.get(col_mapping.get("grupoproduto_codigo")),
                reg.get(col_mapping.get("grupoproduto_descricao")),
                reg.get(col_mapping.get("produtomarca_marcacod")),
                record_id
            ))
            if cursor.rowcount > 0:
                contador_update += 1

        conexao.commit()
        logger.info(f"Importação GrupoProduto: {contador_update} atualizados, {contador_sem_id} ignorados (sem id)")

        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        cursor.close()
        conexao.close()
        return jsonify({"success": True, "message": f"Importação concluída! {contador_update} atualizados, {contador_sem_id} ignorados."})

    except Exception as e:
        import traceback
        logger.error(f"Erro importar grupoproduto: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        try:
            if 'conexao' in locals() and conexao:
                conexao.rollback()
                conexao.close()
        except:
            pass
        return jsonify({"success": False, "message": f"Erro: {str(e)}"})


@grupoproduto_bp.route("/update", methods=["POST"])
def update_registro():
    """Atualiza um campo específico do registro (edição inline)"""
    logger.info("=== UPDATE REGISTRO GRUPOPRODUTO ENDPOINT ACESSADO ===")
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        record_id = data.get("id")
        field = data.get("field")
        value = data.get("value")
        if not record_id or not field:
            return jsonify({"success": False, "message": "ID e campo são obrigatórios"})

        colunas_permitidas = ["GrupoProduto_Codigo", "GrupoProduto_Descricao", "ProdutoMarca_MarcaCod"]
        if field not in colunas_permitidas:
            return jsonify({"success": False, "message": f"Campo {field} não permitido"})

        if "projeto_selecionado" not in session:
            return jsonify({"success": False, "message": "Nenhum projeto selecionado"})

        projeto = session["projeto_selecionado"]
        banco_usuario = projeto.get("DadosGX")
        if not banco_usuario:
            return jsonify({"success": False, "message": "Banco não configurado"})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({"success": False, "message": f"Falha na conexão com o banco: {banco_usuario}"})

        cursor = conexao.cursor()
        try:
            cursor.execute(f"SELECT TOP 1 {field} FROM GrupoProduto_DePara WHERE id = ?", (record_id,))
            res = cursor.fetchone()
            if not res:
                return jsonify({"success": False, "message": "Registro não encontrado"})
        except Exception as e:
            logger.error(f"Erro ao verificar campo: {str(e)}")
            return jsonify({"success": False, "message": f"Campo {field} não existe"})

        cursor.execute(f"UPDATE GrupoProduto_DePara SET {field} = ? WHERE id = ?", (value, record_id))

        # se alterou o código WF, tentar atualizar a descrição automática
        if field == "GrupoProduto_Codigo":
            projeto = session.get("projeto_selecionado")
            projeto_id = projeto.get("ProjetoID") if projeto and isinstance(projeto, dict) else None
            if projeto_id:
                banco_homo = obter_banco_homo(projeto_id)
                if banco_homo and value and value != "S/DePara":
                    nova_desc = obter_descricao_wf(banco_homo, value)
                    if nova_desc:
                        cursor.execute("UPDATE GrupoProduto_DePara SET GrupoProduto_Descricao = ? WHERE id = ?", (nova_desc, record_id))

        if cursor.rowcount == 0:
            conexao.rollback()
            return jsonify({"success": False, "message": "Registro não encontrado ou não modificado"})

        conexao.commit()
        cursor.close()
        conexao.close()
        return jsonify({"success": True, "message": "Registro atualizado"})
    except Exception as e:
        logger.error(f"Erro update registro grupoproduto: {str(e)}")
        if conexao:
            try:
                conexao.rollback()
            except:
                pass
        return jsonify({"success": False, "message": f"Erro: {str(e)}"})
    finally:
        try:
            if cursor:
                cursor.close()
        except:
            pass
        try:
            if conexao:
                conexao.close()
        except:
            pass


@grupoproduto_bp.route("/update_batch", methods=["POST"])
def update_batch():
    """Atualiza vários registros em lote"""
    logger.info("=== UPDATE BATCH GRUPOPRODUTO ENDPOINT ACESSADO ===")
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get("updates", [])
        if not updates:
            return jsonify({"success": False, "message": "Nenhuma atualização fornecida"})

        if "projeto_selecionado" not in session:
            return jsonify({"success": False, "message": "Nenhum projeto selecionado"})

        projeto = session["projeto_selecionado"]
        banco_usuario = projeto.get("DadosGX")
        projeto_id = projeto.get("ProjetoID")
        if not banco_usuario:
            return jsonify({"success": False, "message": "Banco não configurado"})

        banco_homo = obter_banco_homo(projeto_id)
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({"success": False, "message": f"Falha na conexão com o banco: {banco_usuario}"})

        cursor = conexao.cursor()
        success = 0
        errors = []
        for upd in updates:
            try:
                record_id = upd.get("id")
                field = upd.get("field")
                value = upd.get("value")
                if not record_id or not field:
                    errors.append(f"Dados incompletos para update: {upd}")
                    continue
                if field not in ["GrupoProduto_Codigo", "GrupoProduto_Descricao", "ProdutoMarca_MarcaCod"]:
                    errors.append(f"Campo não permitido: {field}")
                    continue
                cursor.execute(f"UPDATE GrupoProduto_DePara SET {field} = ? WHERE id = ?", (value, record_id))
                if field == "GrupoProduto_Codigo" and banco_homo:
                    if value and value != "S/DePara":
                        nova = obter_descricao_wf(banco_homo, value)
                        if nova:
                            cursor.execute("UPDATE GrupoProduto_DePara SET GrupoProduto_Descricao = ? WHERE id = ?", (nova, record_id))
                if cursor.rowcount > 0:
                    success += 1
                else:
                    errors.append(f"Registro não encontrado: {record_id}")
            except Exception as e:
                errors.append(f"Erro ao atualizar {record_id}: {str(e)}")

        conexao.commit()
        cursor.close()
        conexao.close()
        resp = {"success": True, "message": f"{success} atualizações realizadas", "errors": errors[:10]}
        return jsonify(resp)
    except Exception as e:
        logger.error(f"Erro batch grupoproduto: {str(e)}")
        if conexao:
            try:
                conexao.rollback()
            except:
                pass
        return jsonify({"success": False, "message": f"Erro: {str(e)}"})
    finally:
        try:
            if cursor:
                cursor.close()
        except:
            pass
        try:
            if conexao:
                conexao.close()
        except:
            pass


@grupoproduto_bp.route("/get_descricao_wf/<codigo>")
def get_descricao_wf(codigo):
    """Retorna a descrição do código na base WF (BancoHomo)"""
    try:
        if "projeto_selecionado" not in session:
            return jsonify({"success": False, "message": "Nenhum projeto selecionado"})
        projeto = session["projeto_selecionado"]
        projeto_id = projeto.get("ProjetoID")
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({"success": False, "message": "Banco homólogo não configurado"})
        descricao = obter_descricao_wf(banco_homo, codigo)
        if descricao:
            return jsonify({"success": True, "descricao": descricao})
        else:
            return jsonify({"success": False, "message": "Código não encontrado na base WF"})
    except Exception as e:
        logger.error(f"Erro get_descricao_wf grupoproduto: {str(e)}")
        return jsonify({"success": False, "message": f"Erro: {str(e)}"})


# Execução direta para testes
if __name__ == "__main__":
    from flask import Flask
    app = Flask(__name__)
    app.register_blueprint(grupoproduto_bp, url_prefix="/grupoproduto")
    app.run(debug=True)