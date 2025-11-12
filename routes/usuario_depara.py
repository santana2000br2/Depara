# usuario_depara.py
from flask import (
    Blueprint, render_template, redirect, url_for, session, flash,
    request, jsonify, send_file
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

usuario_depara_bp = Blueprint('usuario_depara', __name__)

# -------------------------
# Helpers / utilitários
# -------------------------
def obter_banco_homo(projeto_id):
    """Retorna BancoHomo do Projeto (mesma lógica usada nos outros módulos)."""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("obter_banco_homo: falha ao conectar ao banco principal")
            return None
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return row[0] if row and row[0] else None
    except Exception as e:
        logger.error(f"obter_banco_homo (usuario) -> {e}", exc_info=True)
        return None


def obter_codigos_wf(banco_homo):
    """Retorna lista de Usuario_Codigo existentes na tabela WF (usuario)."""
    try:
        if not banco_homo:
            return []
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"obter_codigos_wf: falha na conexão com banco_homo={banco_homo}")
            return []
        cursor = conexao.cursor()
        cursor.execute("SELECT Usuario_Codigo FROM usuario")
        rows = cursor.fetchall()
        cursor.close()
        conexao.close()
        codigos = [str(r[0]) for r in rows if r and r[0] is not None]
        logger.info(f"obter_codigos_wf (usuario): encontrados {len(codigos)} códigos")
        return codigos
    except Exception as e:
        logger.error(f"obter_codigos_wf (usuario) -> {e}", exc_info=True)
        return []


def obter_descricao_wf(banco_homo, codigo):
    """
    Retorna a descrição/nome (Usuario_Nome e Usuario_Identificador) para um Usuario_Codigo
    na tabela WF (usuario). Aqui retornamos preferencialmente Usuario_Nome.
    """
    try:
        if not banco_homo or not codigo:
            return None
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"obter_descricao_wf: falha na conexão banco_homo={banco_homo}")
            return None
        cursor = conexao.cursor()
        cursor.execute("SELECT Usuario_Nome, Usuario_Identificador FROM usuario WHERE Usuario_Codigo = ?", codigo)
        row = cursor.fetchone()
        cursor.close()
        conexao.close()
        if row:
            nome = row[0] if row[0] is not None else None
            ident = row[1] if len(row) > 1 and row[1] is not None else None
            # retornar a combinação ou apenas o nome conforme necessidade
            return nome or ident
        return None
    except Exception as e:
        logger.error(f"obter_descricao_wf (usuario) codigo={codigo} -> {e}", exc_info=True)
        return None


def atualizar_descricoes_apos_importacao(banco_usuario, banco_homo):
    """
    Atualiza Usuario_Nome/Usuario_Identificador na tabela Usuario_depara a partir do banco homólogo.
    """
    conexao = None
    cursor = None
    conexao_aux = None
    cursor_aux = None
    try:
        if not banco_homo:
            logger.info("atualizar_descricoes_apos_importacao: banco_homo não configurado, pulando.")
            return

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error("atualizar_descricoes_apos_importacao: falha ao conectar banco_usuario")
            return

        cursor = conexao.cursor()
        cursor.execute("""
            SELECT id, Usuario_Codigo, Usuario_Nome, Usuario_Identificador
            FROM Usuario_depara
            WHERE Usuario_Codigo IS NOT NULL AND Usuario_Codigo != 'S/DePara'
        """)
        rows = cursor.fetchall()
        logger.info(f"atualizar_descricoes_apos_importacao: {len(rows)} registros para verificação")

        if not rows:
            cursor.close()
            conexao.close()
            return

        conexao_aux = conectar_segunda_base(banco_homo)
        if not conexao_aux:
            logger.error("atualizar_descricoes_apos_importacao: falha ao conectar banco_homo")
            cursor.close()
            conexao.close()
            return
        cursor_aux = conexao_aux.cursor()

        updates = 0
        for r in rows:
            idr = r[0]
            codigo = r[1]
            desc_atual = r[2] if len(r) > 2 else None
            ident_atual = r[3] if len(r) > 3 else None
            if not codigo:
                continue
            cursor_aux.execute("SELECT Usuario_Nome, Usuario_Identificador FROM usuario WHERE Usuario_Codigo = ?", (codigo,))
            res = cursor_aux.fetchone()
            if res:
                nome_wf = res[0] if res[0] is not None else None
                ident_wf = res[1] if len(res) > 1 and res[1] is not None else None
                # atualiza se diferente
                if nome_wf and nome_wf != desc_atual:
                    try:
                        cursor.execute("UPDATE Usuario_depara SET Usuario_Nome = ? WHERE id = ?", (nome_wf, idr))
                        updates += 1
                    except Exception as ex:
                        logger.warning(f"Falha ao atualizar Usuario_Nome id={idr}: {ex}")
                # atualizar identificador se necessário
                if ident_wf and ident_wf != ident_atual:
                    try:
                        cursor.execute("UPDATE Usuario_depara SET Usuario_Identificador = ? WHERE id = ?", (ident_wf, idr))
                        updates += 1
                    except Exception as ex:
                        logger.warning(f"Falha ao atualizar Usuario_Identificador id={idr}: {ex}")

        if updates > 0:
            conexao.commit()
            logger.info(f"atualizar_descricoes_apos_importacao: atualizadas {updates} colunas")
        else:
            logger.info("atualizar_descricoes_apos_importacao: nenhuma atualização necessária")

        cursor_aux.close()
        conexao_aux.close()
        cursor.close()
        conexao.close()
    except Exception as e:
        logger.error(f"atualizar_descricoes_apos_importacao (usuario) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass


# -------------------------
# ROTAS
# -------------------------
@usuario_depara_bp.route('/')

def index():
    conexao = None
    cursor = None
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        projeto_nome = projeto.get('NomeProjeto', 'N/A')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('usuario.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Usuario_depara")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros_dict = [dict(zip(colunas, r)) for r in registros]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        cursor.close()
        conexao.close()

        return render_template('usuario.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf)
    except Exception as e:
        logger.error(f"index (usuario) -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro ao carregar usuários: {e}', 'error')
        return render_template('usuario.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])


@usuario_depara_bp.route('/exportar')
def exportar_usuario():
    """Exporta Usuario_depara inteiro com coloração em Usuario_Codigo"""
    conexao = None
    cursor = None
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('usuario.index'))

        cursor = conexao.cursor()
        cursor.execute("SELECT fun_cd, fun_nm, Usuario_Codigo, Usuario_Nome, Usuario_Identificador FROM Usuario_depara")
        registros = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="Usuario_depara")

        # cabeçalhos amigáveis
        mapeamento = {
            'fun_cd': 'Codigo de Origem',
            'fun_nm': 'Descrição de origem',
            'Usuario_Codigo': 'Usuario_Codigo',
            'Usuario_Nome': 'Usuario_Nome',
            'Usuario_Identificador': 'Usuario_Identificador'
        }
        colunas_amigaveis = [mapeamento.get(c, c) for c in colunas]
        for idx, h in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for rnum, row in enumerate(registros, 2):
            for cnum, val in enumerate(row, 1):
                v = val.strip() if isinstance(val, str) else val
                ws.cell(row=rnum, column=cnum, value=v)
                # Usuario_Codigo é coluna 3 (1-based)
                if cnum == 3:
                    cell = ws.cell(row=rnum, column=cnum)
                    if not v or v == '':
                        cell.fill = laranja
                    elif v == 'S/DePara':
                        cell.fill = amarelo
                    elif str(v) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        # ajustar largura
        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        cursor.close()
        conexao.close()

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Usuario_depara.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"exportar_usuario -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro na exportação: {e}', 'error')
        return redirect(url_for('usuario.index'))


@usuario_depara_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_usuario_filtrados():
    """Exporta os registros filtrados (recebidos do front) em Excel com coloração."""
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo)

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title="Usuario_Filtrado")

        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for rnum, registro in enumerate(registros_filtrados, 2):
            for cnum, header in enumerate(headers, 1):
                val = registro.get(header, '')
                if isinstance(val, str):
                    val = val.strip()
                cell = ws.cell(row=rnum, column=cnum, value=val)
                if header == 'Usuario_Codigo':
                    if not val or val == '':
                        cell.fill = laranja
                    elif val == 'S/DePara':
                        cell.fill = amarelo
                    elif val and str(val) in codigos_wf:
                        cell.fill = verde
                    else:
                        cell.fill = vermelho

        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Usuario_Filtrado.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument-spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"exportar_usuario_filtrados -> {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro: {e}'}), 500


@usuario_depara_bp.route('/export_wf')
def export_wf():
    """
    Exporta a tabela WF 'usuario' com:
    select Usuario_Codigo, Usuario_Identificador, Usuario_Nome, Usuario_Ativo from usuario
    """
    conexao = None
    cursor = None
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('usuario.index'))

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('usuario.index'))

        cursor = conexao.cursor()
        cursor.execute("SELECT Usuario_Codigo, Usuario_Identificador, Usuario_Nome, Usuario_Ativo FROM usuario")
        rows = cursor.fetchall()
        colunas = [c[0] for c in cursor.description]
        registros = [dict(zip(colunas, r)) for r in rows]
        cursor.close()
        conexao.close()

        df = pd.DataFrame(registros, columns=colunas)
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Usuario_WF', index=False)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Usuario_WF.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"export_wf (usuario) -> {e}", exc_info=True)
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        flash(f'Erro na exportação WF: {e}', 'error')
        return redirect(url_for('usuario.index'))


@usuario_depara_bp.route('/importar', methods=['POST'])
def importar_usuario():
    """
    Importa Excel para Usuario_depara.
    Suporta cabeçalhos amigáveis.
    Faz UPDATE quando fun_cd já existe, senão INSERT.
    """
    conexao = None
    cursor = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        filename = arquivo.filename
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use .xlsx ou .xls'})

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para o projeto'})

        logger.info(f"importar_usuario: iniciando importação para banco {banco_usuario}")

        # Ler Excel
        try:
            df = pd.read_excel(arquivo)
            df = df.where(pd.notnull(df), None)
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()
        except Exception as e:
            logger.error(f"importar_usuario: erro ao ler excel -> {e}", exc_info=True)
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {e}'})

        logger.info(f"importar_usuario: arquivo lido: {len(registros)} registros, colunas: {colunas_excel}")

        # Mapeamento nomes amigáveis -> técnicos
        mapeamento = {
            'Codigo de Origem': 'fun_cd',
            'Descrição de origem': 'fun_nm',
            'fun_cd': 'fun_cd',
            'fun_nm': 'fun_nm',
            'Usuario_Codigo': 'Usuario_Codigo',
            'Usuario_Nome': 'Usuario_Nome',
            'Usuario_Identificador': 'Usuario_Identificador'
        }
        colunas_normalizadas = [mapeamento.get(c, c) for c in colunas_excel]
        df.columns = colunas_normalizadas

        # Validar colunas obrigatórias
        colunas_necessarias = {'fun_cd', 'fun_nm', 'Usuario_Codigo', 'Usuario_Nome', 'Usuario_Identificador'}
        if not colunas_necessarias.issubset(set(colunas_normalizadas)):
            missing = colunas_necessarias - set(colunas_normalizadas)
            return jsonify({'success': False, 'message': f'Colunas necessárias faltando no arquivo: {", ".join(missing)}'})

        registros = df.to_dict('records')

        # Normalizar tamanhos
        registros_filtrados = []
        for reg in registros:
            r = {}
            r['fun_cd'] = None if reg.get('fun_cd') is None else str(reg.get('fun_cd'))[:100]
            r['fun_nm'] = None if reg.get('fun_nm') is None else str(reg.get('fun_nm'))[:200]
            r['Usuario_Codigo'] = None if reg.get('Usuario_Codigo') is None else str(reg.get('Usuario_Codigo'))[:100]
            r['Usuario_Nome'] = None if reg.get('Usuario_Nome') is None else str(reg.get('Usuario_Nome'))[:200]
            r['Usuario_Identificador'] = None if reg.get('Usuario_Identificador') is None else str(reg.get('Usuario_Identificador'))[:200]
            registros_filtrados.append(r)

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # estatísticas antes
        cursor.execute("SELECT COUNT(*) FROM Usuario_depara")
        antes = cursor.fetchone()
        count_antes = antes[0] if antes else 0

        contador_update = 0
        contador_insert = 0

        for reg in registros_filtrados:
            chave = reg.get('fun_cd')
            if chave:
                cursor.execute("SELECT id FROM Usuario_depara WHERE fun_cd = ?", (chave,))
                existe = cursor.fetchone()
                if existe:
                    cursor.execute("""
                        UPDATE Usuario_depara
                        SET fun_nm = ?, Usuario_Codigo = ?, Usuario_Nome = ?, Usuario_Identificador = ?
                        WHERE fun_cd = ?
                    """, (reg.get('fun_nm'), reg.get('Usuario_Codigo'), reg.get('Usuario_Nome'), reg.get('Usuario_Identificador'), chave))
                    contador_update += 1
                else:
                    cursor.execute("""
                        INSERT INTO Usuario_depara (fun_cd, fun_nm, Usuario_Codigo, Usuario_Nome, Usuario_Identificador)
                        VALUES (?, ?, ?, ?, ?)
                    """, (reg.get('fun_cd'), reg.get('fun_nm'), reg.get('Usuario_Codigo'), reg.get('Usuario_Nome'), reg.get('Usuario_Identificador')))
                    contador_insert += 1
            else:
                # sem chave: insere mesmo assim
                cursor.execute("""
                    INSERT INTO Usuario_depara (fun_cd, fun_nm, Usuario_Codigo, Usuario_Nome, Usuario_Identificador)
                    VALUES (?, ?, ?, ?, ?)
                """, (reg.get('fun_cd'), reg.get('fun_nm'), reg.get('Usuario_Codigo'), reg.get('Usuario_Nome'), reg.get('Usuario_Identificador')))
                contador_insert += 1

        conexao.commit()
        logger.info(f"importar_usuario: atualizados={contador_update}, inseridos={contador_insert}")

        # Atualizar descrições a partir do banco homólogo
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        cursor.execute("SELECT COUNT(*) FROM Usuario_depara")
        depois = cursor.fetchone()
        count_depois = depois[0] if depois else 0

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': f'Importação concluída! {contador_update} atualizados, {contador_insert} inseridos. Antes: {count_antes}, Depois: {count_depois}.'})
    except Exception as e:
        logger.error(f"importar_usuario (usuario) -> {e}", exc_info=True)
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if 'cursor' in locals() and cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if 'conexao' in locals() and conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro na importação: {e}'}), 500


@usuario_depara_bp.route('/update', methods=['POST'])
def update_registro():
    """Atualiza um único registro (Usuario_Codigo ou Usuario_Nome ou Usuario_Identificador)."""
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not record_id or not field:
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'}), 400

        campos_permitidos = ['Usuario_Codigo', 'Usuario_Nome', 'Usuario_Identificador']
        if field not in campos_permitidos:
            return jsonify({'success': False, 'message': 'Campo não permitido para edição'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'}), 500
        cursor = conexao.cursor()

        if field == 'Usuario_Codigo':
            if value is not None and value != 'S/DePara' and not str(value).isdigit():
                return jsonify({'success': False, 'message': 'Usuario_Codigo deve ser numérico ou "S/DePara"'}), 400

        cursor.execute(f"UPDATE Usuario_depara SET {field} = ? WHERE id = ?", (value, record_id))
        conexao.commit()

        # se alterou Usuario_Codigo, tentar buscar Usuario_Nome no WF e atualizar
        if field == 'Usuario_Codigo' and value and value != 'S/DePara':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                desc = obter_descricao_wf(banco_homo, value)
                if desc:
                    try:
                        cursor.execute("UPDATE Usuario_depara SET Usuario_Nome = ? WHERE id = ?", (desc, record_id))
                        conexao.commit()
                    except Exception as ex:
                        logger.warning(f"update_registro: falha ao atualizar Usuario_Nome automaticamente: {ex}")

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
    except Exception as e:
        logger.error(f"update_registro (usuario) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {e}'}), 500


@usuario_depara_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Atualização em lote para registros editados no front-end (aceita lista de updates)."""
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização enviada'}), 400

        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão: {banco_usuario}'}), 500
        cursor = conexao.cursor()

        contador = 0
        for upd in updates:
            record_id = upd.get('id')
            field = upd.get('field')
            value = upd.get('value')
            if not record_id or not field:
                continue
            if field not in ['Usuario_Codigo', 'Usuario_Nome', 'Usuario_Identificador']:
                continue
            if field == 'Usuario_Codigo' and value is not None and value != 'S/DePara' and not str(value).isdigit():
                continue
            cursor.execute(f"UPDATE Usuario_depara SET {field} = ? WHERE id = ?", (value, record_id))
            contador += 1

        conexao.commit()

        # após atualização em lote, tentar atualizar descrições a partir do banco homólogo
        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            atualizar_descricoes_apos_importacao(banco_usuario, banco_homo)

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': f'{contador} registros atualizados com sucesso'})
    except Exception as e:
        logger.error(f"update_batch (usuario) -> {e}", exc_info=True)
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass
        try:
            if cursor is not None:
                cursor.close()
        except Exception:
            pass
        try:
            if conexao is not None:
                conexao.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar em lote: {e}'}), 500


@usuario_depara_bp.route('/get_descricao_wf/<codigo>', methods=['GET'])
def get_descricao_wf_endpoint(codigo):
    """Endpoint AJAX para obter Usuario_Nome/Usuario_Identificador a partir do WF."""
    try:
        projeto = session.get('projeto_selecionado')
        if not projeto:
            return jsonify({'descricao': None})
        projeto_id = projeto.get('ProjetoID')
        banco_homo = obter_banco_homo(projeto_id)
        desc = obter_descricao_wf(banco_homo, codigo)
        return jsonify({'descricao': desc})
    except Exception as e:
        logger.error(f"get_descricao_wf_endpoint (usuario) -> {e}", exc_info=True)
        return jsonify({'descricao': None})
