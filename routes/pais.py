from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import tempfile
import os
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

pais_bp = Blueprint("pais", __name__)

def obter_banco_homo(projeto_id):
    """Função para obter o BancoHomo diretamente do banco de dados"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None
        
        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela Pais do banco homólogo"""
    try:
        if not banco_homo:
            return []
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Pais_Codigo FROM Pais")
        registros = cursor.fetchall()
        
        # Criar uma lista com todos os códigos (convertidos para string para comparação)
        codigos = [str(registro[0]) for registro in registros if registro[0] is not None]
        
        cursor.close()
        conexao.close()
        
        logger.info(f"Encontrados {len(codigos)} códigos na base WF")
        return codigos
        
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF: {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela Pais do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        
        cursor = conexao.cursor()
        cursor.execute("SELECT Pais_Nome FROM Pais WHERE Pais_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        
        cursor.close()
        conexao.close()
        
        if resultado and resultado[0]:
            return resultado[0]
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo}: {str(e)}")
        return None

def atualizar_descricoes_apos_importacao(banco_usuario, projeto_id):
    """
    Atualiza o campo Pais_Nome na tabela Pais_DePara,
    com base no Pais_Codigo e nas descrições da base de homologação (BancoHomo).
    """
    conexao_usuario = None
    cursor_usuario = None
    try:
        logger.info("🔄 Iniciando atualização de descrições (Pais_Nome) após importação...")

        # Conexão com o banco do usuário (DadosGX)
        conexao_usuario = conectar_segunda_base(banco_usuario)
        if not conexao_usuario:
            logger.error(f"❌ Falha ao conectar ao banco do usuário: {banco_usuario}")
            return False

        cursor_usuario = conexao_usuario.cursor()

        # Obter códigos únicos de País existentes na tabela DePara
        cursor_usuario.execute("""
            SELECT DISTINCT Pais_Codigo
            FROM Pais_DePara
            WHERE Pais_Codigo IS NOT NULL AND Pais_Codigo <> 'S/DePara'
        """)
        codigos = [row[0] for row in cursor_usuario.fetchall()]

        if not codigos:
            logger.warning("⚠️ Nenhum código de país encontrado para atualizar.")
            return True

        logger.info(f"✅ {len(codigos)} códigos encontrados para atualização.")

        # Conectar na base de homologação (BancoHomo)
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            logger.error("❌ Banco de homologação não configurado.")
            return False

        conexao_homo = conectar_segunda_base(banco_homo)
        if not conexao_homo:
            logger.error(f"❌ Falha ao conectar na base de homologação: {banco_homo}")
            return False

        cursor_homo = conexao_homo.cursor()

        # Atualizar um por um
        total_atualizados = 0
        for codigo in codigos:
            try:
                cursor_homo.execute("SELECT Pais_Nome FROM Pais WHERE Pais_Codigo = ?", (codigo,))
                resultado = cursor_homo.fetchone()
                nome_pais = resultado[0] if resultado else None

                if nome_pais:
                    cursor_usuario.execute(
                        "UPDATE Pais_DePara SET Pais_Nome = ? WHERE Pais_Codigo = ?",
                        (nome_pais, codigo)
                    )
                    total_atualizados += 1
            except Exception as e:
                logger.warning(f"Erro ao atualizar País '{codigo}': {e}")

        conexao_usuario.commit()

        logger.info(f"✅ Atualização de descrições concluída. {total_atualizados} registros atualizados.")
        cursor_usuario.close()
        conexao_usuario.close()
        cursor_homo.close()
        conexao_homo.close()
        return True

    except Exception as e:
        logger.error(f"❌ Erro durante atualização de descrições: {e}", exc_info=True)
        if conexao_usuario:
            conexao_usuario.rollback()
        return False

    finally:
        try:
            if cursor_usuario:
                cursor_usuario.close()
            if conexao_usuario:
                conexao_usuario.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexões: {e}")


@pais_bp.route("/")
def index():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        logger.info(f"Tentando conectar ao banco: {banco_usuario} para o projeto: {projeto_nome}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('pais.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])
        
        # Executar consulta - PRIMEIRO VAMOS VERIFICAR AS COLUNAS EXISTENTES
        cursor = conexao.cursor()
        
        # Verificar quais colunas existem na tabela
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'Pais_DePara'
        """)
        colunas_existentes = [row[0] for row in cursor.fetchall()]
        logger.info(f"Colunas existentes na tabela Pais_DePara: {colunas_existentes}")
        
        # Construir a consulta baseada nas colunas existentes
        if 'Pais_Nome' in colunas_existentes:
            cursor.execute("SELECT * FROM Pais_DePara")
        else:
            # Se Pais_Nome não existe, usar Pais_Descricao ou apenas as colunas básicas
            cursor.execute("SELECT id, pais_cd, pais_ds, Pais_Codigo FROM Pais_DePara")
        
        registros = cursor.fetchall()
        
        # Obter nomes das colunas
        colunas = [column[0] for column in cursor.description]
        
        # Converter para dicionários
        registros_dict = [dict(zip(colunas, row)) for row in registros]
        
        logger.info(f"Encontrados {len(registros_dict)} registros")
        
        # Obter códigos da base WF para comparação
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        return render_template('pais.html', 
                             registros=registros_dict, 
                             colunas=colunas,
                             projeto_nome=projeto_nome,
                             banco_usuario=banco_usuario,
                             codigos_wf=codigos_wf,
                             banco_homo=banco_homo)
        
    except Exception as e:
        logger.error(f"Erro em pais: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('pais.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@pais_bp.route('/exportar')
def exportar_pais():
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('pais.index'))
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('pais.index'))
        
        cursor = conexao.cursor()
        
        # VERIFICAR COLUNAS EXISTENTES NA TABELA
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'Pais_DePara'
        """)
        colunas_existentes = [row[0] for row in cursor.fetchall()]
        logger.info(f"Colunas existentes para exportação: {colunas_existentes}")
        
        # Construir a consulta baseada nas colunas existentes
        if 'Pais_Nome' in colunas_existentes:
            cursor.execute("SELECT pais_cd, pais_ds, Pais_Codigo, Pais_Nome FROM Pais_DePara")
            colunas_originais = ['pais_cd', 'pais_ds', 'Pais_Codigo', 'Pais_Nome']
            mapeamento_colunas = {
                'pais_cd': 'Cod. País',
                'pais_ds': 'País Descrição',
                'Pais_Codigo': 'Pais_Codigo', 
                'Pais_Nome': 'Pais_Nome'
            }
        else:
            # Usar apenas colunas básicas se Pais_Nome não existir
            cursor.execute("SELECT pais_cd, pais_ds, Pais_Codigo FROM Pais_DePara")
            colunas_originais = ['pais_cd', 'pais_ds', 'Pais_Codigo']
            mapeamento_colunas = {
                'pais_cd': 'Cod. País',
                'pais_ds': 'País Descrição',
                'Pais_Codigo': 'Pais_Codigo'
            }
        
        registros = cursor.fetchall()
        
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook e worksheet
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet explicitamente
        ws = wb.create_sheet(title="Pais_DePara")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna Pais_Codigo (coluna 3)
                if col_num == 3:  # Pais_Codigo
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = "Pais_DePara.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar pais: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('pais.index'))

@pais_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_pais_filtrados():
    try:
        # Obter dados da requisição
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter códigos WF para colorir
        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []
        
        # Criar workbook
        wb = Workbook()
        
        # Remover a sheet padrão se existir
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)
        
        # Criar uma nova worksheet
        ws = wb.create_sheet(title="Pais_Filtrado")
        
        # Adicionar cabeçalhos
        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            # Estilizar cabeçalhos
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Definir cores
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Para campos vazios
        
        # Adicionar dados com cores
        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                # Remover espaços extras dos valores
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                
                # Aplicar cores na coluna Pais_Codigo
                if header == 'Pais_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja  # Laranja para campos vazios
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar para buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="Pais_Filtrado.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar pais filtrada: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@pais_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela Pais do banco homólogo (BancoHomo)"""
    try:
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        # Obter o BancoHomo diretamente do banco de dados
        banco_homo = obter_banco_homo(projeto_id)
        
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('pais.index'))
        
        logger.info(f"Exportando tabela WF do banco homólogo: {banco_homo}")
        
        # Conectar ao banco homólogo
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('pais.index'))
        
        # Executar consulta na tabela Pais do banco homólogo
        cursor = conexao.cursor()
        cursor.execute("SELECT Pais_Codigo, Pais_Nome, Pais_Ativo FROM Pais")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        
        # Converter para lista de dicionários
        registros_list = [dict(zip(colunas, row)) for row in registros]
        
        # Fechar recursos
        cursor.close()
        conexao.close()
        
        # Criar DataFrame pandas
        df = pd.DataFrame(registros_list, columns=colunas)
        
        # Criar buffer para o arquivo Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Pais_WF', index=False)
        
        output.seek(0)
        
        # Nome do arquivo
        nome_arquivo = "Pais_WF.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF: {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('pais.index'))

@pais_bp.route('/importar', methods=['POST'])
def importar_pais():
    """Importa a planilha de Países (tolerante a variações de cabeçalho) e sincroniza automaticamente o campo Pais_Nome."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado.'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado.'})
        if not (arquivo.filename.endswith('.xlsx') or arquivo.filename.endswith('.xls')):
            return jsonify({'success': False, 'message': 'Formato inválido. Use .xlsx ou .xls.'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado.'})

        projeto = session['projeto_selecionado']
        banco_usuario = projeto.get('DadosGX')
        projeto_id = projeto.get('ProjetoID')
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto.'})

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        cursor = conexao.cursor()

        # 🟢 Lê a planilha
        df = pd.read_excel(arquivo, dtype=str)
        df = df.replace({pd.NaT: None, 'nan': None, 'NaN': None, '': None})
        df = df.where(pd.notnull(df), None)

        # Normaliza os nomes de colunas
        df.columns = [c.strip().lower().replace(' ', '').replace('_', '') for c in df.columns]
        colunas_norm = df.columns.tolist()
        logger.info(f"Colunas normalizadas: {colunas_norm}")

        # Função auxiliar para buscar coluna equivalente
        def encontrar_coluna(possiveis):
            for possivel in possiveis:
                possivel_norm = possivel.strip().lower().replace(' ', '').replace('_', '')
                for c in df.columns:
                    if possivel_norm in c or c in possivel_norm:
                        return c
            return None

        # Definir possíveis nomes para cada coluna
        col_pais_cd = encontrar_coluna(['codigoanterior', 'codanterior', 'paiscd', 'codigo', 'cgpais'])
        col_pais_codigo = encontrar_coluna(['paiscodigo', 'codigowf', 'codwf'])
        col_pais_descricao = encontrar_coluna(['paisdescricao', 'paisds', 'descricaopais', 'paisnome', 'nomepais'])

        # Verifica se pelo menos código e descrição foram encontrados
        if not col_pais_cd or not col_pais_codigo:
            return jsonify({
                'success': False,
                'message': f'Colunas obrigatórias não encontradas. '
                           f'A planilha deve conter colunas equivalentes a: Codigo Anterior e Pais_Codigo.'
            })

        registros = df.to_dict('records')
        contador_inseridos = 0
        contador_atualizados = 0

        for registro in registros:
            pais_cd = registro.get(col_pais_cd)
            pais_codigo = registro.get(col_pais_codigo)
            pais_descricao = registro.get(col_pais_descricao) if col_pais_descricao else None

            if not pais_cd:
                continue

            cursor.execute("SELECT COUNT(*) FROM Pais_DePara WHERE pais_cd = ?", (pais_cd,))
            resultado = cursor.fetchone()
            existe = resultado[0] if resultado else 0

            if existe:
                cursor.execute("""
                    UPDATE Pais_DePara
                    SET Pais_Codigo = ?, pais_ds = ?
                    WHERE pais_cd = ?
                """, (pais_codigo, pais_descricao, pais_cd))
                contador_atualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO Pais_DePara (pais_cd, Pais_Codigo, pais_ds)
                    VALUES (?, ?, ?)
                """, (pais_cd, pais_codigo, pais_descricao))
                contador_inseridos += 1

        conexao.commit()
        cursor.close()
        conexao.close()

        logger.info(f"Importação concluída — Atualizados: {contador_atualizados}, Inseridos: {contador_inseridos}")

        # 🟢 Atualiza automaticamente os nomes de país (Pais_Nome)
        atualizar_descricoes_apos_importacao(banco_usuario, projeto_id)

        mensagem = (f"Importação concluída com sucesso! "
                    f"{contador_atualizados} atualizados, {contador_inseridos} inseridos. "
                    f"Nomes atualizados automaticamente.")
        return jsonify({'success': True, 'message': mensagem})

    except Exception as e:
        logger.error(f"Erro ao importar países: {e}", exc_info=True)
        if 'conexao' in locals() and conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro na importação: {str(e)}'})


        
@pais_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline"""
    logger.info("=== UPDATE REGISTRO PAIS ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados recebidos: {data}")
        
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')
        
        if not record_id or not field:
            logger.error("ID ou campo não fornecidos")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            logger.error("Nenhum projeto selecionado na sessão")
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            logger.error("Banco não configurado para este projeto")
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        logger.info(f"Atualizando registro {record_id}, campo {field} para valor '{value}' no banco {banco_usuario}")
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            logger.error(f"Falha na conexão com o banco: {banco_usuario}")
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        
        # Verificar se a tabela existe e tem a coluna
        try:
            cursor.execute(f"SELECT TOP 1 {field} FROM Pais_DePara WHERE id = ?", (record_id,))
            resultado = cursor.fetchone()
            if not resultado:
                logger.error(f"Registro não encontrado: {record_id}")
                return jsonify({'success': False, 'message': 'Registro não encontrado'})
        except Exception as e:
            logger.error(f"Erro ao verificar registro: {str(e)}")
            return jsonify({'success': False, 'message': f'Campo {field} não existe na tabela'})
        
        # NOVA LÓGICA: Se estiver atualizando Pais_Codigo, buscar descrição automaticamente
        descricao_wf = None
        nome_coluna_descricao = 'Pais_Nome'  # Vamos tentar usar Pais_Nome
        
        if field == 'Pais_Codigo' and value and value != 'S/DePara' and value.strip() != '':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo:
                descricao_wf = obter_descricao_wf(banco_homo, value.strip())
                logger.info(f"Descrição WF encontrada para código {value}: {descricao_wf}")
        
        # Atualizar registro - usando id como chave primária
        if descricao_wf and nome_coluna_descricao:
            # Se encontrou descrição, atualizar ambos campos
            query = f"UPDATE Pais_DePara SET Pais_Codigo = ?, {nome_coluna_descricao} = ? WHERE id = ?"
            logger.info(f"Executando query: {query} com valores: ({value}, {descricao_wf}, {record_id})")
            cursor.execute(query, (value, descricao_wf, record_id))
        else:
            # Atualizar apenas o campo especificado
            query = f"UPDATE Pais_DePara SET {field} = ? WHERE id = ?"
            logger.info(f"Executando query: {query} com valores: ({value}, {record_id})")
            cursor.execute(query, (value, record_id))
        
        # Verificar se alguma linha foi afetada
        if cursor.rowcount == 0:
            logger.warning(f"Nenhuma linha afetada pela atualização do registro {record_id}")
            if conexao:
                conexao.rollback()
            return jsonify({'success': False, 'message': 'Registro não encontrado ou não modificado'})
        
        # Commit da transação
        conexao.commit()
        
        logger.info(f"Registro {record_id} atualizado com sucesso")
        
        # Retornar a descrição se foi atualizada
        response = {'success': True, 'message': 'Registro atualizado com sucesso'}
        if descricao_wf:
            response['descricao_wf'] = descricao_wf
            
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Erro ao atualizar registro: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Fazer rollback em caso de erro apenas se a conexão existir
        if conexao:
            try:
                conexao.rollback()
            except Exception as rollback_error:
                logger.error(f"Erro ao fazer rollback: {rollback_error}")
        
        return jsonify({'success': False, 'message': f'Erro ao atualizar registro: {str(e)}'})
    
    finally:
        # Fechar recursos de forma segura
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            logger.error(f"Erro ao fechar cursor: {e}")
        
        try:
            if conexao:
                conexao.close()
        except Exception as e:
            logger.error(f"Erro ao fechar conexão: {e}")

@pais_bp.route('/update_batch', methods=['POST'])
def update_batch():
    """Endpoint para atualizar múltiplos registros de uma vez"""
    logger.info("=== UPDATE BATCH ENDPOINT ACESSADO ===")
    
    conexao = None
    cursor = None
    try:
        # Obter dados da requisição
        data = request.get_json()
        logger.info(f"Dados batch recebidos: {len(data.get('updates', []))} atualizações")
        
        updates = data.get('updates', [])
        
        if not updates:
            return jsonify({'success': False, 'message': 'Nenhuma atualização fornecida'})
        
        # Obter o projeto da sessão
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})
        
        # Conectar ao banco
        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})
        
        cursor = conexao.cursor()
        success_count = 0
        error_count = 0
        error_messages = []
        
        # Processar cada atualização
        for update in updates:
            try:
                record_id = update.get('id')
                field = update.get('field')
                value = update.get('value')
                
                if not record_id or not field:
                    error_count += 1
                    error_messages.append(f"ID ou campo não fornecidos para atualização: {update}")
                    continue
                
                # NOVA LÓGICA: Se estiver atualizando Pais_Codigo, buscar descrição automaticamente
                descricao_wf = None
                nome_coluna_descricao = 'Pais_Nome'
                
                if field == 'Pais_Codigo' and value and value != 'S/DePara' and value.strip() != '':
                    banco_homo = obter_banco_homo(projeto_id)
                    if banco_homo:
                        descricao_wf = obter_descricao_wf(banco_homo, value.strip())
                
                # Atualizar registro
                if descricao_wf and nome_coluna_descricao:
                    query = f"UPDATE Pais_DePara SET Pais_Codigo = ?, {nome_coluna_descricao} = ? WHERE id = ?"
                    cursor.execute(query, (value, descricao_wf, record_id))
                else:
                    query = f"UPDATE Pais_DePara SET {field} = ? WHERE id = ?"
                    cursor.execute(query, (value, record_id))
                
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"Registro não encontrado: {record_id}")
                    
            except Exception as e:
                error_count += 1
                error_messages.append(f"Erro ao atualizar {record_id}: {str(e)}")
        
        # Commit de todas as atualizações
        conexao.commit()
        
        logger.info(f"Batch update concluído: {success_count} sucessos, {error_count} erros")
        
        response = {
            'success': True,
            'message': f'Atualizações concluídas: {success_count} sucessos, {error_count} erros',
            'success_count': success_count,
            'error_count': error_count
        }
        
        if error_messages:
            response['error_details'] = error_messages[:10]  # Limitar a 10 mensagens de erro
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Erro no batch update: {str(e)}")
        if conexao:
            conexao.rollback()
        return jsonify({'success': False, 'message': f'Erro no batch update: {str(e)}'})
    
    finally:
        if cursor:
            cursor.close()
        if conexao:
            conexao.close()

@pais_bp.route('/get_descricao_wf/<codigo>')
def get_descricao_wf(codigo):
    """Endpoint para obter a descrição de um código da base WF"""
    try:
        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})
        
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')
        
        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            return jsonify({'success': False, 'message': 'Banco homólogo não configurado'})
        
        descricao = obter_descricao_wf(banco_homo, codigo)
        
        if descricao:
            return jsonify({
                'success': True,
                'descricao': descricao
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Código não encontrado na base WF'
            })
            
    except Exception as e:
        logger.error(f"Erro ao buscar descrição WF: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao buscar descrição: {str(e)}'})