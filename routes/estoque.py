from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    jsonify,
    send_file,
)
from db.connection import conectar_segunda_base, conectar_banco
from logger import logger
from utils.excel_utils import export_to_excel, import_from_excel
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

estoque_bp = Blueprint("estoque", __name__)

def obter_banco_homo(projeto_id):
    """Obter BancoHomo (mesma lógica usada nos outros módulos)"""
    try:
        conn = conectar_banco()
        if not conn:
            logger.error("Falha ao conectar ao banco principal para obter BancoHomo")
            return None

        cursor = conn.cursor()
        cursor.execute("SELECT BancoHomo FROM Projeto WHERE ProjetoID = ?", projeto_id)
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()

        if resultado and resultado[0]:
            return resultado[0]
        return None
    except Exception as e:
        logger.error(f"Erro ao obter BancoHomo: {str(e)}")
        return None

def obter_codigos_wf(banco_homo):
    """Obtém todos os códigos da tabela estoque do banco homólogo"""
    try:
        if not banco_homo:
            return []
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return []
        cursor = conexao.cursor()
        cursor.execute("SELECT Estoque_Codigo FROM estoque")
        registros = cursor.fetchall()
        cursor.close()
        conexao.close()
        codigos = [str(r[0]) for r in registros if r and r[0] is not None]
        logger.info(f"Encontrados {len(codigos)} códigos na base WF (estoque)")
        return codigos
    except Exception as e:
        logger.error(f"Erro ao obter códigos WF (estoque): {str(e)}")
        return []

def obter_descricao_wf(banco_homo, codigo):
    """Obtém a descrição de um código específico da tabela estoque do banco homólogo"""
    try:
        if not banco_homo or not codigo:
            return None
        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            logger.error(f"Falha na conexão com o banco homólogo: {banco_homo}")
            return None
        cursor = conexao.cursor()
        cursor.execute("SELECT Estoque_Descricao FROM estoque WHERE Estoque_Codigo = ?", codigo)
        resultado = cursor.fetchone()
        cursor.close()
        conexao.close()
        if resultado and resultado[0]:
            return resultado[0]
        return None
    except Exception as e:
        logger.error(f"Erro ao obter descrição WF para código {codigo} (estoque): {str(e)}")
        return None

@estoque_bp.route("/")
def index():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_nome = projeto_selecionado.get('NomeProjeto', 'N/A')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return render_template('estoque.html', registros=[], colunas=[], projeto_nome=projeto_nome, banco_usuario=banco_usuario, codigos_wf=[])

        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM Estoque_DePara")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        registros_dict = [dict(zip(colunas, row)) for row in registros]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        cursor.close()
        conexao.close()

        return render_template('estoque.html',
                               registros=registros_dict,
                               colunas=colunas,
                               projeto_nome=projeto_nome,
                               banco_usuario=banco_usuario,
                               codigos_wf=codigos_wf,
                               banco_homo=banco_homo)
    except Exception as e:
        logger.error(f"Erro em estoque.index: {str(e)}")
        flash(f'Erro: {str(e)}', 'error')
        return render_template('estoque.html', registros=[], colunas=[], projeto_nome='N/A', banco_usuario='N/A', codigos_wf=[])

@estoque_bp.route('/exportar')
def exportar_estoque():
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            flash('Banco não configurado para este projeto.', 'error')
            return redirect(url_for('estoque.index'))

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            flash(f'Falha na conexão com o banco: {banco_usuario}', 'error')
            return redirect(url_for('estoque.index'))

        cursor = conexao.cursor()
        cursor.execute("""
            SELECT est_cd, est_ds, Migra_Estoque, Estoque_Codigo,
                   Estoque_Descricao, Origem, Estoque_Sigla
            FROM Estoque_DePara
        """)
        registros = cursor.fetchall()
        colunas_originais = [column[0] for column in cursor.description]

        mapeamento_colunas = {
            'est_cd': 'Codigo de Origem',
            'est_ds': 'Descrição de origem',
            'Migra_Estoque': 'Migra_Estoque',
            'Estoque_Codigo': 'Estoque_Codigo',
            'Estoque_Descricao': 'Estoque_Descricao',
            'Origem': 'Origem',
            'Estoque_Sigla': 'Estoque_Sigla'
        }
        colunas_amigaveis = [mapeamento_colunas.get(col, col) for col in colunas_originais]

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)
        ws = wb.create_sheet(title="Estoque_DePara")

        for col_num, coluna in enumerate(colunas_amigaveis, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                if valor and isinstance(valor, str):
                    valor = valor.strip()
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                # Estoque_Codigo está na posição 4 na query acima
                if col_num == 4:
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        cursor.close()
        conexao.close()

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Estoque_DePara.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar estoque: {str(e)}")
        flash(f'Erro na exportação: {str(e)}', 'error')
        return redirect(url_for('estoque.index'))

@estoque_bp.route('/exportar_filtrados', methods=['POST'])
def exportar_estoque_filtrados():
    try:
        data = request.get_json()
        registros_filtrados = data.get('registros', [])
        headers = data.get('headers', [])
        if not registros_filtrados:
            return jsonify({'success': False, 'message': 'Nenhum registro para exportar'}), 400

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        codigos_wf = obter_codigos_wf(banco_homo) if banco_homo else []

        wb = Workbook()
        if wb.sheetnames and 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)
        ws = wb.create_sheet(title="Estoque_Filtrado")

        for col_num, coluna in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for row_num, registro in enumerate(registros_filtrados, 2):
            for col_num, header in enumerate(headers, 1):
                valor = registro.get(header, '')
                if isinstance(valor, str):
                    valor = valor.strip()
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                if header == 'Estoque_Codigo':
                    if not valor or valor == '':
                        cell.fill = laranja
                    elif valor == 'S/DePara':
                        cell.fill = amarelo
                    elif valor and str(valor) in codigos_wf:
                        cell.fill = verde
                    elif valor and str(valor) not in codigos_wf:
                        cell.fill = vermelho

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Estoque_Filtrado.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar estoque filtrado: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro na exportação: {str(e)}'}), 500

@estoque_bp.route('/export_wf')
def export_wf():
    """Exporta a tabela estoque da base homóloga (WF)"""
    try:
        if 'projeto_selecionado' not in session:
            flash('Nenhum projeto selecionado.', 'error')
            return redirect(url_for('auth.selecionar_projeto'))

        projeto_selecionado = session['projeto_selecionado']
        projeto_id = projeto_selecionado.get('ProjetoID')

        banco_homo = obter_banco_homo(projeto_id)
        if not banco_homo:
            flash('Banco homólogo não configurado para este projeto.', 'error')
            return redirect(url_for('estoque.index'))

        conexao = conectar_segunda_base(banco_homo)
        if not conexao:
            flash(f'Falha na conexão com o banco homólogo: {banco_homo}', 'error')
            return redirect(url_for('estoque.index'))

        cursor = conexao.cursor()
        # SELECT solicitado: Estoque_Codigo, Estoque_Descricao, Estoque_Tipo, Estoque_Ativo
        cursor.execute("select Estoque_Codigo, Estoque_Descricao, Estoque_Tipo, Estoque_Ativo from estoque")
        registros = cursor.fetchall()
        colunas = [column[0] for column in cursor.description]
        registros_list = [dict(zip(colunas, row)) for row in registros]
        cursor.close()
        conexao.close()

        df = pd.DataFrame(registros_list, columns=colunas)
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Estoque_WF', index=False)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="Estoque_WF.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Erro ao exportar tabela WF (estoque): {str(e)}")
        flash(f'Erro na exportação da tabela WF: {str(e)}', 'error')
        return redirect(url_for('estoque.index'))

@estoque_bp.route('/importar', methods=['POST'])
def importar_estoque():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo = request.files['file']
        if not arquivo or not arquivo.filename:
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        filename = arquivo.filename
        if not (filename and (filename.endswith('.xlsx') or filename.endswith('.xls'))):
            return jsonify({'success': False, 'message': 'Formato de arquivo inválido. Use .xlsx ou .xls'})

        if 'projeto_selecionado' not in session:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'})

        projeto_selecionado = session['projeto_selecionado']
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        if not banco_usuario:
            return jsonify({'success': False, 'message': 'Banco não configurado para este projeto'})

        logger.info(f"Iniciando importação de estoque para o banco: {banco_usuario}")

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'})

        cursor = conexao.cursor()

        colunas_importacao = [
            'est_cd',
            'est_ds',
            'Migra_Estoque',
            'Estoque_Codigo',
            'Estoque_Descricao',
            'Origem',
            'Estoque_Sigla'
        ]

        try:
            df = pd.read_excel(arquivo)
            df = df.where(pd.notnull(df), None)
            registros = df.to_dict('records')
            colunas_excel = df.columns.tolist()
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel (estoque): {str(e)}")
            return jsonify({'success': False, 'message': f'Erro ao ler arquivo Excel: {str(e)}'})

        logger.info(f"Arquivo Excel lido: {len(registros)} registros, colunas: {colunas_excel}")

        # Mapeamento amigável -> técnico
        mapeamento_colunas_excel = {
            'Codigo de Origem': 'est_cd',
            'Descrição de origem': 'est_ds',
            'Migra_Estoque': 'Migra_Estoque',
            'Estoque_Codigo': 'Estoque_Codigo',
            'Estoque_Descricao': 'Estoque_Descricao',
            'Origem': 'Origem',
            'Estoque_Sigla': 'Estoque_Sigla'
        }

        colunas_excel_normalizadas = []
        for col in colunas_excel:
            if col in mapeamento_colunas_excel:
                colunas_excel_normalizadas.append(mapeamento_colunas_excel[col])
            else:
                colunas_excel_normalizadas.append(col)

        df.columns = colunas_excel_normalizadas

        colunas_excel_set = set(colunas_excel_normalizadas)
        colunas_necessarias = set(colunas_importacao)
        if not colunas_necessarias.issubset(colunas_excel_set):
            missing = colunas_necessarias - colunas_excel_set
            return jsonify({'success': False, 'message': f'Colunas necessárias faltando no arquivo: {", ".join(missing)}'})

        registros = df.to_dict('records')

        registros_filtrados = []
        for registro in registros:
            registro_filtrado = {}
            for col in colunas_importacao:
                valor = registro.get(col)
                if valor is None:
                    registro_filtrado[col] = None
                else:
                    str_valor = str(valor)
                    if col in ['est_cd', 'est_ds', 'Estoque_Descricao']:
                        registro_filtrado[col] = str_valor[:150]
                    elif col == 'Migra_Estoque':
                        registro_filtrado[col] = str_valor[:3]
                    elif col in ['Estoque_Codigo', 'Estoque_Sigla']:
                        registro_filtrado[col] = str_valor[:100]
                    elif col == 'Origem':
                        registro_filtrado[col] = str_valor[:200]
                    else:
                        registro_filtrado[col] = str_valor
            registros_filtrados.append(registro_filtrado)

        logger.info(f"Registros filtrados e tratados: {len(registros_filtrados)}")

        cursor.execute("SELECT COUNT(*) FROM Estoque_DePara")
        result_antes = cursor.fetchone()
        count_antes = result_antes[0] if result_antes else 0
        logger.info(f"Registros antes da importação: {count_antes}")

        contador_atualizacoes = 0
        contador_insercoes = 0

        for registro in registros_filtrados:
            codigo_origem = registro.get('est_cd')
            novo_codigo_wf = registro.get('Estoque_Codigo')
            nova_descricao = registro.get('Estoque_Descricao')

            if codigo_origem:
                cursor.execute("SELECT id FROM Estoque_DePara WHERE est_cd = ?", (codigo_origem,))
                resultado = cursor.fetchone()
                if resultado:
                    cursor.execute("""
                        UPDATE Estoque_DePara
                        SET est_ds = ?, Migra_Estoque = ?, Estoque_Codigo = ?, Estoque_Descricao = ?, Origem = ?, Estoque_Sigla = ?
                        WHERE est_cd = ?
                    """, (
                        registro.get('est_ds'),
                        registro.get('Migra_Estoque'),
                        novo_codigo_wf,
                        nova_descricao,
                        registro.get('Origem'),
                        registro.get('Estoque_Sigla'),
                        codigo_origem
                    ))
                    contador_atualizacoes += 1
                else:
                    colunas_str = ', '.join(colunas_importacao)
                    placeholders = ', '.join(['?' for _ in colunas_importacao])
                    valores = [registro.get(c) for c in colunas_importacao]
                    cursor.execute(f"INSERT INTO Estoque_DePara ({colunas_str}) VALUES ({placeholders})", valores)
                    contador_insercoes += 1
            else:
                colunas_str = ', '.join(colunas_importacao)
                placeholders = ', '.join(['?' for _ in colunas_importacao])
                valores = [registro.get(c) for c in colunas_importacao]
                cursor.execute(f"INSERT INTO Estoque_DePara ({colunas_str}) VALUES ({placeholders})", valores)
                contador_insercoes += 1

        conexao.commit()
        logger.info(f"UPDATEs: {contador_atualizacoes}, INSERTs: {contador_insercoes}")

        banco_homo = obter_banco_homo(projeto_id)
        if banco_homo:
            logger.info("Atualizando descrições de estoque com base no banco homólogo...")
            # percorre e atualiza descrições com base no banco homólogo
            conexao_aux = conectar_segunda_base(banco_usuario)
            if conexao_aux:
                cursor_aux = conexao_aux.cursor()
                cursor_aux.execute("""
                    SELECT id, Estoque_Codigo, Estoque_Descricao 
                    FROM Estoque_DePara 
                    WHERE Estoque_Codigo IS NOT NULL AND Estoque_Codigo != 'S/DePara'
                """)
                rows = cursor_aux.fetchall()
                for row in rows:
                    id_reg, codigo_wf, desc_atual = row
                    desc_wf = obter_descricao_wf(banco_homo, codigo_wf)
                    if desc_wf and desc_wf != desc_atual:
                        cursor_aux.execute("UPDATE Estoque_DePara SET Estoque_Descricao = ? WHERE id = ?", (desc_wf, id_reg))
                conexao_aux.commit()
                cursor_aux.close()
                conexao_aux.close()

        cursor.execute("SELECT COUNT(*) FROM Estoque_DePara")
        result_depois = cursor.fetchone()
        count_depois = result_depois[0] if result_depois else 0
        logger.info(f"Registros depois da importação: {count_depois}")

        # Exemplo para debug (SQL Server TOP)
        try:
            cursor.execute("SELECT TOP 3 * FROM Estoque_DePara")
            exemplos = cursor.fetchall()
            logger.info(f"Exemplos após import: {exemplos}")
        except Exception:
            # alguns DB engines não suportam TOP - ignorar erro de debug
            pass

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': f'Importação concluída! {contador_atualizacoes} atualizados, {contador_insercoes} inseridos. Total: {count_depois}.'})
    except Exception as e:
        logger.error(f"Erro ao importar estoque: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # rollback / fechar conexões com segurança se existirem
        if 'conexao' in locals() and conexao is not None:
            try:
                conexao.rollback()
            except Exception:
                pass
            try:
                conexao.close()
            except Exception:
                pass
        return jsonify({'success': False, 'message': f'Erro na importação: {str(e)}'})

@estoque_bp.route('/update', methods=['POST'])
def update_registro():
    """Endpoint para atualizar um registro individual via edição inline (aceita Estoque_Codigo e Estoque_Descricao)"""
    logger.info("=== UPDATE REGISTRO ESTOQUE ENDPOINT ACESSADO ===")
    conexao = None
    cursor = None
    try:
        data = request.get_json()
        logger.info(f"Dados recebidos (update estoque): {data}")

        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not record_id or not field:
            logger.error("ID ou campo não fornecidos (update estoque)")
            return jsonify({'success': False, 'message': 'ID e campo são obrigatórios'}), 400

        colunas_permitidas = ['Estoque_Codigo', 'Estoque_Descricao']
        if field not in colunas_permitidas:
            logger.error(f"Tentativa de editar campo não permitido: {field}")
            return jsonify({'success': False, 'message': 'Campo não permitido para edição'}), 400

        projeto_selecionado = session.get('projeto_selecionado')
        if not projeto_selecionado:
            return jsonify({'success': False, 'message': 'Nenhum projeto selecionado'}), 400
        banco_usuario = projeto_selecionado.get('DadosGX')
        projeto_id = projeto_selecionado.get('ProjetoID')

        conexao = conectar_segunda_base(banco_usuario)
        if not conexao:
            return jsonify({'success': False, 'message': f'Falha na conexão com o banco: {banco_usuario}'}), 500

        cursor = conexao.cursor()

        if field == 'Estoque_Codigo':
            if value != 'S/DePara' and value is not None and not str(value).isdigit():
                return jsonify({'success': False, 'message': 'Estoque_Codigo deve ser somente números ou "S/DePara"'}), 400

        cursor.execute(f"UPDATE Estoque_DePara SET {field} = ? WHERE id = ?", (value, record_id))
        conexao.commit()
        logger.info(f"Registro id={record_id} atualizado: {field} = {value}")

        if field == 'Estoque_Codigo':
            banco_homo = obter_banco_homo(projeto_id)
            if banco_homo and value and value != 'S/DePara':
                descricao_wf = obter_descricao_wf(banco_homo, value)
                if descricao_wf:
                    cursor.execute("UPDATE Estoque_DePara SET Estoque_Descricao = ? WHERE id = ?", (descricao_wf, record_id))
                    conexao.commit()
                    logger.info(f"Descrição atualizada automaticamente para id={record_id} com valor '{descricao_wf}'")

        cursor.close()
        conexao.close()
        return jsonify({'success': True, 'message': 'Registro atualizado com sucesso'})
    except Exception as e:
        logger.error(f"Erro no endpoint update (estoque): {str(e)}")
        # rollback e fechar com segurança
        if conexao is not None:
            try:
                conexao.rollback()
            except Exception as ex:
                logger.warning(f"Falha ao executar rollback: {str(ex)}")
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conexao is not None:
            try:
                conexao.close()
            except Exception:
                pass
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {str(e)}'}), 500
